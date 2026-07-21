"""Add Tech Stack tab to Education Partnerships Excel."""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

path = r"C:\Users\USER\Desktop\LMT\ops\LMT-EDUCATION-PARTNERSHIPS.xlsx"
wb = load_workbook(path)

navy = PatternFill(start_color="0E1C2F", end_color="0E1C2F", fill_type="solid")
red_bg = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
green_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
yellow_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
red_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
light_blue = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
white_font = Font(color="FFFFFF", bold=True, size=11)
normal_font = Font(size=11)
bold_font = Font(bold=True, size=11)
big_bold = Font(bold=True, size=13)
wrap = Alignment(wrap_text=True, vertical="top")
border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def add_header(ws, row, headers, fill=navy):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def add_row(ws, row, data, fill=None):
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = normal_font
        cell.alignment = wrap
        cell.border = border
        if fill:
            cell.fill = fill


# Remove existing Tech Stack tab if it exists
if "Tech Stack + Burn Rate" in wb.sheetnames:
    del wb["Tech Stack + Burn Rate"]

ws = wb.create_sheet("Tech Stack + Burn Rate")
ws.sheet_properties.tabColor = "C0392B"

h = ["Category", "Tool", "Monthly", "Annual", "Billing", "Status", "Action", "Potential Savings"]
add_header(ws, 1, h, red_bg)

rows = [
    # AI / Development
    ["AI / Dev", "Claude Max (Code + Chat)", "$100.00", "$1,200", "Monthly", "ACTIVE", "Keep — primary tool", ""],
    ["AI / Dev", "Claude API Console", "VARIABLE", "??", "Per-use", "DUPLICATE", "CANCEL NOW — double billing", "Unknown — check bank"],
    ["AI / Dev", "Claude Pro ($20)", "$20.00", "$240", "Monthly", "CHECK", "May be redundant with Max", "-$20/mo if redundant"],
    ["AI / Dev", "Manus Pro", "$20.00", "$240", "Monthly", "CHECK", "Still using? Cancel if not", "-$20/mo if unused"],
    ["AI / Dev", "Grammarly Premium", "$12.00", "$144", "Monthly", "ACTIVE", "Keep", ""],
    ["AI / Dev", "Coursera", "$0-59", "$0-708", "Monthly", "CHECK", "Free or paid plan?", "-$59/mo if on paid"],

    # Video / Audio / Design
    ["Video/Audio", "HeyGen", "$24.00", "$288", "Monthly", "ACTIVE", "No edu discount available", ""],
    ["Video/Audio", "ElevenLabs Creator", "$18.33", "$220", "Annual", "ACTIVE", "Apply Impact Program NOW", "-$18/mo if approved (free)"],
    ["Video/Audio", "Canva Pro", "$13.00", "$156", "Monthly", "ACTIVE", "Free with 501c3 sponsor", "-$13/mo with sponsor"],
    ["Video/Audio", "Adobe Creative Cloud", "$20.83", "$250", "Annual", "ACTIVE", "$20K grant if sponsor found", "-$21/mo + $20K grant"],
    ["Video/Audio", "Snagit/Camtasia", "$16.58", "$199", "Annual", "ACTIVE", "Keep", ""],

    # Communication / Email
    ["Comms", "Google Workspace Enterprise", "$29.00", "$348", "Monthly", "ACTIVE", "Free with MS Nonprofit if sponsor", "-$29/mo with sponsor"],
    ["Comms", "Mailchimp Starter", "$13.00", "$156", "Monthly", "ACTIVE", "15% discount with 501c3", "-$2/mo with sponsor"],
    ["Comms", "Zoom", "$0", "$0", "Free plan", "ACTIVE", "50% off paid with 501c3", ""],

    # Website / Platform
    ["Platform", "Bluehost (WP Hosting)", "~$13.00", "~$156", "Monthly", "ACTIVE", "Keep", ""],
    ["Platform", "LearnDash", "$18.67", "$224", "Annual Nov 2", "ACTIVE", "Keep — course platform", ""],
    ["Platform", "BuddyBoss", "$16.67", "$200", "Annual Nov 2", "ACTIVE", "Keep — community platform", ""],

    # Marketing / Social
    ["Marketing", "LinkedIn Premium", "~$60.00", "~$720", "Monthly", "ACTIVE — exp ~May 14", "Renew decision based on pipeline", ""],
    ["Marketing", "Metricool", "$0", "$0", "Free", "ACTIVE", "Analytics — connected", ""],
    ["Marketing", "Buffer", "$0", "$0", "Not connected", "NOT ACTIVE", "Connect when posting 10+/wk", ""],
    ["Marketing", "Calendly", "$0", "$0", "Free plan", "ACTIVE", "Keep", ""],

    # Podcast
    ["Podcast", "Buzzsprout", "$0", "$0", "Free trial", "TRIAL — exp Jul 11", "Upgrade to $19/mo or lose episodes", "+$19/mo after Jul 11"],

    # Domain / Email
    ["Domain", "learnmoretechnologies.com", "~$2.00", "~$24", "Annual", "ACTIVE", "Keep", ""],
    ["Domain", "learnmo.com", "~$2.00", "~$24", "Annual", "ACTIVE", "Keep", ""],
]

for i, r in enumerate(rows):
    if "DUPLICATE" in r[5] or "CANCEL" in r[6]:
        fill = red_fill
    elif "CHECK" in r[5] or "TRIAL" in r[5]:
        fill = yellow_fill
    elif "NOT ACTIVE" in r[5]:
        fill = light_blue
    else:
        fill = None
    add_row(ws, i + 2, r, fill)

# Summary rows
summary_row = len(rows) + 3
ws.cell(row=summary_row, column=1, value="MONTHLY TOTALS").font = big_bold
ws.cell(row=summary_row + 1, column=1, value="Current monthly burn (estimated):").font = bold_font
ws.cell(row=summary_row + 1, column=3, value="~$395-$435").font = Font(bold=True, size=12, color="C0392B")

ws.cell(row=summary_row + 2, column=1, value="Current annual burn (estimated):").font = bold_font
ws.cell(row=summary_row + 2, column=3, value="~$4,700-$5,200").font = Font(bold=True, size=12, color="C0392B")

ws.cell(row=summary_row + 4, column=1, value="IMMEDIATE SAVINGS (cancel/check):").font = big_bold
ws.cell(row=summary_row + 5, column=1, value="Cancel Claude API duplicate").font = normal_font
ws.cell(row=summary_row + 5, column=3, value="Unknown").font = normal_font
ws.cell(row=summary_row + 6, column=1, value="Cancel Claude Pro if redundant with Max").font = normal_font
ws.cell(row=summary_row + 6, column=3, value="-$20/mo").font = normal_font
ws.cell(row=summary_row + 7, column=1, value="Cancel Manus if unused").font = normal_font
ws.cell(row=summary_row + 7, column=3, value="-$20/mo").font = normal_font
ws.cell(row=summary_row + 8, column=1, value="ElevenLabs Impact (if approved)").font = normal_font
ws.cell(row=summary_row + 8, column=3, value="-$18/mo").font = normal_font
ws.cell(row=summary_row + 9, column=1, value="Immediate savings potential:").font = bold_font
ws.cell(row=summary_row + 9, column=3, value="-$58-$78/mo").font = Font(bold=True, size=12, color="109F35")

ws.cell(row=summary_row + 11, column=1, value="WITH FISCAL SPONSOR (501c3):").font = big_bold
ws.cell(row=summary_row + 12, column=1, value="Anthropic 75% off").font = normal_font
ws.cell(row=summary_row + 12, column=3, value="-$75/mo").font = normal_font
ws.cell(row=summary_row + 13, column=1, value="Canva free").font = normal_font
ws.cell(row=summary_row + 13, column=3, value="-$13/mo").font = normal_font
ws.cell(row=summary_row + 14, column=1, value="Microsoft free (replace Google WS)").font = normal_font
ws.cell(row=summary_row + 14, column=3, value="-$29/mo").font = normal_font
ws.cell(row=summary_row + 15, column=1, value="Adobe discount + $20K grant").font = normal_font
ws.cell(row=summary_row + 15, column=3, value="-$21/mo + $20K").font = normal_font
ws.cell(row=summary_row + 16, column=1, value="Total with sponsor:").font = bold_font
ws.cell(row=summary_row + 16, column=3, value="-$138/mo + grants").font = Font(bold=True, size=12, color="109F35")

ws.cell(row=summary_row + 18, column=1, value="BEST CASE MONTHLY BURN:").font = big_bold
ws.cell(row=summary_row + 18, column=3, value="~$200/mo ($2,400/yr)").font = Font(bold=True, size=14, color="109F35")

ws.cell(row=summary_row + 19, column=1, value="vs current:").font = bold_font
ws.cell(row=summary_row + 19, column=3, value="~$420/mo ($5,000/yr)").font = Font(bold=True, size=14, color="C0392B")

ws.cell(row=summary_row + 20, column=1, value="ANNUAL SAVINGS:").font = big_bold
ws.cell(row=summary_row + 20, column=3, value="$2,600/yr + grants").font = Font(bold=True, size=14, color="109F35")

# Set widths
for i, w in enumerate([14, 30, 12, 12, 16, 20, 35, 22]):
    ws.column_dimensions[chr(65 + i)].width = w

wb.save(path)
print(f"Updated: {path}")
