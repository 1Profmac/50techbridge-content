"""Build LMT Workforce Outreach Excel tracker."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

navy = PatternFill(start_color="0E1C2F", end_color="0E1C2F", fill_type="solid")
gold = PatternFill(start_color="C8942E", end_color="C8942E", fill_type="solid")
orange = PatternFill(start_color="E8733A", end_color="E8733A", fill_type="solid")
green_bg = PatternFill(start_color="109F35", end_color="109F35", fill_type="solid")
green_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
yellow_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
white_font = Font(color="FFFFFF", bold=True, size=11)
bold_font = Font(bold=True, size=11)
normal_font = Font(size=11)
wrap = Alignment(wrap_text=True, vertical="top")
border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def add_header(ws, row, headers, fill=navy):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def add_row(ws, row, data, fill=None):
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = normal_font
        cell.alignment = wrap
        cell.border = border
        if fill:
            cell.fill = fill


def set_widths(ws, widths):
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65 + i)].width = w


# === SHEET 1: BATCH 1 ===
ws1 = wb.active
ws1.title = "Batch 1 - WIOA"
ws1.sheet_properties.tabColor = "0E1C2F"
h1 = ["Target", "Email", "Phone", "Channel", "Sent", "Reply", "Next Action"]
add_header(ws1, 1, h1)
rows1 = [
    ["David Goldstein (C2GPS)", "david@c2gps.net", "(512) 577-3030", "Phone Apr 20 + Email Apr 22", "Apr 22", "", "Follow up Apr 25"],
    ["Jason Helmick (C2GPS)", "jason@c2gps.net", "(512) 577-3030", "Email Apr 22", "Apr 22", "", "Follow up Apr 25"],
    ["Yael Lawson (WSCA CEO)", "yael.lawson@wfscapitalarea.com", "", "Email Apr 22", "Apr 22", "", "Follow up Apr 25"],
    ["WS Alamo", "etp@wsalamo.org", "(210) 224-4357", "Email Apr 22", "Apr 22", "", "Follow up Apr 25"],
    ["Juliet Stipeche (Gulf Coast)", "juliet.stipeche@wrksolutions.net", "", "Email Apr 22", "Apr 22", "", "Follow up Apr 25"],
]
for i, r in enumerate(rows1):
    add_row(ws1, i + 2, r, green_fill)
set_widths(ws1, [30, 35, 18, 28, 10, 10, 22])

# === SHEET 2: BATCH 2 ===
ws2 = wb.create_sheet("Batch 2 - Partners")
ws2.sheet_properties.tabColor = "C8942E"
add_header(ws2, 1, h1, gold)
rows2 = [
    ["Jasmin Vargas (Austin Free-Net)", "TBD (bounced)", "(512) 236-8225", "Phone Apr 22 (VM)", "Called", "", "Get email, follow up Apr 28"],
    ["Goodwill Central TX", "digitalaccess@gwctx.org", "", "Email", "", "", "Send this week"],
    ["Darryl Greer (AARP/Senior Planet)", "Dgreer@aarp.org", "(210) 920-9323", "Phone Apr 23 + Email Apr 22", "Both", "", "Follow up Apr 28"],
    ["Connected Dallas", "connecteddallas.org", "", "Email", "", "", "Find program director"],
    ["Harris County Broadband", "broadband@harriscountytx.gov", "(713) 755-5000", "Email", "", "", "Send this week"],
]
for i, r in enumerate(rows2):
    fill = green_fill if r[4] else yellow_fill
    add_row(ws2, i + 2, r, fill)
set_widths(ws2, [32, 30, 18, 28, 10, 10, 28])

# === SHEET 3: BATCH 3 ===
ws3 = wb.create_sheet("Batch 3 - Employers")
ws3.sheet_properties.tabColor = "E8733A"
h3 = ["#", "Company", "Contact", "Phone", "Call Day", "Status", "Script / Hook"]
add_header(ws3, 1, h3, orange)
rows3 = [
    ["1", "Texas Mutual Insurance", "Larry Martin, SVP HR", "(512) 224-4700", "Wed Apr 23", "NOT CALLED", "Insurance = oldest workforce. AI tools with zero training."],
    ["2", "National Western Life", "Kim Gibbs, VP HR", "(512) 836-1010", "Wed Apr 23", "NOT CALLED", "Small company = fast decision. Tenured workforce."],
    ["3", "City of Austin HR", "Rebecca Kennedy, Deputy Dir", "(512) 974-3400", "Thu Apr 24", "NOT CALLED", "14K employees. MBE procurement. Diversity spend."],
    ["4", "TxDOT", "Christopher Young, HR Dir", "(512) 463-8588", "Thu Apr 24", "NOT CALLED", "13K employees. Chairs Training subcommittee."],
    ["5", "LCRA", "Stephanie Taylor, HR Mgr", "(512) 578-4004", "Fri Apr 25", "NOT CALLED", "Utility. Half workforce retiring in 5 years."],
    ["6", "Travis County", "HR Director", "(512) 854-9165", "Fri Apr 25", "NOT CALLED", "MBE procurement goals. Digitizing everything."],
    ["7", "St. David's / HCA", "Todd Steward", "(512) 544-5000", "Mon Apr 28", "NOT CALLED", "Healthcare. Nurses 50+ need AI. Supplier diversity."],
    ["8", "Austin ISD", "Pamela Hall, Exec Dir HR", "(512) 414-1700", "Mon Apr 28", "NOT CALLED", "11K employees. Veteran teachers need AI literacy."],
    ["9", "Baylor Scott & White", "Regional VP HR", "(512) 509-0100", "Tue Apr 29", "NOT CALLED", "Expanding in Austin. Supplier diversity."],
    ["10", "Ascension Seton", "Jay Huckabee, HR Exec", "LinkedIn", "Tue Apr 29", "NOT CALLED", "Largest private employer Austin. Supplier diversity."],
]
for i, r in enumerate(rows3):
    add_row(ws3, i + 2, r)
set_widths(ws3, [5, 24, 28, 18, 14, 14, 50])

# === SHEET 4: STRATEGIC ===
ws4 = wb.create_sheet("Batch 4 - Strategic")
ws4.sheet_properties.tabColor = "109F35"
h4 = ["Target", "Org", "Phone / Email", "Channel", "Status", "Next Action"]
add_header(ws4, 1, h4, green_bg)
rows4 = [
    ["Jessica Lemann", "AARP Texas", "(512) 480-2498", "LinkedIn DM + Office VM Apr 23", "Waiting", "Follow up Fri"],
    ["Larry Williams", "Gov contracting / Austin ISD", "LinkedIn", "LinkedIn DM Apr 23", "CONFIRMED Apr 29", "Fair Housing Conference"],
    ["Dr. Angela Mulrooney", "Unleashing Influence", "info@unleashinginfluence.com", "Calendly", "CONFIRMED Apr 30 10AM", "Prep talking points"],
    ["Susannah Munro", "Gov Contracting (SA)", "LinkedIn", "Connection req Apr 23", "Pending accept", "Coffee when accepted"],
    ["Transamerica Institute", "Financial services", "LinkedIn", "Comment exchange Apr 23", "Engaged", "Find research director"],
    ["Morgan DuBose", "Texas Mutual (grants)", "communityaffairs@texasmutual.com", "Webinar + email Apr 22", "Waiting", "Monitor Aug RFP"],
    ["Felicia Brown", "AARP", "LinkedIn", "DM Apr 15", "Responded +", "Follow up"],
    ["Carly Roszkowski", "AARP Foundation HQ", "LinkedIn", "Connection req Apr 15", "Pending", "Wait for accept"],
    ["Tosan Arueyingho", "Black Is Tech", "Email", "Pitch Apr 13", "No reply", "Follow up"],
]
for i, r in enumerate(rows4):
    if "CONFIRMED" in r[4]:
        fill = green_fill
    elif "Waiting" in r[4] or "Pending" in r[4]:
        fill = yellow_fill
    else:
        fill = None
    add_row(ws4, i + 2, r, fill)
set_widths(ws4, [22, 24, 30, 28, 22, 26])

# === SHEET 5: CALL SCHEDULE ===
ws5 = wb.create_sheet("Call Schedule")
ws5.sheet_properties.tabColor = "0E1C2F"
h5 = ["Day", "Call 1", "Phone", "Call 2", "Phone", "Events / Follow-ups"]
add_header(ws5, 1, h5)
rows5 = [
    ["Wed Apr 23 (TODAY)", "Texas Mutual Insurance", "(512) 224-4700", "National Western Life", "(512) 836-1010", ""],
    ["Thu Apr 24", "City of Austin", "(512) 974-3400", "TxDOT", "(512) 463-8588", ""],
    ["Fri Apr 25", "LCRA", "(512) 578-4004", "Travis County", "(512) 854-9165", "Follow up Batch 1 emails"],
    ["Mon Apr 28", "St. David's / HCA", "(512) 544-5000", "Austin ISD", "(512) 414-1700", "Follow up Jasmin, Darryl"],
    ["Tue Apr 29", "Baylor Scott & White", "(512) 509-0100", "Follow-ups", "", "Fair Housing Conference (Larry)"],
    ["Wed Apr 30", "Follow-up calls", "", "", "", "Angela 10AM, Craig Hewitt 11AM"],
]
for i, r in enumerate(rows5):
    fill = gold if i == 0 else None
    add_row(ws5, i + 2, r, fill)
set_widths(ws5, [22, 24, 18, 24, 18, 32])

# === SAVE ===
path = r"C:\Users\USER\Desktop\LMT\ops\LMT-WORKFORCE-OUTREACH.xlsx"
wb.save(path)
print(f"Saved: {path}")
