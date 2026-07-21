"""Build LMT Education Partnerships Excel tracker."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

navy = PatternFill(start_color="0E1C2F", end_color="0E1C2F", fill_type="solid")
green_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
yellow_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
red_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
white_font = Font(color="FFFFFF", bold=True, size=11)
normal_font = Font(size=11)
wrap = Alignment(wrap_text=True, vertical="top")
border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def add_header(ws, row, headers, fill=navy):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def add_row(ws, row, data, fill=None):
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = normal_font
        cell.alignment = wrap
        cell.border = border
        if fill:
            cell.fill = fill


# === SHEET 1: RATED PARTNERSHIPS ===
ws1 = wb.active
ws1.title = "Education Partnerships"

h1 = ["Rank", "Company", "Program", "What You Get", "Value/Year", "501(c)(3)?", "LMT Eligible?", "Contact / URL", "Rating", "Notes"]
add_header(ws1, 1, h1)

rows = [
    ["1", "Cisco", "Networking Academy", "Free IT/AI curriculum + certs for Pioneers", "Free", "NO", "YES - Apply now", "netacad.com", "10/10", "No 501c3 needed. Free curriculum. Apply immediately."],
    ["2", "ElevenLabs", "Impact Program", "Free Pro account for education orgs", "$1,188/yr", "Flexible", "MAYBE - Apply as educator", "elevenlabs.io/impact-program", "9/10", "You use their voice clone. Apply citing community mission."],
    ["3", "Google", "AI Opportunity Fund", "$30M AI literacy grants + Workspace", "$10K-$50K", "YES (grants)", "PARTIAL - Need sponsor", "aiopportunityfund.withgoogle.com", "9/10", "Massive pool. Need 501c3 partner. Mission is exact match."],
    ["4", "Anthropic", "Claude for Nonprofits", "70-75% off Claude tools", "$2,400+/yr", "YES", "NO - Need sponsor", "support.claude.com", "8/10", "You pay full price now. 75% off through nonprofit partner."],
    ["5", "OpenAI", "OpenAI for Nonprofits", "75% off ChatGPT Business", "$1,800+/yr", "YES", "NO - Need sponsor", "openai.com/nonprofits", "8/10", "$50M commitment. Apply through Goodstack."],
    ["6", "Microsoft", "MS for Nonprofits", "75% off M365 + $2K Azure credits", "$10,000+/yr", "YES", "NO - Need sponsor", "microsoft.com/nonprofits", "8/10", "Biggest dollar value across the board."],
    ["7", "Canva", "Canva for Nonprofits", "Free Pro for up to 50 users", "$9,000/yr", "YES", "NO - Need sponsor", "canva.com/canva-for-nonprofits", "7/10", "You already use Canva. Free Pro saves real money."],
    ["8", "Adobe", "Adobe Nonprofits + Fund", "Discounted Creative Cloud + $20K grant", "$5,000+/yr", "YES", "NO - Need sponsor", "adobe.com/nonprofits.html", "7/10", "Grant window Mar 16 - May 16 2026. Move fast."],
    ["9", "Zoom", "Zoom for Nonprofits", "50% off paid plans", "$1,000+/yr", "YES", "NO - Need sponsor", "zoom.com/en/zoom-cares", "6/10", "Moderate savings. Free Zoom covers basic needs."],
    ["10", "Slack", "Slack for Nonprofits", "Free Pro plan (under 250 members)", "$960/yr", "YES", "NO - Need sponsor", "slack.com nonprofit discount", "5/10", "Nice to have, not critical right now."],
    ["11", "Mailchimp", "Nonprofit Discount", "15% off paid plans", "$200-500/yr", "YES", "NO - Need sponsor", "mailchimp.com nonprofit discount", "4/10", "Small discount. Not worth solo effort."],
    ["12", "HeyGen", "None", "NO program exists", "$0", "N/A", "N/A", "No program available", "0/10", "No discount. Use free plan or find alternative."],
]

for i, r in enumerate(rows):
    if "YES - Apply" in r[6] or r[6].startswith("YES"):
        fill = green_fill
    elif "MAYBE" in r[6] or "PARTIAL" in r[6]:
        fill = yellow_fill
    else:
        fill = red_fill
    add_row(ws1, i + 2, r, fill)

for i, w in enumerate([6, 14, 28, 40, 14, 14, 22, 35, 8, 50]):
    ws1.column_dimensions[chr(65 + i)].width = w

# === SHEET 2: INTRO LETTER ===
ws2 = wb.create_sheet("Intro Letter")
ws2.column_dimensions["A"].width = 100

letter = [
    "EDUCATION PARTNERSHIP INTRODUCTORY LETTER",
    "Use this template for each company. Customize the [brackets].",
    "",
    "---",
    "",
    "Subject: Education Partnership Inquiry - AI Training for Adults 50+ | Learn More Technologies",
    "",
    "Dear [Program Name] Team,",
    "",
    "My name is Brian McKinney, founder of Learn More Technologies and the 50+TechBridge program.",
    "We are an MBE-certified AI and digital skills training provider based in Austin, Texas,",
    "serving adults 50 and older - the most digitally underserved population in America.",
    "",
    "I am writing to explore an education partnership with [Company Name].",
    "",
    "Here is why we are a fit for [Program Name]:",
    "",
    "MISSION ALIGNMENT:",
    "- 22 million adults 50+ are digitally underserved in the U.S.",
    "- Our program runs at 3X the industry completion rate",
    "- 74% of participants report increased confidence with technology",
    "- We deliver training at senior centers, libraries, churches, and community orgs",
    "- Free foundational lessons - no paywall between people and learning",
    "",
    "WHAT WE NEED:",
    "- [Specific ask: tool access, credits, curriculum partnership, grant funding]",
    "- These resources would directly serve adults 50+ learning AI for the first time",
    "- Every license/credit translates to a real person gaining digital skills",
    "",
    "WHAT WE BRING:",
    "- MBE certification (State of Texas)",
    "- Proven curriculum deployed across 23 organizations",
    "- 3X industry completion rate with documented outcomes",
    "- Boots-on-the-ground delivery in underserved communities",
    "- A population that [Company] tools are not currently reaching",
    "",
    "We are not asking for charity. We are offering to put [Company] tools in the hands",
    "of the most experienced, most loyal, and most overlooked workforce in America.",
    "Your tools deserve users who will actually use them. Our Pioneers will.",
    "",
    "I would welcome 15 minutes to discuss how a partnership could work.",
    "calendly.com/brianmckinney/new-meeting",
    "",
    "Brian McKinney",
    "Founder, Learn More Technologies | 50+TechBridge",
    "MBE Certified - Austin, Texas",
    "brian@learnmoretechnologies.com",
    "(512) 200-4241",
    "learnmoretechnologies.com",
    "",
    "---",
    "",
    "CUSTOMIZATION NOTES:",
    "",
    "For Cisco: Emphasize free curriculum for Pioneers. Ask about Networking Academy partnership.",
    "For ElevenLabs: Reference your existing voice clone. Ask for Impact Program Pro license.",
    "For Google: Reference AI Opportunity Fund. Emphasize digital inclusion for 50+.",
    "For Anthropic: You are a paying customer. Ask about education/community pricing.",
    "For Microsoft: Lead with MBE. Ask about Azure credits + M365 for training delivery.",
    "For Canva: You already use Canva. Ask about education tier access.",
    "For Adobe: Grant window closes May 16. Apply for $20K Community Fund NOW.",
]

for i, line in enumerate(letter):
    cell = ws2.cell(row=i + 1, column=1, value=line)
    if i == 0 or line.startswith("MISSION") or line.startswith("WHAT WE") or line.startswith("CUSTOMIZATION"):
        cell.font = Font(bold=True, size=12)
    elif line.startswith("For "):
        cell.font = Font(bold=True, size=11)
    else:
        cell.font = normal_font

# === SHEET 3: THE PLAY ===
ws3 = wb.create_sheet("The Play")
ws3.column_dimensions["A"].width = 85

lines = [
    "THE PLAY: HOW LMT ACCESSES THESE PROGRAMS AS A FOR-PROFIT LLC",
    "",
    "PROBLEM: 11 of 12 programs require 501(c)(3). LMT is for-profit.",
    "",
    "SOLUTION 1: FISCAL SPONSOR (unlocks everything)",
    "Find ONE nonprofit partner willing to be your fiscal sponsor.",
    "They apply for the programs. You get the tools through them.",
    "Same blocker as Texas Mutual grant - solve once, unlock 11 programs + grants.",
    "Best candidates: Goodwill Central TX, Austin Free-Net, Capital IDEA, ACC Foundation",
    "",
    "SOLUTION 2: APPLY DIRECTLY WHERE POSSIBLE",
    "- Cisco Networking Academy: No 501(c)(3) needed. Apply NOW.",
    "- ElevenLabs Impact: Flexible eligibility. Apply as education provider.",
    "- Google AI Opportunity Fund: Some tracks open to non-501(c)(3).",
    "- DOL TechAccess: Federal resources for all training providers.",
    "",
    "SOLUTION 3: FORM A NONPROFIT ARM",
    "Create 501(c)(3) alongside LMT LLC.",
    "LLC = paid workshops/contracts. Nonprofit = grants/free programs.",
    "Common social enterprise structure. Takes 3-6 months.",
    "",
    "ESTIMATED TOTAL VALUE IF ALL ACCESSED:",
    "Cisco: Free curriculum (priceless)",
    "ElevenLabs: $1,188/yr",
    "Anthropic: $2,400+/yr",
    "OpenAI: $1,800+/yr",
    "Microsoft: $10,000+/yr",
    "Canva: $9,000/yr",
    "Adobe: $5,000+/yr + $20K grant",
    "Google: $10K-$50K grants",
    "",
    "TOTAL: $30,000 - $80,000+/year in savings and grants",
    "",
    "ACTION ITEMS:",
    "1. Apply to Cisco Networking Academy TODAY (no blocker)",
    "2. Apply to ElevenLabs Impact Program TODAY (no blocker)",
    "3. Secure fiscal sponsor by June 15 (unlocks everything else)",
    "4. Apply to Adobe Community Fund before May 16 deadline",
]

for i, line in enumerate(lines):
    cell = ws3.cell(row=i + 1, column=1, value=line)
    if i == 0 or line.startswith("SOLUTION") or line.startswith("PROBLEM") or line.startswith("ESTIMATED") or line.startswith("ACTION"):
        cell.font = Font(bold=True, size=12)
    elif line.startswith("TOTAL"):
        cell.font = Font(bold=True, size=14, color="109F35")
    else:
        cell.font = normal_font

path = r"C:\Users\USER\Desktop\LMT\ops\LMT-EDUCATION-PARTNERSHIPS.xlsx"
wb.save(path)
print(f"Saved: {path}")
