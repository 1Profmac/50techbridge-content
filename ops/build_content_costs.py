"""Build Content Production Cost Breakdown Excel."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

navy = PatternFill(start_color="0E1C2F", end_color="0E1C2F", fill_type="solid")
gold = PatternFill(start_color="C8942E", end_color="C8942E", fill_type="solid")
orange = PatternFill(start_color="E8733A", end_color="E8733A", fill_type="solid")
green_bg = PatternFill(start_color="109F35", end_color="109F35", fill_type="solid")
green_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
yellow_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
red_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
light_blue = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
white_font = Font(color="FFFFFF", bold=True, size=11)
normal_font = Font(size=11)
bold_font = Font(bold=True, size=11)
big_bold = Font(bold=True, size=13)
huge_bold = Font(bold=True, size=16)
wrap = Alignment(wrap_text=True, vertical="top")
border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def add_header(ws, row, headers, fill=navy):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def add_row(ws, row, data, fill=None):
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = normal_font
        cell.alignment = wrap
        cell.border = border
        if fill:
            cell.fill = fill


# ===== SHEET 1: COST PER CONTENT TYPE =====
ws1 = wb.active
ws1.title = "Cost Per Content Type"

h1 = ["Content Type", "AI Method", "AI Tools Used", "AI Cost/Unit", "DIY Method", "DIY Tools Used", "DIY Cost/Unit", "Time (AI)", "Time (DIY)", "Reach Potential", "Conversion Power"]
add_header(ws1, 1, h1)

rows1 = [
    ["WordPress Blog Post", "Paste from .md, publish via wp-publish.py", "Claude Max + WP", "$0 (already written)", "Same", "Same", "$0", "5 min", "5 min", "HIGH (SEO)", "MED — drives search traffic"],
    ["LinkedIn Article", "Auto-posts from WP on publish", "WP plugin", "$0", "Manual paste to LinkedIn", "LinkedIn", "$0", "0 min (auto)", "10 min", "MED (network)", "HIGH — decision makers read"],
    ["LinkedIn Feed Post", "Write teaser, paste, post", "Claude Max", "$0", "Same", "Same", "$0", "5 min", "5 min", "MED-HIGH", "MED — stops the scroll"],
    ["YouTube Long Video (5-10 min) — AI", "HeyGen talking head + B-Roll", "HeyGen ($24/mo) + Canva ($13/mo)", "$3-5/video", "N/A", "N/A", "N/A", "2-3 hrs", "N/A", "HIGH (algorithm)", "HIGH — authority builder"],
    ["YouTube Long Video — DIY Camera", "Brian on camera + B-Roll overlay", "iPhone + Canva ($13/mo)", "$0-1/video", "Record yourself + edit in Canva", "iPhone + Canva", "$0-1/video", "N/A", "2-4 hrs", "HIGH (algorithm)", "HIGHEST — real face, real trust"],
    ["YouTube Short (< 60 sec) — AI", "HeyGen avatar + text overlay", "HeyGen ($24/mo) + Canva", "$1-2/short", "N/A", "N/A", "N/A", "30-45 min", "N/A", "VERY HIGH (discovery)", "LOW — awareness only"],
    ["YouTube Short — DIY Camera", "Brian on camera, 30-60 sec clip", "iPhone + Canva", "$0", "Record 1 key point + caption", "iPhone + Canva", "$0", "N/A", "15-30 min", "VERY HIGH", "LOW-MED — builds familiarity"],
    ["Podcast Episode (10-20 min)", "ElevenLabs voice clone reads article", "ElevenLabs ($18/mo)", "$0.50-1/ep", "Brian reads into mic", "iPhone/USB mic + Buzzsprout", "$0/ep (free trial)", "30 min", "30-45 min", "MED (Spotify/Apple)", "HIGH — intimacy, trust"],
    ["Podcast Episode — DIY Voice", "Brian talks naturally, no script", "USB mic + Buzzsprout", "$0/ep", "Same", "Same", "$0/ep", "N/A", "20-30 min", "MED", "HIGHEST — authentic voice"],
    ["Email Newsletter", "Paste excerpt + link to article", "Mailchimp ($13/mo)", "$0.50/send", "Same", "Same", "$0.50/send", "10 min", "10 min", "MED (list size)", "HIGH — direct to inbox"],
    ["ElevenLabs Audio Article", "Full article as audio on website", "ElevenLabs ($18/mo)", "$0.50-1/article", "N/A", "N/A", "N/A", "15 min", "N/A", "LOW (website only)", "MED — accessibility"],
]

for i, r in enumerate(rows1):
    fill = green_fill if "$0" in r[3] else yellow_fill if "$0-1" in r[3] or "$0.50" in r[3] else None
    add_row(ws1, i + 2, r, fill)

for i, w in enumerate([22, 35, 28, 14, 28, 22, 14, 12, 12, 16, 18]):
    ws1.column_dimensions[chr(65 + i)].width = w


# ===== SHEET 2: 30 ARTICLES x CONTENT MATRIX =====
ws2 = wb.create_sheet("30 Articles x Content Plan")

h2 = ["#", "Article Title", "WP Blog", "LinkedIn Article", "LI Feed Post", "YT Long (AI)", "YT Long (DIY)", "YT Short (AI)", "YT Short (DIY)", "Podcast (AI)", "Podcast (DIY)", "Email", "Total Assets"]
add_header(ws2, 1, h2, gold)

# Recommended content per article type
articles = [
    ["1", "Why Your 50+ Workforce Is Untapped", "Y", "AUTO", "Y", "", "Y", "", "Y", "", "Y", "Y", "6"],
    ["2", "$850B Cost of Ignoring Workers", "Y", "AUTO", "Y", "Y", "", "Y", "", "Y", "", "Y", "6"],
    ["3", "Digital Divide for Adults Over 50", "Y", "AUTO", "Y", "", "", "", "Y", "Y", "", "Y", "5"],
    ["4", "The Confidence Gap", "Y", "AUTO", "Y", "", "Y", "", "Y", "", "Y", "Y", "6"],
    ["5", "$76 Trillion Blind Spot", "Y", "AUTO", "Y", "Y", "", "Y", "", "Y", "", "Y", "6"],
    ["6", "It's Not About the Data", "Y", "AUTO", "Y", "", "Y", "", "Y", "", "Y", "Y", "6"],
    ["7", "What Is AgeTech? Complete Guide", "Y", "AUTO", "Y", "Y", "", "", "Y", "Y", "", "Y", "6"],
    ["8", "What 23 Orgs Learned", "Y", "AUTO", "Y", "", "Y", "", "Y", "", "Y", "Y", "6"],
    ["9", "Pioneer Stories: Real Results", "Y", "AUTO", "Y", "", "Y", "", "Y", "", "Y", "Y", "6"],
    ["10", "AI Training for Older Workers", "Y", "AUTO", "Y", "Y", "", "Y", "", "Y", "", "Y", "6"],
    ["11", "Corporate AI Workshops ROI", "Y", "AUTO", "Y", "", "Y", "", "Y", "", "Y", "Y", "6"],
    ["12", "Agentic AI and 50+ Workforce", "Y", "AUTO", "Y", "Y", "", "Y", "", "Y", "", "Y", "6"],
    ["13", "50-Year-Old Founders Outperform", "Y", "AUTO", "Y", "", "", "", "Y", "Y", "", "Y", "5"],
    ["14", "Deploy AI Training This Quarter", "Y", "AUTO", "Y", "Y", "", "", "Y", "", "Y", "Y", "6"],
    ["15", "I Started at 65", "Y", "AUTO", "Y", "", "Y", "", "Y", "", "Y", "Y", "6"],
    ["16", "The Agentic Advantage", "Y", "AUTO", "Y", "Y", "", "Y", "", "Y", "", "Y", "6"],
    ["17", "AI Is for Young Tech People?", "Y", "AUTO", "Y", "", "", "", "Y", "Y", "", "Y", "5"],
    ["18", "Will AI Replace Me?", "Y", "AUTO", "Y", "", "Y", "Y", "", "", "Y", "Y", "6"],
    ["19", "Too Old to Learn AI?", "Y", "AUTO", "Y", "", "Y", "", "Y", "", "Y", "Y", "6"],
    ["20", "AI Is Just ChatGPT for Fun?", "Y", "AUTO", "Y", "", "", "Y", "", "Y", "", "Y", "5"],
    ["21", "I Need a Course First?", "Y", "AUTO", "Y", "", "", "", "Y", "Y", "", "Y", "5"],
    ["22", "Start a Business After 50", "Y", "AUTO", "Y", "Y", "", "Y", "", "Y", "", "Y", "6"],
    ["23", "Digital Skills Training Guide", "Y", "AUTO", "Y", "Y", "", "", "Y", "", "Y", "Y", "6"],
    ["24", "WIOA Funds AI Training", "Y", "AUTO", "Y", "", "", "", "Y", "Y", "", "Y", "5"],
    ["25", "WIOA Mandates Digital Skills", "Y", "AUTO", "Y", "", "", "Y", "", "Y", "", "Y", "5"],
    ["26", "How to Write WIOA Proposal", "Y", "AUTO", "Y", "", "", "", "", "Y", "", "Y", "4"],
    ["27", "What Is ETPL?", "Y", "AUTO", "Y", "", "", "", "", "Y", "", "Y", "4"],
    ["28", "MBE + Workforce Contracts", "Y", "AUTO", "Y", "", "", "", "Y", "", "", "Y", "4"],
    ["29", "MBE Vendors in Texas", "Y", "AUTO", "Y", "", "", "", "", "", "", "Y", "3"],
    ["30", "Best AI Tools for Adults 50+", "Y", "AUTO", "Y", "Y", "", "Y", "", "Y", "", "Y", "6"],
]

for i, r in enumerate(articles):
    add_row(ws2, i + 2, r)

# Totals row
total_row = len(articles) + 2
ws2.cell(row=total_row, column=1, value="TOTAL").font = big_bold
ws2.cell(row=total_row, column=2, value="30 articles").font = bold_font
ws2.cell(row=total_row, column=3, value="30").font = bold_font
ws2.cell(row=total_row, column=4, value="30").font = bold_font
ws2.cell(row=total_row, column=5, value="30").font = bold_font
ws2.cell(row=total_row, column=6, value="10").font = bold_font
ws2.cell(row=total_row, column=7, value="10").font = bold_font
ws2.cell(row=total_row, column=8, value="10").font = bold_font
ws2.cell(row=total_row, column=9, value="20").font = bold_font
ws2.cell(row=total_row, column=10, value="15").font = bold_font
ws2.cell(row=total_row, column=11, value="10").font = bold_font
ws2.cell(row=total_row, column=12, value="28").font = bold_font
ws2.cell(row=total_row, column=13, value="~165").font = Font(bold=True, size=14, color="109F35")

for i, w in enumerate([4, 32, 8, 12, 10, 10, 10, 10, 10, 10, 10, 8, 10]):
    ws2.column_dimensions[chr(65 + i)].width = w


# ===== SHEET 3: TOTAL PRODUCTION COST =====
ws3 = wb.create_sheet("Total Production Cost")

h3 = ["Content Type", "Qty (30 articles)", "Method", "Tool Cost/mo", "Per-Unit Cost", "Total Cost", "Total Time"]
add_header(ws3, 1, h3, orange)

costs = [
    ["WordPress Blog Posts", "30", "wp-publish.py (auto)", "$0 (included in Max)", "$0", "$0", "2.5 hrs"],
    ["LinkedIn Articles", "30", "Auto-post from WP", "$0 (WP plugin)", "$0", "$0", "0 hrs (auto)"],
    ["LinkedIn Feed Posts", "30", "Manual teaser posts", "$0", "$0", "$0", "2.5 hrs"],
    ["YouTube Long — AI (HeyGen)", "10", "HeyGen talking head + Canva B-Roll", "$24/mo HeyGen", "$3-5 each", "$30-50", "25 hrs"],
    ["YouTube Long — DIY Camera", "10", "iPhone + Canva edit", "$0 (own Canva)", "$0", "$0", "30 hrs"],
    ["YouTube Shorts — AI", "10", "HeyGen 30-sec clips", "$0 (included in HeyGen)", "$1-2 each", "$10-20", "5 hrs"],
    ["YouTube Shorts — DIY Camera", "20", "iPhone 30-sec clips + Canva captions", "$0", "$0", "$0", "8 hrs"],
    ["Podcast — AI Voice (ElevenLabs)", "15", "Clone reads article, upload Buzzsprout", "$18/mo ElevenLabs", "$0.50-1 each", "$8-15", "7.5 hrs"],
    ["Podcast — DIY Brian Reads", "10", "USB mic, talk naturally, Buzzsprout", "$0 (free trial)", "$0", "$0", "5 hrs"],
    ["Email Newsletter", "28", "Excerpt + link via Mailchimp", "$13/mo Mailchimp", "$0.50 each", "$14", "4.5 hrs"],
    ["Audio Articles (website)", "10", "ElevenLabs renders for site embed", "$0 (included)", "$0.50 each", "$5", "2.5 hrs"],
]

for i, r in enumerate(costs):
    fill = green_fill if r[5] == "$0" else yellow_fill
    add_row(ws3, i + 2, r, fill)

# Summary
s = len(costs) + 3
ws3.cell(row=s, column=1, value="PRODUCTION SUMMARY").font = huge_bold

ws3.cell(row=s+1, column=1, value="Total content pieces from 30 articles:").font = bold_font
ws3.cell(row=s+1, column=3, value="~165 pieces").font = Font(bold=True, size=14, color="109F35")

ws3.cell(row=s+2, column=1, value="Total production time:").font = bold_font
ws3.cell(row=s+2, column=3, value="~92 hours (spread over 30 days = 3 hrs/day)").font = bold_font

ws3.cell(row=s+3, column=1, value="Total variable production cost:").font = bold_font
ws3.cell(row=s+3, column=3, value="$67-104 (one-time, on top of monthly subscriptions)").font = bold_font

ws3.cell(row=s+5, column=1, value="MONTHLY TOOL COST FOR CONTENT:").font = huge_bold
ws3.cell(row=s+6, column=1, value="HeyGen (talking head videos)").font = normal_font
ws3.cell(row=s+6, column=3, value="$24/mo").font = normal_font
ws3.cell(row=s+7, column=1, value="ElevenLabs (voice clone + podcasts)").font = normal_font
ws3.cell(row=s+7, column=3, value="$18/mo").font = normal_font
ws3.cell(row=s+8, column=1, value="Canva (graphics, video editing, thumbnails)").font = normal_font
ws3.cell(row=s+8, column=3, value="$13/mo").font = normal_font
ws3.cell(row=s+9, column=1, value="Mailchimp (email newsletters)").font = normal_font
ws3.cell(row=s+9, column=3, value="$13/mo").font = normal_font
ws3.cell(row=s+10, column=1, value="Buzzsprout (podcast hosting)").font = normal_font
ws3.cell(row=s+10, column=3, value="$0 (free trial til Jul 11, then $19/mo)").font = normal_font
ws3.cell(row=s+11, column=1, value="Claude Max (writing, scripts, wp-publish)").font = normal_font
ws3.cell(row=s+11, column=3, value="$100/mo (already paying)").font = normal_font
ws3.cell(row=s+12, column=1, value="CONTENT STACK MONTHLY TOTAL:").font = big_bold
ws3.cell(row=s+12, column=3, value="$68/mo ($168 with Claude Max)").font = Font(bold=True, size=14, color="C0392B")

ws3.cell(row=s+14, column=1, value="COST PER CONTENT PIECE:").font = huge_bold
ws3.cell(row=s+15, column=1, value="165 pieces / $168 monthly tools =").font = bold_font
ws3.cell(row=s+15, column=3, value="$1.02 per piece").font = Font(bold=True, size=16, color="109F35")

ws3.cell(row=s+17, column=1, value="vs hiring a content team:").font = bold_font
ws3.cell(row=s+17, column=3, value="$5,000-$15,000/mo for same output").font = Font(bold=True, size=12, color="C0392B")

for i, w in enumerate([35, 18, 38, 22, 14, 14, 30]):
    ws3.column_dimensions[chr(65 + i)].width = w


# ===== SHEET 4: STRATEGY RANKING =====
ws4 = wb.create_sheet("Strategy Ranking")

h4 = ["Rank", "Content Type", "Reach", "Trust", "Conversion", "Cost", "Time", "Overall Score", "Verdict"]
add_header(ws4, 1, h4, navy)

strategy = [
    ["1", "YouTube Long (DIY Camera)", "HIGH", "HIGHEST", "HIGH", "FREE", "2-4 hrs", "10/10", "MOST POTENT. Your face, your voice, your conviction. Nothing beats it."],
    ["2", "WordPress Blog + LinkedIn Article", "HIGH (SEO)", "HIGH", "MED-HIGH", "FREE", "5 min (auto)", "9/10", "FOUNDATION. Google finds you. Decision makers read you. Already built."],
    ["3", "Podcast (DIY Voice)", "MED", "HIGHEST", "HIGH", "FREE", "$20-30 min", "9/10", "INTIMATE. Voice builds trust faster than text. Just talk."],
    ["4", "YouTube Shorts (DIY Camera)", "VERY HIGH", "MED", "LOW-MED", "FREE", "15-30 min", "8/10", "DISCOVERY. Algorithm feeds new eyeballs. Quick to make. Stack daily."],
    ["5", "LinkedIn Feed Posts", "MED-HIGH", "MED", "MED", "FREE", "5 min", "8/10", "DAILY PRESENCE. Keeps you visible. Teasers drive to articles."],
    ["6", "Email Newsletter", "MED", "HIGH", "HIGH", "$13/mo", "10 min", "7/10", "DIRECT LINE. Your list, your audience. Converts when list grows."],
    ["7", "YouTube Long (HeyGen AI)", "HIGH", "MED", "MED", "$24/mo", "2-3 hrs", "6/10", "SCALABLE but less authentic. Use for B-Roll and explainers, not keynotes."],
    ["8", "Podcast (ElevenLabs AI Voice)", "MED", "MED", "MED", "$18/mo", "30 min", "6/10", "EFFICIENT but not your real voice. Use for article audio versions."],
    ["9", "YouTube Shorts (HeyGen AI)", "VERY HIGH", "LOW-MED", "LOW", "$24/mo", "30-45 min", "5/10", "AI avatar feels less authentic in Shorts. DIY camera wins here."],
    ["10", "Audio Articles (ElevenLabs)", "LOW", "MED", "LOW", "$18/mo", "15 min", "4/10", "ACCESSIBILITY play. Good for website, low discovery."],
]

for i, r in enumerate(strategy):
    if int(r[0]) <= 3:
        fill = green_fill
    elif int(r[0]) <= 6:
        fill = yellow_fill
    else:
        fill = red_fill
    add_row(ws4, i + 2, r, fill)

# Bottom line
b = len(strategy) + 3
ws4.cell(row=b, column=1, value="THE PLAY:").font = huge_bold
ws4.cell(row=b+1, column=1, value="TIER 1 (do every day, FREE):").font = big_bold
ws4.cell(row=b+2, column=1, value="Blog + LinkedIn auto-post (5 min) + LinkedIn teaser (5 min) + 1 DIY Short (15 min) = 25 min/day").font = bold_font
ws4.cell(row=b+4, column=1, value="TIER 2 (do weekly, FREE or cheap):").font = big_bold
ws4.cell(row=b+5, column=1, value="1 YouTube Long DIY (2-4 hrs) + 1 Podcast DIY (30 min) + 1 Email newsletter (10 min) = 3-5 hrs/week").font = bold_font
ws4.cell(row=b+7, column=1, value="TIER 3 (use AI to fill gaps):").font = big_bold
ws4.cell(row=b+8, column=1, value="HeyGen videos for articles you don't want to film. ElevenLabs for articles you don't want to read. Supplement, don't replace.").font = bold_font
ws4.cell(row=b+10, column=1, value="COST OF THE MOST POTENT STRATEGY:").font = huge_bold
ws4.cell(row=b+11, column=1, value="Tier 1 + Tier 2 = $0-13/mo (just Mailchimp + tools you already have)").font = Font(bold=True, size=14, color="109F35")
ws4.cell(row=b+12, column=1, value="Add Tier 3 for scale = $55/mo (HeyGen + ElevenLabs)").font = bold_font

for i, w in enumerate([6, 28, 10, 10, 12, 10, 12, 12, 60]):
    ws4.column_dimensions[chr(65 + i)].width = w


path = r"C:\Users\USER\Desktop\LMT\ops\LMT-CONTENT-PRODUCTION-COSTS.xlsx"
wb.save(path)
print(f"Saved: {path}")
