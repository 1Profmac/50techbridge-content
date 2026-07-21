"""
Build LMT Master Daily Playbook — single Excel workbook with all daily/weekly ops.
Replaces 6 separate markdown files with one master reference.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============================================================
# STYLES
# ============================================================
navy_fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
gold_fill = PatternFill(start_color="C5972C", end_color="C5972C", fill_type="solid")
light_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
red_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")

header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
gold_header_font = Font(name="Calibri", bold=True, color="1B2A4A", size=12)
title_font = Font(name="Calibri", bold=True, size=14)
bold_font = Font(name="Calibri", bold=True, size=11)
normal_font = Font(name="Calibri", size=11)
small_font = Font(name="Calibri", size=10, italic=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header_row(ws, row, cols, fill=navy_fill, font=header_font):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def style_data_row(ws, row, cols, alt=False):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = normal_font
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        cell.border = thin_border
        if alt:
            cell.fill = light_gray

# ============================================================
# TAB 1: DAILY PLAYBOOK
# ============================================================
ws = wb.active
ws.title = "Daily Playbook"
ws.sheet_properties.tabColor = "1B2A4A"

# Column widths
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 45
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 15

# Title
ws.merge_cells('A1:E1')
ws['A1'] = "LMT DAILY PLAYBOOK — The Rule: Calls First. LinkedIn Second. Content Last."
ws['A1'].font = title_font

ws.merge_cells('A2:E2')
ws['A2'] = "2 phone calls before anything else. Every day. No exceptions."
ws['A2'].font = Font(name="Calibri", bold=True, size=11, color="CC0000")

# Headers
headers = ["Block", "Time", "Tasks", "Duration", "Status"]
for i, h in enumerate(headers, 1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, 5)

# Daily blocks
blocks = [
    ["1", "8:00–8:30", "PHONE CALLS — Make 2 calls minimum.\n• Work through weekly call list in order\n• Leave voicemails\n• Send follow-up emails immediately after\n• Use voicemail script (see Calls tab)", "30 min", ""],
    ["2", "8:30–8:55", "LINKEDIN MORNING HUNT\n• Work Warm Signal Stack (10 min)\n• Send 15–20 connection requests to Tier 1 buyers (10 min)\n• Send 3–5 follow-up DMs with specific ask (5 min)", "25 min", ""],
    ["3", "8:55–9:10", "CONTENT + DISTRIBUTION\n• Publish 1 article from ready queue\n• Post to LinkedIn (auto or manual)\n• Post to FB Business page\n• Comment on 2–3 FB group posts (helpful, no links)", "15 min", ""],
    ["4", "9:10–9:20", "PIPELINE + METRICS\n• Check Gmail for replies (workforce pipeline label)\n• Update outreach tracker\n• Log calls made\n• Check platform metrics (daily quick scan)", "10 min", ""],
]

for i, block in enumerate(blocks):
    row = 5 + i
    for j, val in enumerate(block):
        ws.cell(row=row, column=j+1, value=val)
    style_data_row(ws, row, 5, alt=(i % 2 == 1))

# Midday + Evening (content days only)
row = 10
ws.cell(row=row, column=1, value="")
ws.merge_cells(f'A{row}:E{row}')
ws.cell(row=row, column=1, value="CONTENT DAYS ONLY (Tue, Thu) — add these blocks:")
ws.cell(row=row, column=1).font = bold_font

extra_blocks = [
    ["5", "Midday", "MIDDAY SIGNAL\n• Comment on 3 posts by Tier 1 decision-makers\n• Save 1 post for tomorrow's DM context", "10 min", ""],
    ["6", "Evening", "EVENING SHIP\n• Post 1 piece of content\n• Log DMs/accepts in correspondence log\n• Update CEO-DASHBOARD scoreboard (Mondays only)\n• Close LinkedIn — no scrolling after this", "10 min", ""],
]

for i, block in enumerate(extra_blocks):
    row = 11 + i
    for j, val in enumerate(block):
        ws.cell(row=row, column=j+1, value=val)
    style_data_row(ws, row, 5, alt=(i % 2 == 1))

# Focus Lock
row = 14
ws.merge_cells(f'A{row}:E{row}')
ws.cell(row=row, column=1, value="FOCUS LOCK — Read aloud every morning before Block 1")
ws.cell(row=row, column=1).font = bold_font
ws.cell(row=row, column=1).fill = gold_fill

focus_items = [
    "I am MBE-certified. 347 completions. 3X industry completion rate.",
    "Agentic 50 / 50+ Tech Bridge. The only niche: tech training for adults 50+.",
    "Host, not attendee. Teacher, not student. Prime, not sub.",
    "If it doesn't move toward a signed contract or PO, I don't do it.",
    "Phone calls first. Applications second. LinkedIn third. Content last.",
]
for i, item in enumerate(focus_items):
    row = 15 + i
    ws.merge_cells(f'A{row}:E{row}')
    ws.cell(row=row, column=1, value=f"• {item}")
    ws.cell(row=row, column=1).font = normal_font

# ============================================================
# TAB 2: WEEKLY SCHEDULE
# ============================================================
ws2 = wb.create_sheet("Weekly Schedule")
ws2.sheet_properties.tabColor = "C5972C"

ws2.column_dimensions['A'].width = 12
ws2.column_dimensions['B'].width = 20
ws2.column_dimensions['C'].width = 50
ws2.column_dimensions['D'].width = 15

ws2.merge_cells('A1:D1')
ws2['A1'] = "WEEKLY SCHEDULE — Learn More Technologies"
ws2['A1'].font = title_font

headers = ["Day", "Focus", "Tasks", "Total Time"]
for i, h in enumerate(headers, 1):
    ws2.cell(row=3, column=i, value=h)
style_header_row(ws2, 3, 4)

days = [
    ["MONDAY", "Content Batch", "• Morning Hunt (25 min)\n• Batch week's LinkedIn posts (2 hrs)\n• Publish 1 article from queue (30 min)\n• Rest of day: pitches, calls, proposals", "~3 hrs"],
    ["TUESDAY", "Heavy Production", "• Full 3-block LinkedIn (45 min)\n• YouTube / Short production — HeyGen or iPhone (2–4 hrs)\n• Afternoon: rest or short admin", "~5 hrs"],
    ["WEDNESDAY", "Phone + Admin", "• Morning Hunt (25 min)\n• Phone calls — work through call list (60 min)\n• Applications: STARS, SCSEP, open RFPs (60 min)\n• Rest of day: flex", "~2.5 hrs"],
    ["THURSDAY", "Content Day", "• Full 3-block LinkedIn (45 min)\n• Publish Thursday series post (30 min)\n• Warm contact follow-ups\n• Lunch & learn pitch drafting", "~2 hrs"],
    ["FRIDAY", "Review + Close", "• Morning Hunt (25 min)\n• Weekly review + CEO-DASHBOARD scoreboard (30 min)\n• Git commit (15 min)\n• Flex: calls, prep for Monday batch", "~1.5 hrs"],
    ["SAT/SUN", "Off / Flex", "• No mandatory work\n• Caregiving, rest, optional backlog\n• Exception: if Tier 1 replies, take the call", "0"],
]

for i, day in enumerate(days):
    row = 4 + i
    for j, val in enumerate(day):
        ws2.cell(row=row, column=j+1, value=val)
    style_data_row(ws2, row, 4, alt=(i % 2 == 1))
    ws2.cell(row=row, column=1).font = bold_font

# Weekly targets
row = 12
ws2.merge_cells(f'A{row}:D{row}')
ws2.cell(row=row, column=1, value="WEEKLY TARGETS")
ws2.cell(row=row, column=1).font = title_font

targets_headers = ["Metric", "Target", "Actual", "Notes"]
row = 13
for i, h in enumerate(targets_headers, 1):
    ws2.cell(row=row, column=i, value=h)
style_header_row(ws2, row, 4)

targets = [
    ["Phone calls made", "10/week (2/day)", "", ""],
    ["Lunch-and-learns offered", "5/week", "", ""],
    ["LinkedIn requests sent", "75–100/week", "", ""],
    ["Articles published", "2/week", "", ""],
    ["YouTube Shorts published", "1/week", "", ""],
    ["Applications filed", "1+/week", "", ""],
    ["Pipeline conversations", "2/week", "", ""],
    ["Pipeline value added", "$5K+/week", "", ""],
]

for i, t in enumerate(targets):
    row = 14 + i
    for j, val in enumerate(t):
        ws2.cell(row=row, column=j+1, value=val)
    style_data_row(ws2, row, 4, alt=(i % 2 == 1))

# ============================================================
# TAB 3: CALLS & CONTACTS
# ============================================================
ws3 = wb.create_sheet("Calls & Contacts")
ws3.sheet_properties.tabColor = "CC0000"

ws3.column_dimensions['A'].width = 5
ws3.column_dimensions['B'].width = 25
ws3.column_dimensions['C'].width = 25
ws3.column_dimensions['D'].width = 18
ws3.column_dimensions['E'].width = 18
ws3.column_dimensions['F'].width = 30

ws3.merge_cells('A1:F1')
ws3['A1'] = "CALL LIST — Work Top to Bottom, 2 Calls/Day Minimum"
ws3['A1'].font = title_font

headers = ["#", "Company", "Contact", "Phone", "Status", "Notes"]
for i, h in enumerate(headers, 1):
    ws3.cell(row=3, column=i, value=h)
style_header_row(ws3, 3, 6)

contacts = [
    [1, "Texas Mutual Insurance", "Larry Martin, SVP HR", "(512) 224-4700", "Not called", ""],
    [2, "National Western Life", "Kim Gibbs, VP HR", "(512) 836-1010", "Not called", ""],
    [3, "City of Austin HR", "Rebecca Kennedy, Deputy Dir", "(512) 974-3400", "Not called", ""],
    [4, "TxDOT", "Christopher Young, HR Dir", "(512) 463-8588", "Not called", ""],
    [5, "LCRA", "Stephanie Taylor, HR Mgr", "(512) 578-4004", "Not called", ""],
    [6, "Travis County", "HR Director", "(512) 854-9165", "Not called", ""],
    [7, "St. David's / HCA", "Todd Steward", "(512) 544-5000", "Not called", ""],
    [8, "Austin ISD", "Pamela Hall, Exec Dir HR", "(512) 414-1700", "Not called", ""],
    [9, "Baylor Scott & White", "Regional VP HR", "(512) 509-0100", "Not called", ""],
    [10, "Austin Free-Net", "Jasmin Vargas", "(512) 236-8225", "Called Apr 22 (VM)", ""],
]

for i, c in enumerate(contacts):
    row = 4 + i
    for j, val in enumerate(c):
        ws3.cell(row=row, column=j+1, value=val)
    style_data_row(ws3, row, 6, alt=(i % 2 == 1))

# Voicemail script
row = 16
ws3.merge_cells(f'A{row}:F{row}')
ws3.cell(row=row, column=1, value="VOICEMAIL SCRIPT")
ws3.cell(row=row, column=1).font = bold_font
ws3.cell(row=row, column=1).fill = gold_fill

row = 17
ws3.merge_cells(f'A{row}:F{row}')
ws3.cell(row=row, column=1, value='"Hi [Name], this is Brian McKinney with Learn More Technologies in Austin. We\'re MBE-certified — AI and digital skills training for adults 50+, 3X industry completion rate. I\'d like to offer your team a free 60-minute lunch-and-learn. No cost, no commitment — just a demo of what we do. My number is 512-200-4241. I\'ll follow up by email."')
ws3.cell(row=row, column=1).font = normal_font
ws3.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
ws3.row_dimensions[row].height = 60

# Live script
row = 19
ws3.merge_cells(f'A{row}:F{row}')
ws3.cell(row=row, column=1, value="PHONE SCRIPT (when someone picks up)")
ws3.cell(row=row, column=1).font = bold_font
ws3.cell(row=row, column=1).fill = gold_fill

scripts = [
    'OPENING (10 sec): "Hi [Name], this is Brian McKinney with Learn More Technologies. I\'m MBE-certified, based in Austin. Do you have 60 seconds?"',
    'IF YES (30 sec): "I train adults 50+ in AI and digital skills. 3X industry completion rate. I know your organization has experienced employees who could benefit from AI training, and I\'d like to offer a free 60-minute lunch-and-learn for your team. No cost, no obligation — I bring everything. If your team likes it, we talk about a paid workshop. If not, you got a free hour of AI training. Can I send you a one-page overview?"',
    'IF THEY SAY SEND IT: "Great — what\'s the best email? I\'ll send the overview and a link to book 15 minutes on my calendar. Thank you, [Name]." → Send Workshop Sell Sheet + Calendly link. Follow up in 3 days.',
    'IF NOT INTERESTED: "I understand. If you know anyone in HR or L&D who\'s thinking about AI training for experienced employees, I\'d appreciate the referral. Thanks for your time."',
]

for i, s in enumerate(scripts):
    row = 20 + i
    ws3.merge_cells(f'A{row}:F{row}')
    ws3.cell(row=row, column=1, value=s)
    ws3.cell(row=row, column=1).font = normal_font
    ws3.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
    ws3.row_dimensions[row].height = 45

# ============================================================
# TAB 4: ACTIVE PROJECTS
# ============================================================
ws4 = wb.create_sheet("Active Projects")
ws4.sheet_properties.tabColor = "4CAF50"

ws4.column_dimensions['A'].width = 30
ws4.column_dimensions['B'].width = 40
ws4.column_dimensions['C'].width = 15
ws4.column_dimensions['D'].width = 15
ws4.column_dimensions['E'].width = 20

ws4.merge_cells('A1:E1')
ws4['A1'] = "ACTIVE PROJECTS & TO-DO"
ws4['A1'].font = title_font

headers = ["Task", "Details", "Deadline", "Status", "Owner"]
for i, h in enumerate(headers, 1):
    ws4.cell(row=3, column=i, value=h)
style_header_row(ws4, 3, 5)

projects = [
    ["Clean LearnDash bot accounts", "Delete random-string usernames from WP Users + LearnDash", "2026-05-02", "TODO", "Brian"],
    ["Create /startfreetoday redirect", "Page or redirect → /courses/50techbridge/", "2026-05-02", "TODO", "Brian"],
    ["Update landing page copy", 'Change "six free modules" to "three free lessons"', "2026-05-02", "TODO", "Brian"],
    ["Delete test accounts", "Remove test@learnmo.com from Mailchimp + WP", "2026-05-02", "TODO", "Brian"],
    ["Launch YouTube video", "ElevenLabs + HeyGen + Canva. Scripts ready.", "2026-05-02", "TODO", "Brian"],
    ["Launch YouTube Short", "15-20 sec vertical. Script ready.", "2026-05-02", "TODO", "Brian"],
    ["Publish WP article", "3 Free Digital Skills Lessons. Draft ready.", "2026-05-02", "TODO", "Brian"],
    ["LinkedIn post", "Announce free lessons. Draft ready.", "2026-05-02", "TODO", "Brian"],
    ["Facebook post", "Share free lessons link. Draft ready.", "2026-05-02", "TODO", "Brian"],
    ["Activate hCaptcha", "Prevent future bot signups on join form", "2026-05-02", "TODO", "Brian"],
    ["Update BuddyBoss plugins", "Platform 2.21.0→2.21.1, Pro 2.13.1→2.13.2", "2026-05-02", "TODO", "Brian"],
    ["TWC Chapter 132 Exemption", "8-step filing process", "In progress", "IN PROGRESS", "Brian"],
    ["ETPL Application", "5-step process after exemption", "After Ch.132", "BLOCKED", "Brian"],
    ["STARS Supplier Portal", "Register for AARP RFP access", "ASAP", "TODO", "Brian"],
    ["SCSEP Host Agency", "Apply at ha.scsep.org — $46.9M grant", "ASAP", "TODO", "Brian"],
    ["Credential rotation", "WP password, GitHub token — done. Mailchimp key — FIXED today.", "DONE", "DONE", "Brian"],
    ["Weekly lives", "Once-a-week live stream — platform TBD", "TBD", "PLANNING", "Brian"],
]

for i, p in enumerate(projects):
    row = 4 + i
    for j, val in enumerate(p):
        ws4.cell(row=row, column=j+1, value=val)
    style_data_row(ws4, row, 5, alt=(i % 2 == 1))
    # Color code status
    status_cell = ws4.cell(row=row, column=4)
    if status_cell.value == "DONE":
        status_cell.fill = green_fill
    elif status_cell.value == "BLOCKED":
        status_cell.fill = red_fill

# ============================================================
# TAB 5: SITE HEALTH
# ============================================================
ws5 = wb.create_sheet("Site Health")
ws5.sheet_properties.tabColor = "2196F3"

ws5.column_dimensions['A'].width = 35
ws5.column_dimensions['B'].width = 40
ws5.column_dimensions['C'].width = 15
ws5.column_dimensions['D'].width = 20

ws5.merge_cells('A1:D1')
ws5['A1'] = "SITE HEALTH — Weekly Check (every Monday, 15 min)"
ws5['A1'].font = title_font

headers = ["Check", "What to Do", "Frequency", "Status"]
for i, h in enumerate(headers, 1):
    ws5.cell(row=3, column=i, value=h)
style_header_row(ws5, 3, 4)

checks = [
    ["Run site-health-check.py", "python Desktop/LMT/ops/site-health-check.py", "Daily", ""],
    ["SSL certificate", "Auto-checked by script. Warns at 14 days.", "Daily", ""],
    ["Critical pages loading", "Homepage, /courses/, /join-now/, /train, /workforce", "Weekly", ""],
    ["Mailchimp API", "Auto-checked by script. Catches disabled keys.", "Daily", ""],
    ["WordPress REST API", "Auto-checked by script.", "Daily", ""],
    ["Google indexing", "Search site:learnmoretechnologies.com", "Weekly", ""],
    ["Google Search Console", "Crawl errors, mobile issues, search performance", "Weekly", ""],
    ["LearnDash enrollments", "Check course enrollment tracking", "Weekly", ""],
    ["Mailchimp audience", "New subscribers appearing?", "Weekly", ""],
    ["Bluehost billing/alerts", "Any renewal notices?", "Weekly", ""],
    ["WP plugin updates", "Check Plugins page for red badge", "Weekly", ""],
    ["Error log", "Bluehost File Manager → error_log", "Weekly", ""],
]

for i, c in enumerate(checks):
    row = 4 + i
    for j, val in enumerate(c):
        ws5.cell(row=row, column=j+1, value=val)
    style_data_row(ws5, row, 4, alt=(i % 2 == 1))

# Issues log
row = 18
ws5.merge_cells(f'A{row}:D{row}')
ws5.cell(row=row, column=1, value="ISSUES LOG")
ws5.cell(row=row, column=1).font = bold_font
ws5.cell(row=row, column=1).fill = gold_fill

issue_headers = ["Date", "Issue", "Status", "Resolution"]
row = 19
for i, h in enumerate(issue_headers, 1):
    ws5.cell(row=row, column=i, value=h)
style_header_row(ws5, row, 4)

issues = [
    ["2026-04-27", "Yoast redirect loop /courses/ ↔ /our-courses/", "FIXED", "Deleted Yoast redirect rule"],
    ["2026-04-27", "Mailchimp API key disabled", "FIXED", "New key: LMT-Join-Now-Form"],
    ["2026-04-27", "Error log spam (128k lines)", "FIXED", "Removed debug logging line"],
    ["2026-04-27", "BuddyBoss registration disabled", "FIXED", "Enabled in BB Settings"],
]

for i, issue in enumerate(issues):
    row = 20 + i
    for j, val in enumerate(issue):
        ws5.cell(row=row, column=j+1, value=val)
    style_data_row(ws5, row, 4)
    if ws5.cell(row=row, column=3).value == "FIXED":
        ws5.cell(row=row, column=3).fill = green_fill

# ============================================================
# TAB 6: OFFER LADDER
# ============================================================
ws6 = wb.create_sheet("Offer Ladder")
ws6.sheet_properties.tabColor = "FF9800"

ws6.column_dimensions['A'].width = 8
ws6.column_dimensions['B'].width = 30
ws6.column_dimensions['C'].width = 18
ws6.column_dimensions['D'].width = 40

ws6.merge_cells('A1:D1')
ws6['A1'] = "THE OFFER LADDER"
ws6['A1'].font = title_font

headers = ["Step", "What", "Cost", "Goal"]
for i, h in enumerate(headers, 1):
    ws6.cell(row=3, column=i, value=h)
style_header_row(ws6, 3, 4)

ladder = [
    [1, "Free lunch-and-learn (60 min)", "Free", "Get in the door"],
    [2, "Paid workshop (half-day)", "$3,500", "First revenue"],
    [3, "Full-day workshop", "$7,500", "Repeat + referral"],
    [4, "Retainer", "$2K–$5K/mo", "Recurring revenue"],
    [5, "WIOA/grant contract", "$30K–$160K/qtr", "Scale play"],
]

for i, l in enumerate(ladder):
    row = 4 + i
    for j, val in enumerate(l):
        ws6.cell(row=row, column=j+1, value=val)
    style_data_row(ws6, row, 4, alt=(i % 2 == 1))

row = 10
ws6.merge_cells(f'A{row}:D{row}')
ws6.cell(row=row, column=1, value="Steps 1–3 can happen in 30 days. Steps 4–5 are the long game (ETPL, Texas Mutual, AARP).")
ws6.cell(row=row, column=1).font = bold_font

# THE MATH
row = 12
ws6.merge_cells(f'A{row}:D{row}')
ws6.cell(row=row, column=1, value="THE MATH")
ws6.cell(row=row, column=1).font = bold_font
ws6.cell(row=row, column=1).fill = gold_fill

math_items = [
    "10 calls/week → 3 conversations → 1 lunch-and-learn offered",
    "4 lunch-and-learns/month → 1 converts to paid workshop",
    "1 paid workshop = $3,500–$7,500",
    "First revenue in 30 days if you start calling today",
]
for i, m in enumerate(math_items):
    row = 13 + i
    ws6.merge_cells(f'A{row}:D{row}')
    ws6.cell(row=row, column=1, value=f"• {m}")
    ws6.cell(row=row, column=1).font = normal_font

# ============================================================
# SAVE
# ============================================================
output_path = r"C:\Users\USER\Desktop\LMT\ops\LMT-MASTER-PLAYBOOK.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
print("Tabs: Daily Playbook | Weekly Schedule | Calls & Contacts | Active Projects | Site Health | Offer Ladder")
