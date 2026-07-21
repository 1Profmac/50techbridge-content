import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Content Command Center"

hf = Font(bold=True, color="FFFFFF", size=10)
hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
blue = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
orange = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
red = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
gray = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
purple = PatternFill(start_color="D5A6E6", end_color="D5A6E6", fill_type="solid")
redf = Font(bold=True, color="FFFFFF", size=10)
bdr = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
al = Alignment(wrap_text=True, vertical="top", horizontal="center")
al_left = Alignment(wrap_text=True, vertical="top")

# Series colors
S850 = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SBLOG = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
SCAR = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
S30Y = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
SOVER = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
SAARP = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
SINFL = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
SBIZ = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
SAGEN = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
SWIOA = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
sfont = Font(bold=True, color="FFFFFF", size=11)

headers = [
    "#", "Content Title", "Pri",
    "LinkedIn\nPost", "LI Post\nDate",
    "LinkedIn\nVideo", "LI Video\nDate",
    "LI News-\nletter", "LI NL\nDate",
    "LI\nCarousel", "LI Car\nDate",
    "WordPress\nBlog", "WP\nDate",
    "YouTube\nVideo", "YT Vid\nDate",
    "YouTube\nShort", "YT Short\nDate",
    "Podcast", "Pod\nDate",
    "Facebook", "FB\nDate",
]
widths = [5, 42, 6, 10, 11, 10, 11, 10, 11, 10, 11, 10, 11, 10, 11, 10, 11, 10, 11, 10, 11]

for c, (h, w) in enumerate(zip(headers, widths), 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = hf
    cell.fill = hfill
    cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
    cell.border = bdr
    ws.column_dimensions[get_column_letter(c)].width = w
ws.row_dimensions[1].height = 45

NC = len(headers)

def color_cell(cell, val):
    if isinstance(val, str):
        v = val.upper().strip()
        if v in ("POSTED", "LIVE", "PUBLISHED"): cell.fill = green
        elif v == "READY": cell.fill = blue
        elif v == "DRAFT": cell.fill = yellow
        elif v == "TO BUILD": cell.fill = orange
        elif v in ("TONIGHT", "POST NOW", "NOW"): cell.fill = red; cell.font = redf
        elif v == "REPURPOSE": cell.fill = purple
        elif v == "--": cell.fill = gray

def sumrow(ws, row, title, stats, sfill):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = sfont; cell.fill = sfill; cell.alignment = Alignment(horizontal="left", vertical="center"); cell.border = bdr
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=NC)
    c2 = ws.cell(row=row, column=4, value=stats)
    c2.font = Font(bold=True, color="FFFFFF", size=9); c2.fill = sfill; c2.border = bdr; c2.alignment = al
    ws.row_dimensions[row].height = 28
    return row + 1

def add(ws, row, data, rh=35):
    for c, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.border = bdr
        cell.alignment = al if c > 2 else al_left
        if c == 3:
            cell.alignment = al
            if isinstance(val, str):
                if val == "H": cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"); cell.font = Font(bold=True, color="FFFFFF", size=9)
                elif val == "M": cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif val == "L": cell.fill = gray
        elif c in (4,6,8,10,12,14,16,18,20): color_cell(cell, val)
    ws.row_dimensions[row].height = rh
    return row + 1

row = 2
groups = []

# =============== $850 BILLION SERIES ===============
row = sumrow(ws, row, "$850 BILLION VIDEO SERIES", "6 posted | 15 ready | 1 to build", S850)
gs = row
row = add(ws, row, ["1","Part 1 — Cost of Ignoring Workers","H","POSTED","Apr 5","READY","May 6","--","","--","","--","","POSTED","Apr 5","READY","May 23","READY","May 3","REPURPOSE",""])
row = add(ws, row, ["2","Part 2 — The Investment Gap","H","READY","May 5","READY","May 12","--","","TO BUILD","May 5","--","","POSTED","Apr 6","READY","May 2","READY","May 7","REPURPOSE",""])
row = add(ws, row, ["3","Part 3 — Winning Organizations","H","READY","May 19","READY","May 19","--","","TO BUILD","May 12","--","","POSTED","Apr 7","READY","May 28","READY","May 14","REPURPOSE",""])
row = add(ws, row, ["4","Part 4 — The Solution","H","READY","May 20","READY","May 22","--","","--","","--","","POSTED","Apr 13","READY","May 30","TO BUILD","","REPURPOSE",""])
row = add(ws, row, ["5","Bridge — It's About the People","H","READY","May 13","READY","May 15","--","","--","","PUBLISHED","Apr 29","READY","Upload","POSTED","Check","READY","May 21","REPURPOSE",""])
row = add(ws, row, ["6","$50 Billion Lost Short","M","READY","May 1","--","","--","","--","","--","","--","","POSTED","Apr 20","--","","REPURPOSE",""])
row = add(ws, row, ["7","AI Hiring Bias Text Post","H","TONIGHT","Apr 29","--","","--","","--","","--","","--","","--","","--","","--",""])
row = add(ws, row, ["8","Investment Gap Math Text","M","READY","May 5","--","","--","","--","","--","","--","","--","","--","","--",""])
row = add(ws, row, ["9","Flagship $850B Long Article","H","--","","--","","READY","May 8","--","","READY","May 9","--","","--","","--","","REPURPOSE",""])
groups.append((gs, row-1))
row += 1

# =============== 30 YEARS SERIES ===============
row = sumrow(ws, row, "30 YEARS THAT AI NEVER WILL (8 parts)", "All written | Podcast scripts ready | Starts May 6", S30Y)
gs = row
titles30 = [
    ["1","I Started at 65. Here's What They Don't Tell You","H"],
    ["2","The Agentic Advantage","H"],
    ["3","AI Is for Young Tech People? Wrong.","M"],
    ["4","Will AI Replace Me? Not If You're Over 50","M"],
    ["5","Too Old to Learn AI? You Can Write an Email.","M"],
    ["6","AI Is Just ChatGPT? It's a Business Team for $120/mo","M"],
    ["7","I Need an AI Course First? No. 7 Days + Problem.","M"],
    ["8","Start a Business After 50 — 7-Day AI Plan","H"],
]
for i, t in enumerate(titles30):
    dt = f"May {6+i*7}" if i < 4 else f"Jun {3+(i-4)*7}"
    row = add(ws, row, [t[0], t[1], t[2], "READY",dt,"REPURPOSE","","--","","--","","READY",dt,"REPURPOSE","","REPURPOSE","","READY",dt,"REPURPOSE",""])
groups.append((gs, row-1))
row += 1

# =============== OVERLOOKED WORKFORCE ===============
row = sumrow(ws, row, "THE OVERLOOKED WORKFORCE (8 LinkedIn shorts)", "All written | 2x/week Tue+Thu", SOVER)
gs = row
for i in range(1,9):
    d = f"May {5 + (i-1)*3}"
    row = add(ws, row, [str(i), f"Overlooked Workforce Post #{i}","M","READY",d,"--","","--","","--","","--","","--","","--","","--","","REPURPOSE",""])
groups.append((gs, row-1))
row += 1

# =============== AARP TRUTH ===============
row = sumrow(ws, row, "AARP TRUTH SERIES", "Video finished | Ready to publish", SAARP)
gs = row
row = add(ws, row, ["1","AARP Truth — Main Video","H","READY","May 10","READY","May 10","--","","--","","--","","READY","May 10","READY","May 10","READY","May 10","REPURPOSE",""])
groups.append((gs, row-1))
row += 1

# =============== INFLUENCE ===============
row = sumrow(ws, row, "INFLUENCE IS NOT WHO YOU KNOW", "Video finished | Launch May 24", SINFL)
gs = row
row = add(ws, row, ["1","Influence Is Not Who You Know — Full Video","H","READY","May 24","READY","May 24","--","","--","","--","","READY","May 24","READY","May 24","READY","May 24","REPURPOSE",""])
groups.append((gs, row-1))
row += 1

# =============== BUSINESS AFTER 50 ===============
row = sumrow(ws, row, "HOW TO START A BUSINESS AFTER 50", "Published", SBIZ)
gs = row
row = add(ws, row, ["1","Start a Business Online After 50","M","PUBLISHED","","PUBLISHED","","--","","--","","PUBLISHED","","PUBLISHED","","PUBLISHED","","--","","REPURPOSE",""])
groups.append((gs, row-1))
row += 1

# =============== AGENTIC AGE ===============
row = sumrow(ws, row, "AGENTIC AGE SERIES (2 parts)", "Written | Ready", SAGEN)
gs = row
row = add(ws, row, ["1","The Agentic Advantage: Building Business After 50","M","READY","May 15","--","","--","","--","","READY","May 15","--","","--","","--","","--",""])
row = add(ws, row, ["2","Agentic AI and Experience Advantage","M","READY","May 22","--","","--","","--","","READY","May 22","--","","--","","--","","--",""])
groups.append((gs, row-1))
row += 1

# =============== WIOA CLUSTER ===============
row = sumrow(ws, row, "WIOA / MBE ARTICLES CLUSTER (6 pieces)", "All written | SEO-optimized", SWIOA)
gs = row
wioa = [
    ["1","How WIOA Funds Can Pay for AI Training for Adults 50+","H","May 14"],
    ["2","WIOA Already Mandates Digital Skills Training","H","May 15"],
    ["3","How to Write a WIOA Training Proposal That Gets Funded","H","May 16"],
    ["4","What Is ETPL and How to Get Listed","H","May 19"],
    ["5","MBE Certification and Workforce Contracts","M","May 20"],
    ["6","MBE Certified Workforce Training Vendors in Texas","M","May 21"],
]
for w in wioa:
    row = add(ws, row, [w[0],w[1],w[2],"REPURPOSE","","--","","--","","--","","READY",w[3],"--","","--","","--","","--",""])
groups.append((gs, row-1))
row += 1

# =============== BLOG SERIES ===============
row = sumrow(ws, row, "WORDPRESS BLOG SERIES (Day 01-16)", "6 published | 10 drafts", SBLOG)
gs = row
blogs = [
    ["D01","Why Your 50+ Workforce Is Biggest Untapped Asset","M","PUBLISHED","Apr 21"],
    ["D02","The $850 Billion Cost of Ignoring Workers","H","PUBLISHED","Apr 23"],
    ["D03","The Confidence Gap: AI Training Fails Workers 50+","M","PUBLISHED","Apr 27"],
    ["D04","The Digital Divide for Adults Over 50","M","PUBLISHED","Apr 25"],
    ["D05","The $76 Trillion Blind Spot","M","PUBLISHED","Apr 25"],
    ["D06","It's Not About the Data. It Never Was.","H","PUBLISHED","Apr 29"],
    ["D07","What Is AgeTech? Complete Guide 2026","M","DRAFT","Apr 30"],
    ["D08","What 23 Organizations Learned","H","DRAFT","May 1"],
    ["D09","Pioneer Stories: Real Results","H","DRAFT","May 2"],
    ["D10","AI Training for Older Workers","H","DRAFT","May 5"],
    ["D11","Corporate AI Workshops: ROI Guide","H","DRAFT","May 6"],
    ["D12","Agentic AI and the 50+ Workforce","M","DRAFT","May 7"],
    ["D13","Why 50-Year-Old Founders Succeed 2.8X","M","DRAFT","May 8"],
    ["D14","Deploy AI Training This Quarter","H","DRAFT","May 9"],
    ["D15","I Started at 65","H","DRAFT","May 12"],
    ["D16","The Agentic Advantage for Pros Over 50","M","DRAFT","May 13"],
]
for b in blogs:
    row = add(ws, row, [b[0],b[1],b[2],"REPURPOSE","","--","","--","","--","",b[3],b[4],"--","","--","","--","","REPURPOSE",""])
groups.append((gs, row-1))
row += 1

# =============== CAROUSEL SERIES ===============
row = sumrow(ws, row, "CAROUSEL SERIES (8 pieces)", "1 ready | 7 to build in Canva", SCAR)
gs = row
cars = [
    ["C01","$850B Workforce Gap","H","READY","Apr 30"],
    ["C02","Investment Gap — 5 Stats","H","TO BUILD","May 5"],
    ["C03","What Winning Organizations Do","M","TO BUILD","May 12"],
    ["C04","347 Pioneers — Real Results","H","TO BUILD","May 19"],
    ["C05","The $76 Trillion Opportunity","M","TO BUILD","May 26"],
    ["C06","AI Is Not Replacing You","M","TO BUILD","Jun 4"],
    ["C07","I Started at 65","M","TO BUILD","Jun 9"],
    ["C08","Why 50-Year-Old Founders Win","L","TO BUILD","Jun 16"],
]
for ca in cars:
    row = add(ws, row, [ca[0],ca[1],ca[2],"--","","--","","--","",ca[3],ca[4],"--","","--","","--","","--","","--",""])
groups.append((gs, row-1))
row += 2

# =============== COLOR KEY ===============
sumrow(ws, row, "COLOR KEY", "", hfill)
row += 1
for label, fill, font in [
    ("GREEN = POSTED / PUBLISHED / LIVE", green, Font(size=10)),
    ("BLUE = READY to post (file exists)", blue, Font(size=10)),
    ("YELLOW = DRAFT in WordPress", yellow, Font(size=10)),
    ("ORANGE = TO BUILD (needs Canva/recording)", orange, Font(size=10)),
    ("RED = POST NOW / TONIGHT", red, redf),
    ("PURPLE = REPURPOSE from existing content", purple, Font(size=10)),
    ("GRAY = N/A for this content", gray, Font(size=10)),
]:
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=11)
    c = ws.cell(row=row, column=2, value=label); c.fill = fill; c.font = font; c.border = bdr
    ws.row_dimensions[row].height = 20; row += 1

row += 1
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=11)
ws.cell(row=row, column=2, value="SERIES COLORS (header bars)").font = Font(bold=True, size=11)
row += 1
for label, fill in [
    ("BLUE = $850 Billion Video Series", S850),
    ("RED = 30 Years That AI Never Will", S30Y),
    ("CYAN = The Overlooked Workforce", SOVER),
    ("ORANGE = AARP Truth Series", SAARP),
    ("PURPLE = Influence Is Not Who You Know", SINFL),
    ("GREEN = Business After 50", SBIZ),
    ("NAVY = Agentic Age Series", SAGEN),
    ("GRAY = WIOA/MBE Cluster", SWIOA),
    ("FOREST = WordPress Blog Series", SBLOG),
    ("GOLD = Carousel Series", SCAR),
]:
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=11)
    c = ws.cell(row=row, column=2, value=label); c.fill = fill; c.font = Font(bold=True, color="FFFFFF"); c.border = bdr
    ws.row_dimensions[row].height = 20; row += 1

# Grouping
ws.sheet_properties.outlinePr = openpyxl.worksheet.properties.Outline(summaryBelow=False, summaryRight=False)
for gs, ge in groups:
    for r in range(gs, ge+1):
        ws.row_dimensions[r].outlineLevel = 1

ws.freeze_panes = "D2"
fp = r"C:\Users\USER\Desktop\LMT\850-Billion-Series\MASTER-CONTENT-GRID.xlsx"
wb.save(fp)
print("Saved: " + fp)
