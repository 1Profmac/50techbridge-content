import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Content Grid"

# Styles
hf = Font(bold=True, color="FFFFFF", size=11)
hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
blue = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
orange = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
red = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
gray = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
redf = Font(bold=True, color="FFFFFF", size=10)
bdr = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
al = Alignment(wrap_text=True, vertical="top", horizontal="center")
al_left = Alignment(wrap_text=True, vertical="top")

# Series stripe colors
s850 = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
sblog = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
scarousel = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
secfont = Font(bold=True, size=12, color="FFFFFF")

# Platform headers in order of importance
headers = [
    "Series", "#", "Content Title",
    "WP Article\nStatus", "WP Article\nDate",
    "LinkedIn Post\nStatus", "LinkedIn Post\nDate",
    "YouTube Video\nStatus", "YouTube Video\nDate",
    "YouTube Short\nStatus", "YouTube Short\nDate",
    "Podcast\nStatus", "Podcast\nDate",
    "FB Post\nStatus", "FB Post\nDate"
]
widths = [12, 5, 45, 12, 13, 12, 13, 12, 13, 12, 13, 12, 13, 12, 13]

for c, (h, w) in enumerate(zip(headers, widths), 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = hf
    cell.fill = hfill
    cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
    cell.border = bdr
    ws.column_dimensions[get_column_letter(c)].width = w

ws.row_dimensions[1].height = 45

def color_cell(cell, val):
    if isinstance(val, str):
        v = val.upper()
        if v in ("POSTED", "LIVE", "PUBLISHED"):
            cell.fill = green
        elif v == "READY":
            cell.fill = blue
        elif v == "DRAFT":
            cell.fill = yellow
        elif v == "TO BUILD":
            cell.fill = orange
        elif v in ("TONIGHT", "POST NOW"):
            cell.fill = red
            cell.font = redf
        elif v == "--":
            cell.fill = gray
        else:
            cell.fill = white

def sec(ws, row, title, fill):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=15)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = secfont
    cell.fill = fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = bdr
    return row + 1

def add(ws, row, data, sfill, rh=40):
    # data = [series, num, title, wp_status, wp_date, li_status, li_date, yt_status, yt_date, short_status, short_date, pod_status, pod_date, fb_status, fb_date]
    for c, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.border = bdr
        cell.alignment = al if c > 3 else al_left
        if c == 1:
            cell.fill = sfill
            cell.font = Font(bold=True, color="FFFFFF", size=9)
            cell.alignment = al
        elif c in (4,6,8,10,12,14):  # status columns
            color_cell(cell, val)
        elif c in (5,7,9,11,13,15):  # date columns
            cell.fill = white
        else:
            cell.fill = white
    ws.row_dimensions[row].height = rh
    return row + 1

row = 2

# ===== $850 BILLION VIDEO SERIES =====
row = sec(ws, row, "$850 BILLION VIDEO SERIES", s850)

row = add(ws, row, [
    "$850B", "1", "Part 1 — Cost of Ignoring Experienced Workers",
    "--", "",
    "POSTED", "Sat Apr 5",
    "POSTED", "Sat Apr 5",
    "READY", "Fri May 23",
    "READY", "Sat May 3",
    "--", ""
], s850)

row = add(ws, row, [
    "$850B", "2", "Part 2 — The Investment Gap",
    "--", "",
    "READY", "Mon May 5",
    "POSTED", "Sun Apr 6",
    "READY", "Fri May 2",
    "READY", "Wed May 7",
    "--", ""
], s850)

row = add(ws, row, [
    "$850B", "3", "Part 3 — What Winning Organizations Do",
    "--", "",
    "READY", "Mon May 19",
    "POSTED", "Mon Apr 7",
    "READY", "Wed May 28",
    "READY", "Wed May 14",
    "--", ""
], s850)

row = add(ws, row, [
    "$850B", "4", "Part 4 — The Solution: 50+TechBridge",
    "--", "",
    "READY", "Tue May 20",
    "POSTED", "Sun Apr 13",
    "READY", "Fri May 30",
    "TO BUILD", "",
    "--", ""
], s850)

row = add(ws, row, [
    "$850B", "5", "Bridge — It's Not About the Data / About the People",
    "PUBLISHED", "Tue Apr 29",
    "READY", "Tue May 13",
    "READY", "Upload",
    "POSTED", "Check YT",
    "READY", "Wed May 21",
    "--", ""
], s850)

row = add(ws, row, [
    "$850B", "6", "$50 Billion Lost (Short)",
    "--", "",
    "READY", "Thu May 1",
    "--", "",
    "POSTED", "Mon Apr 20",
    "--", "",
    "--", ""
], s850)

row = add(ws, row, [
    "$850B", "7", "AI Hiring Bias — Text Post",
    "--", "",
    "TONIGHT", "Tue Apr 29",
    "--", "",
    "--", "",
    "--", "",
    "--", ""
], s850)

row = add(ws, row, [
    "$850B", "8", "Investment Gap Math — Text Post",
    "--", "",
    "READY", "Mon May 5",
    "--", "",
    "--", "",
    "--", "",
    "--", ""
], s850)

row = add(ws, row, [
    "$850B", "9", "Flagship — $850B Long Article",
    "READY", "Fri May 9",
    "READY", "Thu May 8",
    "--", "",
    "--", "",
    "--", "",
    "--", ""
], s850)

row += 1

# ===== WORDPRESS BLOG SERIES =====
row = sec(ws, row, "WORDPRESS BLOG SERIES (Day 01-16)", sblog)

blogs = [
    ["D01", "Why Your 50+ Workforce Is Your Biggest Untapped Asset", "PUBLISHED", "Mon Apr 21"],
    ["D02", "The $850 Billion Cost of Ignoring Your Experienced Workers", "PUBLISHED", "Wed Apr 23"],
    ["D03", "The Confidence Gap: Why AI Training Fails Workers Over 50+", "PUBLISHED", "Sun Apr 27"],
    ["D04", "The Digital Divide for Adults Over 50", "PUBLISHED", "Fri Apr 25"],
    ["D05", "The $76 Trillion Blind Spot: Why Marketers Ignore Adults 50+", "PUBLISHED", "Fri Apr 25"],
    ["D06", "It's Not About the Data. It Never Was.", "PUBLISHED", "Tue Apr 29"],
    ["D07", "What Is AgeTech? The Complete Guide for 2026", "DRAFT", "Wed Apr 30"],
    ["D08", "What 23 Organizations Learned Deploying 50+TechBridge", "DRAFT", "Thu May 1"],
    ["D09", "Pioneer Stories: Real Results from Real People", "DRAFT", "Fri May 2"],
    ["D10", "AI Training for Older Workers: What Employers Need to Know", "DRAFT", "Mon May 5"],
    ["D11", "Corporate AI Workshops: The ROI Guide", "DRAFT", "Tue May 6"],
    ["D12", "Agentic AI and the 50+ Workforce", "DRAFT", "Wed May 7"],
    ["D13", "Why 50-Year-Old Founders Are 2.8X More Likely to Succeed", "DRAFT", "Thu May 8"],
    ["D14", "How to Deploy AI Training This Quarter", "DRAFT", "Fri May 9"],
    ["D15", "I Started at 65. Here's What They Don't Tell You.", "DRAFT", "Mon May 12"],
    ["D16", "The Agentic Advantage: AI Agents for Pros Over 50", "DRAFT", "Tue May 13"],
]

for b in blogs:
    row = add(ws, row, [
        "Blog", b[0], b[1],
        b[2], b[3],
        "--", "",
        "--", "",
        "--", "",
        "--", "",
        "--", ""
    ], sblog)

row += 1

# ===== CAROUSEL SERIES =====
row = sec(ws, row, "CAROUSEL SERIES (LinkedIn)", scarousel)

carousels = [
    ["C01", "$850B Workforce Gap", "READY", "Wed Apr 30"],
    ["C02", "Investment Gap — 5 Stats", "TO BUILD", "Mon May 5"],
    ["C03", "What Winning Organizations Do", "TO BUILD", "Mon May 12"],
    ["C04", "347 Pioneers — Real Results", "TO BUILD", "Mon May 19"],
    ["C05", "The $76 Trillion Opportunity", "TO BUILD", "Mon May 26"],
    ["C06", "AI Is Not Replacing You", "TO BUILD", "Wed Jun 4"],
    ["C07", "I Started at 65", "TO BUILD", "Mon Jun 9"],
    ["C08", "Why 50-Year-Old Founders Win", "TO BUILD", "Mon Jun 16"],
]

for ca in carousels:
    row = add(ws, row, [
        "Carousel", ca[0], ca[1],
        "--", "",
        ca[2], ca[3],
        "--", "",
        "--", "",
        "--", "",
        "--", ""
    ], scarousel)

row += 2

# ===== COLOR KEY =====
row = sec(ws, row, "COLOR KEY", hfill)
legend = [
    ("POSTED / PUBLISHED / LIVE = Done, it's out there", green, Font()),
    ("READY = File exists, date is set, just post it", blue, Font()),
    ("DRAFT = In WordPress drafts, hit Publish", yellow, Font()),
    ("TO BUILD = Needs to be created (Canva, record, etc.)", orange, Font()),
    ("TONIGHT / POST NOW = Do it right now", red, redf),
    ("-- = Not applicable for this content piece", gray, Font()),
]
for label, fill, font in legend:
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=9)
    cell = ws.cell(row=row, column=3, value=label)
    cell.fill = fill
    cell.font = font
    cell.border = bdr
    cell.alignment = al_left
    ws.row_dimensions[row].height = 22
    row += 1

row += 1
# Series key
ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=9)
ws.cell(row=row, column=3, value="SERIES (left column color):").font = Font(bold=True, size=11)
row += 1
for label, fill in [("BLUE = $850 Billion Video Series", s850), ("GREEN = WordPress Blog Series", sblog), ("GOLD = Carousel Series", scarousel)]:
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=9)
    cell = ws.cell(row=row, column=3, value=label)
    cell.fill = fill
    cell.font = Font(bold=True, color="FFFFFF")
    cell.border = bdr
    ws.row_dimensions[row].height = 22
    row += 1

ws.freeze_panes = "D2"

fp = r"C:\Users\USER\Desktop\LMT\850-Billion-Series\MASTER-CONTENT-GRID.xlsx"
wb.save(fp)
print("Saved: " + fp)
