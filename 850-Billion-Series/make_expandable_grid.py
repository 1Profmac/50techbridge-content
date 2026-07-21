import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Content Command Center"

# ============ STYLES ============
hf = Font(bold=True, color="FFFFFF", size=11)
hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
blue = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
orange = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
red = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
gray = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
purple = PatternFill(start_color="D5A6E6", end_color="D5A6E6", fill_type="solid")
redf = Font(bold=True, color="FFFFFF", size=10)
bdr = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
al = Alignment(wrap_text=True, vertical="top", horizontal="center")
al_left = Alignment(wrap_text=True, vertical="top")

# Series colors
s850_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
sblog_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
scar_fill = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
sstand_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
sfont = Font(bold=True, color="FFFFFF", size=12)
sfont_sm = Font(bold=True, color="FFFFFF", size=9)

# Summary row
sum_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
sum_font = Font(bold=True, color="FFFFFF", size=11)

# Headers: Platform order by relevance
# LinkedIn Post, LinkedIn Video, LinkedIn Newsletter, LinkedIn Carousel, WP Blog, YouTube Video, YouTube Short, Podcast, FB Post, Email Newsletter
headers = [
    "#", "Content Title", "Priority",
    "LinkedIn\nPost", "LinkedIn Post\nDate",
    "LinkedIn\nVideo", "LinkedIn Video\nDate",
    "LinkedIn\nNewsletter", "LI Newsletter\nDate",
    "LinkedIn\nCarousel", "LI Carousel\nDate",
    "WordPress\nBlog", "WP Blog\nDate",
    "YouTube\nVideo", "YouTube Video\nDate",
    "YouTube\nShort", "YT Short\nDate",
    "Buzzsprout\nPodcast", "Podcast\nDate",
    "Facebook\nPost", "FB Post\nDate",
    "Email\nNewsletter", "Email\nDate"
]
widths = [5, 44, 9, 11, 12, 11, 12, 11, 12, 11, 12, 11, 12, 11, 12, 11, 12, 11, 12, 11, 12, 11, 12]

for c, (h, w) in enumerate(zip(headers, widths), 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = hf
    cell.fill = hfill
    cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
    cell.border = bdr
    ws.column_dimensions[get_column_letter(c)].width = w
ws.row_dimensions[1].height = 50

def color_cell(cell, val):
    if isinstance(val, str):
        v = val.upper().strip()
        if v in ("POSTED", "LIVE", "PUBLISHED"):
            cell.fill = green
        elif v == "READY":
            cell.fill = blue
        elif v == "DRAFT":
            cell.fill = yellow
        elif v == "TO BUILD":
            cell.fill = orange
        elif v in ("TONIGHT", "POST NOW", "NOW"):
            cell.fill = red
            cell.font = redf
        elif v == "REPURPOSE":
            cell.fill = purple
        elif v == "--":
            cell.fill = gray
        else:
            cell.fill = white

def summary_row(ws, row, title, total, posted, ready, draft, tobuild, sfill):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = sfont
    cell.fill = sfill
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = bdr
    stats = f"{posted} posted | {ready} ready | {draft} draft | {tobuild} to build"
    cell2 = ws.cell(row=row, column=3, value=stats)
    cell2.font = Font(bold=True, color="FFFFFF", size=9)
    cell2.fill = sfill
    cell2.border = bdr
    cell2.alignment = Alignment(horizontal="center", wrap_text=True)
    for c in range(4, 24):
        ws.cell(row=row, column=c).fill = sfill
        ws.cell(row=row, column=c).border = bdr
    ws.row_dimensions[row].height = 30
    return row + 1

def add(ws, row, data, rh=38):
    for c, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.border = bdr
        if c <= 2:
            cell.alignment = al_left
        elif c == 3:
            cell.alignment = al
            # Priority coloring
            if isinstance(val, str):
                if val.upper() == "HIGH":
                    cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                elif val.upper() == "MED":
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif val.upper() == "LOW":
                    cell.fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        elif c in (4,6,8,10,12,14,16,18,20,22):  # status cols
            cell.alignment = al
            color_cell(cell, val)
        else:  # date cols
            cell.alignment = al
            cell.fill = white
    ws.row_dimensions[row].height = rh
    return row + 1

row = 2

# ===============================================
# $850 BILLION VIDEO SERIES
# ===============================================
row = summary_row(ws, row, "$850 BILLION VIDEO SERIES (9 pieces)", 9, 6, 15, 0, 1, s850_fill)
group_start_1 = row

# data: #, title, priority, li_post_s, li_post_d, li_vid_s, li_vid_d, li_news_s, li_news_d, li_car_s, li_car_d, wp_s, wp_d, yt_s, yt_d, short_s, short_d, pod_s, pod_d, fb_s, fb_d, email_s, email_d

row = add(ws, row, [
    "1", "Part 1 — Cost of Ignoring Experienced Workers", "HIGH",
    "POSTED", "Sat Apr 5", "READY", "Tue May 6", "--", "", "--", "",
    "--", "", "POSTED", "Sat Apr 5", "READY", "Fri May 23",
    "READY", "Sat May 3", "REPURPOSE", "", "--", ""
])
row = add(ws, row, [
    "2", "Part 2 — The Investment Gap", "HIGH",
    "READY", "Mon May 5", "READY", "Mon May 12", "--", "", "TO BUILD", "Mon May 5",
    "--", "", "POSTED", "Sun Apr 6", "READY", "Fri May 2",
    "READY", "Wed May 7", "REPURPOSE", "", "--", ""
])
row = add(ws, row, [
    "3", "Part 3 — What Winning Organizations Do", "HIGH",
    "READY", "Mon May 19", "READY", "Mon May 19", "--", "", "TO BUILD", "Mon May 12",
    "--", "", "POSTED", "Mon Apr 7", "READY", "Wed May 28",
    "READY", "Wed May 14", "REPURPOSE", "", "--", ""
])
row = add(ws, row, [
    "4", "Part 4 — The Solution: 50+TechBridge", "HIGH",
    "READY", "Tue May 20", "READY", "Thu May 22", "--", "", "--", "",
    "--", "", "POSTED", "Sun Apr 13", "READY", "Fri May 30",
    "TO BUILD", "", "REPURPOSE", "", "--", ""
])
row = add(ws, row, [
    "5", "Bridge — It's About the People", "HIGH",
    "READY", "Tue May 13", "READY", "Thu May 15", "--", "", "--", "",
    "PUBLISHED", "Tue Apr 29", "READY", "Upload", "POSTED", "Check YT",
    "READY", "Wed May 21", "REPURPOSE", "", "--", ""
])
row = add(ws, row, [
    "6", "$50 Billion Lost (Short)", "MED",
    "READY", "Thu May 1", "--", "", "--", "", "--", "",
    "--", "", "--", "", "POSTED", "Mon Apr 20",
    "--", "", "REPURPOSE", "", "--", ""
])
row = add(ws, row, [
    "7", "AI Hiring Bias — Text Post", "HIGH",
    "TONIGHT", "Tue Apr 29", "--", "", "--", "", "--", "",
    "--", "", "--", "", "--", "",
    "--", "", "--", "", "--", ""
])
row = add(ws, row, [
    "8", "Investment Gap Math — Text Post", "MED",
    "READY", "Mon May 5", "--", "", "--", "", "--", "",
    "--", "", "--", "", "--", "",
    "--", "", "--", "", "--", ""
])
row = add(ws, row, [
    "9", "Flagship — $850B Long Article", "HIGH",
    "--", "", "--", "", "READY", "Thu May 8", "--", "",
    "READY", "Fri May 9", "--", "", "--", "",
    "--", "", "REPURPOSE", "", "REPURPOSE", ""
])

group_end_1 = row - 1

row += 1

# ===============================================
# WORDPRESS BLOG SERIES
# ===============================================
row = summary_row(ws, row, "WORDPRESS BLOG SERIES (16 pieces)", 16, 6, 0, 10, 0, sblog_fill)
group_start_2 = row

blogs = [
    ["D01", "Why Your 50+ Workforce Is Your Biggest Untapped Asset", "MED", "PUBLISHED", "Mon Apr 21"],
    ["D02", "The $850 Billion Cost of Ignoring Your Experienced Workers", "HIGH", "PUBLISHED", "Wed Apr 23"],
    ["D03", "The Confidence Gap: Why AI Training Fails Workers Over 50+", "MED", "PUBLISHED", "Sun Apr 27"],
    ["D04", "The Digital Divide for Adults Over 50", "MED", "PUBLISHED", "Fri Apr 25"],
    ["D05", "The $76 Trillion Blind Spot: Why Marketers Ignore Adults 50+", "MED", "PUBLISHED", "Fri Apr 25"],
    ["D06", "It's Not About the Data. It Never Was.", "HIGH", "PUBLISHED", "Tue Apr 29"],
    ["D07", "What Is AgeTech? The Complete Guide for 2026", "MED", "DRAFT", "Wed Apr 30"],
    ["D08", "What 23 Organizations Learned Deploying 50+TechBridge", "HIGH", "DRAFT", "Thu May 1"],
    ["D09", "Pioneer Stories: Real Results from Real People", "HIGH", "DRAFT", "Fri May 2"],
    ["D10", "AI Training for Older Workers: What Employers Need to Know", "HIGH", "DRAFT", "Mon May 5"],
    ["D11", "Corporate AI Workshops: The ROI Guide", "HIGH", "DRAFT", "Tue May 6"],
    ["D12", "Agentic AI and the 50+ Workforce", "MED", "DRAFT", "Wed May 7"],
    ["D13", "Why 50-Year-Old Founders Are 2.8X More Likely to Succeed", "MED", "DRAFT", "Thu May 8"],
    ["D14", "How to Deploy AI Training This Quarter", "HIGH", "DRAFT", "Fri May 9"],
    ["D15", "I Started at 65. Here's What They Don't Tell You.", "HIGH", "DRAFT", "Mon May 12"],
    ["D16", "The Agentic Advantage: AI Agents for Pros Over 50", "MED", "DRAFT", "Tue May 13"],
]

for b in blogs:
    row = add(ws, row, [
        b[0], b[1], b[2],
        "REPURPOSE", "", "--", "", "--", "", "--", "",
        b[3], b[4], "--", "", "--", "",
        "--", "", "REPURPOSE", "", "REPURPOSE", ""
    ])

group_end_2 = row - 1

row += 1

# ===============================================
# CAROUSEL SERIES
# ===============================================
row = summary_row(ws, row, "CAROUSEL SERIES (8 pieces)", 8, 0, 1, 0, 7, scar_fill)
group_start_3 = row

carousels = [
    ["C01", "$850B Workforce Gap", "HIGH", "READY", "Wed Apr 30"],
    ["C02", "Investment Gap — 5 Stats", "HIGH", "TO BUILD", "Mon May 5"],
    ["C03", "What Winning Organizations Do", "MED", "TO BUILD", "Mon May 12"],
    ["C04", "347 Pioneers — Real Results", "HIGH", "TO BUILD", "Mon May 19"],
    ["C05", "The $76 Trillion Opportunity", "MED", "TO BUILD", "Mon May 26"],
    ["C06", "AI Is Not Replacing You", "MED", "TO BUILD", "Wed Jun 4"],
    ["C07", "I Started at 65", "MED", "TO BUILD", "Mon Jun 9"],
    ["C08", "Why 50-Year-Old Founders Win", "LOW", "TO BUILD", "Mon Jun 16"],
]

for ca in carousels:
    row = add(ws, row, [
        ca[0], ca[1], ca[2],
        "--", "", "--", "", "--", "", ca[3], ca[4],
        "--", "", "--", "", "--", "",
        "--", "", "--", "", "--", ""
    ])

group_end_3 = row - 1

row += 1

# ===============================================
# STANDALONE / ONE-OFF CONTENT
# ===============================================
row = summary_row(ws, row, "STANDALONE CONTENT (future)", 0, 0, 0, 0, 0, sstand_fill)
group_start_4 = row

row = add(ws, row, [
    "S01", "LinkedIn Newsletter — Launch \"The Workforce Math\"", "HIGH",
    "--", "", "--", "", "TO BUILD", "May 2026", "--", "",
    "--", "", "--", "", "--", "",
    "--", "", "--", "", "--", ""
])
row = add(ws, row, [
    "S02", "Workshop Proof Reel (90 sec sizzle)", "HIGH",
    "REPURPOSE", "", "TO BUILD", "", "--", "", "--", "",
    "--", "", "TO BUILD", "", "TO BUILD", "",
    "--", "", "REPURPOSE", "", "--", ""
])
row = add(ws, row, [
    "S03", "Pioneer Testimonial Clips (3-5 clips)", "MED",
    "REPURPOSE", "", "TO BUILD", "", "--", "", "--", "",
    "--", "", "--", "", "TO BUILD", "",
    "--", "", "REPURPOSE", "", "--", ""
])
row = add(ws, row, [
    "S04", "Weekly LinkedIn Live — TBD format", "LOW",
    "--", "", "TO BUILD", "", "--", "", "--", "",
    "--", "", "--", "", "--", "",
    "--", "", "REPURPOSE", "", "--", ""
])

group_end_4 = row - 1

row += 2

# ===============================================
# COLOR KEY
# ===============================================
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
cell = ws.cell(row=row, column=2, value="COLOR KEY — Status")
cell.font = Font(bold=True, size=12, color="1F4E79")
cell.border = bdr
row += 1

legend = [
    ("POSTED / PUBLISHED / LIVE", green, Font(size=10)),
    ("READY — file exists, date set, just post it", blue, Font(size=10)),
    ("DRAFT — in WordPress, hit Publish", yellow, Font(size=10)),
    ("TO BUILD — needs Canva, recording, or production", orange, Font(size=10)),
    ("TONIGHT / POST NOW — do it right now", red, redf),
    ("REPURPOSE — can be created from existing content", purple, Font(size=10)),
    ("-- = not applicable", gray, Font(size=10)),
]
for label, fill, font in legend:
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=10)
    cell = ws.cell(row=row, column=3, value=label)
    cell.fill = fill
    cell.font = font
    cell.border = bdr
    cell.alignment = al_left
    ws.row_dimensions[row].height = 22
    row += 1

row += 1
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
cell = ws.cell(row=row, column=2, value="COLOR KEY — Priority")
cell.font = Font(bold=True, size=12, color="1F4E79")
cell.border = bdr
row += 1

for label, fill, font in [
    ("HIGH — revenue-driving or audience-building", PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"), Font(bold=True, color="FFFFFF")),
    ("MED — supports the brand, schedule when time allows", PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"), Font()),
    ("LOW — nice to have, park for later", PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid"), Font()),
]:
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=10)
    cell = ws.cell(row=row, column=3, value=label)
    cell.fill = fill
    cell.font = font
    cell.border = bdr
    cell.alignment = al_left
    ws.row_dimensions[row].height = 22
    row += 1

row += 1
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
cell = ws.cell(row=row, column=2, value="SERIES (header bar color)")
cell.font = Font(bold=True, size=12, color="1F4E79")
cell.border = bdr
row += 1

for label, fill in [
    ("BLUE — $850 Billion Video Series", s850_fill),
    ("GREEN — WordPress Blog Series", sblog_fill),
    ("GOLD — Carousel Series", scar_fill),
    ("PURPLE — Standalone / One-Off Content", sstand_fill),
]:
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=10)
    cell = ws.cell(row=row, column=3, value=label)
    cell.fill = fill
    cell.font = Font(bold=True, color="FFFFFF")
    cell.border = bdr
    ws.row_dimensions[row].height = 22
    row += 1

# ============ GROUPING (expandable +/- buttons) ============
ws.sheet_properties.outlinePr = openpyxl.worksheet.properties.Outline(summaryBelow=False, summaryRight=False)

for r in range(group_start_1, group_end_1 + 1):
    ws.row_dimensions[r].outlineLevel = 1
for r in range(group_start_2, group_end_2 + 1):
    ws.row_dimensions[r].outlineLevel = 1
for r in range(group_start_3, group_end_3 + 1):
    ws.row_dimensions[r].outlineLevel = 1
for r in range(group_start_4, group_end_4 + 1):
    ws.row_dimensions[r].outlineLevel = 1

ws.freeze_panes = "D2"

fp = r"C:\Users\USER\Desktop\LMT\850-Billion-Series\MASTER-CONTENT-GRID.xlsx"
wb.save(fp)
print("Saved: " + fp)
