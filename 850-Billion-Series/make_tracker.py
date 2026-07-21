import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============================================================
# SHEET 1: MAIN CONTENT TRACKER
# ============================================================
ws = wb.active
ws.title = "All Content"

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
pub_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
draft_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ready_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
hot_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
hot_font = Font(bold=True, color="FFFFFF")
sec_font = Font(bold=True, size=12, color="1F4E79")
sec_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
bdr = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

headers = ["Day #", "Title", "Platform", "Type", "Status", "Date Published", "Post Day", "File Path", "Next Action"]
widths = [8, 65, 15, 15, 12, 20, 15, 55, 25]

for c, (h, w) in enumerate(zip(headers, widths), 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = bdr
    ws.column_dimensions[get_column_letter(c)].width = w

def section(ws, row, title):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = sec_font
    cell.fill = sec_fill
    cell.alignment = Alignment(horizontal="center")
    return row + 1

def add_rows(ws, row, data, fill, highlight_action=None):
    for d in data:
        is_hot = highlight_action and d[8] == highlight_action
        for c, v in enumerate(d, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.fill = hot_fill if is_hot else fill
            cell.font = hot_font if is_hot else Font()
            cell.border = bdr
            cell.alignment = Alignment(wrap_text=True)
        row += 1
    return row

row = 2
row = section(ws, row, "PUBLISHED")
published = [
    ["Day 01", "Why Your 50+ Workforce Is Your Biggest Untapped Asset", "WordPress", "Blog", "PUBLISHED", "Mon Apr 21", "", "", "Live"],
    ["Day 02", "The $850 Billion Cost of Ignoring Your Experienced Workers", "WordPress", "Blog", "PUBLISHED", "Wed Apr 23", "", "", "Live"],
    ["Day 03", "The Confidence Gap: Why AI Training Fails Workers Over 50+", "WordPress", "Blog", "PUBLISHED", "Sun Apr 27", "", "", "Live"],
    ["Day 04", "The Digital Divide for Adults Over 50: Data, Causes, and Solutions", "WordPress", "Blog", "PUBLISHED", "Fri Apr 25", "", "", "Live"],
    ["Day 05", "The $76 Trillion Blind Spot: Why Marketers Ignore Adults 50+", "WordPress", "Blog", "PUBLISHED", "Fri Apr 25", "", "", "Live"],
    ["--", "Part 1 Cost of Ignoring Workers (Video)", "YouTube", "Video", "POSTED", "Sat Apr 5", "", "", "Live"],
    ["--", "Part 1 LinkedIn Post", "LinkedIn", "Text Post", "POSTED", "Sat Apr 5", "", "", "Live"],
    ["--", "Part 2 Investment Gap (Video)", "YouTube", "Video", "POSTED", "Sun Apr 6", "", "", "Live"],
    ["--", "Part 3 Real People Stories (Video)", "YouTube", "Video", "POSTED", "Mon Apr 7", "", "", "Live"],
    ["--", "Part 4 The Solution (Video)", "YouTube", "Video", "POSTED", "Sun Apr 13", "", "", "Live"],
    ["--", "$50 Billion Lost Short", "YouTube Shorts", "Short Video", "POSTED", "Mon Apr 20", "", "", "Live"],
]
row = add_rows(ws, row, published, pub_fill)
row += 1

row = section(ws, row, "WORDPRESS DRAFTS - Ready to Publish")
drafts = [
    ["Day 06", "Its Not About the Data. It Never Was.", "WordPress", "Blog", "DRAFT", "", "Tue Apr 29", "", "PUBLISH TONIGHT"],
    ["Day 07", "What Is AgeTech? The Complete Guide for 2026", "WordPress", "Blog", "DRAFT", "", "Wed Apr 30", "", ""],
    ["Day 08", "What 23 Organizations Learned Deploying 50+TechBridge", "WordPress", "Blog", "DRAFT", "", "Thu May 1", "", ""],
    ["Day 09", "50+TechBridge Pioneer Stories: Real Results from Real People", "WordPress", "Blog", "DRAFT", "", "Fri May 2", "", ""],
    ["Day 10", "AI Training for Older Workers: What Employers Need to Know", "WordPress", "Blog", "DRAFT", "", "Mon May 5", "", ""],
    ["Day 11", "Corporate AI Workshops for Experienced Employees: The ROI Guide", "WordPress", "Blog", "DRAFT", "", "Tue May 6", "", ""],
    ["Day 12", "Agentic AI and the 50+ Workforce: Why Experience Is the Advantage", "WordPress", "Blog", "DRAFT", "", "Wed May 7", "", ""],
    ["Day 13", "Why 50-Year-Old Founders Are 2.8X More Likely to Succeed", "WordPress", "Blog", "DRAFT", "", "Thu May 8", "", ""],
    ["Day 14", "How to Deploy AI Training for Your 50+ Workforce This Quarter", "WordPress", "Blog", "DRAFT", "", "Fri May 9", "", ""],
    ["Day 15", "I Started at 65. Heres What They Dont Tell You.", "WordPress", "Blog", "DRAFT", "", "Mon May 12", "", ""],
    ["Day 16", "The Agentic Advantage: Why AI Agents Are the Biggest Career Opportunity for Pros Over 50", "WordPress", "Blog", "DRAFT", "", "Tue May 13", "", ""],
]
row = add_rows(ws, row, drafts, draft_fill, "PUBLISH TONIGHT")
row += 1

row = section(ws, row, "UNPUBLISHED - LinkedIn, YouTube Shorts, Buzzsprout (Ready to Post)")
unpub = [
    ["--", "LinkedIn Text - AI Hiring Bias", "LinkedIn", "Text Post", "READY", "", "Tue Apr 29", "LINKEDIN-POSTS-WEEK-APR28.md", "POST TONIGHT"],
    ["--", "Carousel 850B", "LinkedIn", "Carousel", "READY", "", "Wed Apr 30", "CAROSUALS/50plus-techbridge-carousel-850B.pdf", "Upload with hook"],
    ["--", "$50 Billion Lost native video", "LinkedIn", "Native Video", "READY", "", "Thu May 1", "Shorts/$50 Billion Lost/$850 Billion lost.mp4", "Upload MP4 to LinkedIn"],
    ["--", "Part 2 Short", "YouTube Shorts", "Short Video", "READY", "", "Fri May 2", "Part-2/Youtube/Shorts/PART-2-SHORT-FINISHED.mp4", "Upload as Short"],
    ["--", "Podcast EP 1 The Number", "Buzzsprout", "Podcast", "READY", "", "Sat May 3", "podcast/EP-850B-Part1-The-Number.mp3", "Publish episode"],
    ["--", "LinkedIn Text Investment Gap", "LinkedIn", "Text Post", "READY", "", "Mon May 5", "LINKEDIN-POSTS-WEEK-APR28.md", "Copy/paste post"],
    ["--", "Part 1 Square Video", "LinkedIn", "Native Video", "READY", "", "Tue May 6", "Part-1/ALL-FORMATS/PART-1-FINISHED-SQUARE-1080x1080.mp4", "Upload MP4"],
    ["--", "Podcast EP 2 Investment Gap", "Buzzsprout", "Podcast", "READY", "", "Wed May 7", "podcast/EP-850B-Part2-Investment-Gap.mp3", "Publish episode"],
    ["--", "LinkedIn Newsletter Flagship", "LinkedIn", "Newsletter", "READY", "", "Thu May 8", "seomachine/articles/09-The-850-Billion-Cost-FLAGSHIP.md", "Publish as Newsletter"],
    ["--", "Blog Flagship Article", "WordPress", "Blog", "READY", "", "Fri May 9", "seomachine/articles/09-The-850-Billion-Cost-FLAGSHIP.md", "wp-publish.py"],
    ["--", "Part 2 Square Video", "LinkedIn", "Native Video", "READY", "", "Mon May 12", "Part-2/ALL-FORMATS/PART-2-FINISHED-SQUARE-1080x1080.mp4", "Upload MP4"],
    ["--", "Bridge LinkedIn Post", "LinkedIn", "Text Post", "READY", "", "Tue May 13", "Bridge/LINKEDIN/LINKEDIN-POST.txt", "Best post personal"],
    ["--", "Podcast EP 3 Real People", "Buzzsprout", "Podcast", "READY", "", "Wed May 14", "podcast/EP-850B-Part3-Real-People.mp3", "Publish episode"],
    ["--", "Bridge Square Video", "LinkedIn", "Native Video", "READY", "", "Thu May 15", "Bridge/ALL-FORMATS/Its-About-the-People-FINISHED-SQUARE.mp4", "Upload MP4"],
    ["--", "Carousel Repost", "LinkedIn", "Carousel", "READY", "", "Fri May 16", "CAROSUALS/50plus-techbridge-carousel-850B.pdf", "Different hook"],
    ["--", "Part 3 Square Video", "LinkedIn", "Native Video", "READY", "", "Mon May 19", "Part-3/ALL-FORMATS/PART-3-FINISHED-SQUARE-1080x1080.mp4", "Upload MP4"],
    ["--", "Part 4 LinkedIn Post", "LinkedIn", "Text Post", "READY", "", "Tue May 20", "Part-4/LINKEDIN/PART-4-LINKEDIN-POST.txt", "Rewrite hook"],
    ["--", "Podcast EP Bridge", "Buzzsprout", "Podcast", "READY", "", "Wed May 21", "podcast/EP-850B-Bridge-About-the-People.mp3", "Last episode"],
    ["--", "Part 4 Square Video", "LinkedIn", "Native Video", "READY", "", "Thu May 22", "Part-4/ALL-FORMATS/PART-4-FINISHED-SQUARE-1080x1080.mp4", "Upload MP4"],
    ["--", "Part 1 Vertical Short", "YouTube Shorts", "Short Video", "READY", "", "Fri May 23", "Part-1/ALL-FORMATS/PART-1-FINISHED-VERTICAL-1080x1920.mp4", "Upload as Short"],
    ["--", "Part 2 Vertical Short", "YouTube Shorts", "Short Video", "READY", "", "Mon May 26", "Part-2/ALL-FORMATS/PART-2-FINISHED-VERTICAL-1080x1920.mp4", "Upload as Short"],
    ["--", "Part 3 Vertical Short", "YouTube Shorts", "Short Video", "READY", "", "Wed May 28", "Part-3/ALL-FORMATS/PART-3-FINISHED-VERTICAL-1080x1920.mp4", "Upload as Short"],
    ["--", "Part 4 Vertical Short", "YouTube Shorts", "Short Video", "READY", "", "Fri May 30", "Part-4/ALL-FORMATS/PART-4-FINISHED-VERTICAL-1080x1920.mp4", "Upload as Short"],
    ["--", "Bridge YouTube Upload", "YouTube", "Video", "READY", "", "", "Bridge/ALL-FORMATS/Its-About-the-People-FINISHED-LANDSCAPE.mp4", "Upload to channel"],
    ["--", "Part 4 Podcast", "Buzzsprout", "Podcast", "NOT CREATED", "", "", "", "Needs to be produced"],
]
row = add_rows(ws, row, unpub, ready_fill, "POST TONIGHT")

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:I{row-1}"

# ============================================================
# SHEET 2: CAROUSELS
# ============================================================
ws2 = wb.create_sheet("Carousels")

car_headers = ["#", "Carousel Name", "Topic", "Status", "Date Created", "Date Posted", "Platform", "File Path", "Hook Text Used", "Notes"]
car_widths = [5, 45, 30, 12, 15, 15, 12, 50, 40, 30]

# Colors
car_done_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
car_ready_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
car_idea_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
orange_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")

for c, (h, w) in enumerate(zip(car_headers, car_widths), 1):
    cell = ws2.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = bdr
    ws2.column_dimensions[get_column_letter(c)].width = w

row2 = 2

# Section: POSTED
ws2.merge_cells(start_row=row2, start_column=1, end_row=row2, end_column=10)
cell = ws2.cell(row=row2, column=1, value="POSTED")
cell.font = sec_font
cell.fill = sec_fill
cell.alignment = Alignment(horizontal="center")
row2 += 1

# No carousels posted yet
for c in range(1, 11):
    cell = ws2.cell(row=row2, column=c, value="(none posted yet)" if c == 2 else "")
    cell.fill = pub_fill
    cell.border = bdr
row2 += 2

# Section: READY TO POST
ws2.merge_cells(start_row=row2, start_column=1, end_row=row2, end_column=10)
cell = ws2.cell(row=row2, column=1, value="READY TO POST")
cell.font = sec_font
cell.fill = sec_fill
cell.alignment = Alignment(horizontal="center")
row2 += 1

ready_carousels = [
    ["1", "50+TechBridge $850 Billion Carousel", "$850B Workforce Gap", "READY", "Mon Apr 28", "", "LinkedIn", "CAROSUALS/50plus-techbridge-carousel-850B.pdf", "$850 billion. Not a typo...", "First post Wed Apr 30 then repost Fri May 16 with different hook"],
]
for d in ready_carousels:
    for c, v in enumerate(d, 1):
        cell = ws2.cell(row=row2, column=c, value=v)
        cell.fill = car_ready_fill
        cell.border = bdr
        cell.alignment = Alignment(wrap_text=True)
    row2 += 1
row2 += 1

# Section: IDEAS - Future Carousels to Build
ws2.merge_cells(start_row=row2, start_column=1, end_row=row2, end_column=10)
cell = ws2.cell(row=row2, column=1, value="IDEAS - Future Carousels to Build")
cell.font = sec_font
cell.fill = sec_fill
cell.alignment = Alignment(horizontal="center")
row2 += 1

ideas = [
    ["2", "The Investment Gap — 5 Stats That Change Everything", "Part 2 data points", "TO BUILD", "", "", "LinkedIn", "", "", "Pull stats from Part 2 script"],
    ["3", "What Winning Organizations Do Differently", "Part 3 case studies", "TO BUILD", "", "", "LinkedIn", "", "", "BMW, CVS, United Health, Cisco examples"],
    ["4", "347 Pioneers — Real Results", "Pioneer outcomes", "TO BUILD", "", "", "LinkedIn", "", "", "3X completion, 74% confidence, 23 orgs"],
    ["5", "The $76 Trillion Opportunity", "Market size for 50+", "TO BUILD", "", "", "LinkedIn", "", "", "70% disposable income, $7.6T spending"],
    ["6", "AI Is Not Replacing You — Its Waiting For You", "AI for 50+ adults", "TO BUILD", "", "", "LinkedIn", "", "", "Tie to Agentic AI article Day 12"],
    ["7", "I Started at 65 — What They Dont Tell You", "Personal story", "TO BUILD", "", "", "LinkedIn", "", "", "Tie to blog Day 15"],
    ["8", "Why 50-Year-Old Founders Win", "Founder stats", "TO BUILD", "", "", "LinkedIn", "", "", "2.8X more likely to succeed — Day 13"],
]
for d in ideas:
    for c, v in enumerate(d, 1):
        cell = ws2.cell(row=row2, column=c, value=v)
        cell.fill = orange_fill
        cell.border = bdr
        cell.alignment = Alignment(wrap_text=True)
    row2 += 1

ws2.freeze_panes = "A2"
ws2.auto_filter.ref = f"A1:J{row2-1}"

fp = r"C:\Users\USER\Desktop\LMT\850-Billion-Series\MASTER-CONTENT-TRACKER.xlsx"
wb.save(fp)
print("Saved: " + fp)
