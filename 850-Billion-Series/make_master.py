import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Master Content Grid"

# Styles
hf = Font(bold=True, color="FFFFFF", size=11)
hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
blue = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
orange = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
red = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
gray = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
redf = Font(bold=True, color="FFFFFF", size=10)
secfont = Font(bold=True, size=12, color="FFFFFF")
bdr = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
al = Alignment(wrap_text=True, vertical="top", horizontal="center")
al_left = Alignment(wrap_text=True, vertical="top")

# Series colors for left stripe
s850 = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
sblog = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
scarousel = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
sflag = PatternFill(start_color="843C0B", end_color="843C0B", fill_type="solid")

headers = ["Series", "#", "Content Title", "Post Date", "LinkedIn\nPost", "LinkedIn\nVideo", "LinkedIn\nNewsletter", "WordPress\nBlog", "YouTube\nVideo", "YouTube\nShort", "Buzzsprout\nPodcast", "Carousel"]
widths = [14, 5, 48, 14, 14, 14, 14, 14, 14, 14, 14, 14]

for c, (h, w) in enumerate(zip(headers, widths), 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = hf
    cell.fill = hfill
    cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
    cell.border = bdr
    ws.column_dimensions[get_column_letter(c)].width = w

ws.row_dimensions[1].height = 40

row = 2

def sec(ws, row, title, fill):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = secfont
    cell.fill = fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = bdr
    return row + 1

def r(ws, row, data, series_fill, rh=50):
    for c, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.border = bdr
        cell.alignment = al if c > 3 else al_left
        if c == 1:
            cell.fill = series_fill
            cell.font = Font(bold=True, color="FFFFFF", size=9)
            cell.alignment = al
        elif isinstance(val, str):
            v = val.upper()
            if v.startswith("POSTED") or v.startswith("LIVE"):
                cell.fill = green
            elif v.startswith("READY"):
                cell.fill = blue
            elif v.startswith("DRAFT"):
                cell.fill = yellow
            elif v.startswith("TO BUILD"):
                cell.fill = orange
            elif v.startswith("POST NOW") or v.startswith("TONIGHT"):
                cell.fill = red
                cell.font = redf
            elif v == "--":
                cell.fill = gray
                cell.alignment = al
            else:
                cell.fill = white
        else:
            cell.fill = white
    ws.row_dimensions[row].height = rh
    return row + 1

# ===== $850 BILLION VIDEO SERIES =====
row = sec(ws, row, "$850 BILLION VIDEO SERIES", s850)

row = r(ws, row, [
    "$850B", "1", "Part 1 — Cost of Ignoring Experienced Workers",
    "Sat Apr 5",
    "POSTED\nSat Apr 5", "READY\nTue May 6", "--", "--",
    "POSTED\nSat Apr 5", "READY\nFri May 23", "READY\nSat May 3", "--"
], s850)

row = r(ws, row, [
    "$850B", "2", "Part 2 — The Investment Gap",
    "Sun Apr 6",
    "READY\nMon May 5", "READY\nMon May 12", "--", "--",
    "POSTED\nSun Apr 6", "READY\nFri May 2", "READY\nWed May 7",
    "TO BUILD\nMon May 5"
], s850)

row = r(ws, row, [
    "$850B", "3", "Part 3 — What Winning Organizations Do",
    "Mon Apr 7",
    "READY\nMon May 19", "READY\nMon May 19", "--", "--",
    "POSTED\nMon Apr 7", "READY\nWed May 28", "READY\nWed May 14",
    "TO BUILD\nMon May 12"
], s850)

row = r(ws, row, [
    "$850B", "4", "Part 4 — The Solution: 50+TechBridge",
    "Sun Apr 13",
    "READY\nTue May 20", "READY\nThu May 22", "--", "--",
    "POSTED\nSun Apr 13", "READY\nFri May 30", "TO BUILD", "--"
], s850)

row = r(ws, row, [
    "$850B", "5", "Bridge — It's About the People",
    "",
    "READY\nTue May 13", "READY\nThu May 15", "--", "--",
    "READY\nUpload", "READY\nVertical", "READY\nWed May 21", "--"
], s850)

row = r(ws, row, [
    "$850B", "6", "$50 Billion Lost (Short)",
    "Mon Apr 20",
    "READY\nThu May 1", "--", "--", "--",
    "--", "POSTED\nMon Apr 20", "--", "--"
], s850)

row = r(ws, row, [
    "$850B", "7", "AI Hiring Bias — Text Post",
    "Tue Apr 29",
    "TONIGHT\nTue Apr 29", "--", "--", "--",
    "--", "--", "--", "--"
], s850)

row = r(ws, row, [
    "$850B", "8", "Investment Gap Math — Text Post",
    "Mon May 5",
    "READY\nMon May 5", "--", "--", "--",
    "--", "--", "--", "--"
], s850)

row = r(ws, row, [
    "$850B", "9", "Flagship — $850B Long Article",
    "Thu May 8",
    "--", "--", "READY\nThu May 8", "READY\nFri May 9",
    "--", "--", "--", "--"
], s850)

row += 1

# ===== WORDPRESS BLOG SERIES =====
row = sec(ws, row, "WORDPRESS BLOG SERIES (Day 01-16)", sblog)

blogs = [
    ["D01", "Why Your 50+ Workforce Is Your Biggest Untapped Asset", "Mon Apr 21", "LIVE\nMon Apr 21"],
    ["D02", "The $850 Billion Cost of Ignoring Your Experienced Workers", "Wed Apr 23", "LIVE\nWed Apr 23"],
    ["D03", "The Confidence Gap: Why AI Training Fails Workers Over 50+", "Sun Apr 27", "LIVE\nSun Apr 27"],
    ["D04", "The Digital Divide for Adults Over 50", "Fri Apr 25", "LIVE\nFri Apr 25"],
    ["D05", "The $76 Trillion Blind Spot: Why Marketers Ignore Adults 50+", "Fri Apr 25", "LIVE\nFri Apr 25"],
    ["D06", "It's Not About the Data. It Never Was.", "Tue Apr 29", "TONIGHT\nTue Apr 29"],
    ["D07", "What Is AgeTech? The Complete Guide for 2026", "Wed Apr 30", "DRAFT\nWed Apr 30"],
    ["D08", "What 23 Organizations Learned Deploying 50+TechBridge", "Thu May 1", "DRAFT\nThu May 1"],
    ["D09", "Pioneer Stories: Real Results from Real People", "Fri May 2", "DRAFT\nFri May 2"],
    ["D10", "AI Training for Older Workers: What Employers Need to Know", "Mon May 5", "DRAFT\nMon May 5"],
    ["D11", "Corporate AI Workshops: The ROI Guide", "Tue May 6", "DRAFT\nTue May 6"],
    ["D12", "Agentic AI and the 50+ Workforce", "Wed May 7", "DRAFT\nWed May 7"],
    ["D13", "Why 50-Year-Old Founders Are 2.8X More Likely to Succeed", "Thu May 8", "DRAFT\nThu May 8"],
    ["D14", "How to Deploy AI Training This Quarter", "Fri May 9", "DRAFT\nFri May 9"],
    ["D15", "I Started at 65. Here's What They Don't Tell You.", "Mon May 12", "DRAFT\nMon May 12"],
    ["D16", "The Agentic Advantage: AI Agents for Professionals Over 50", "Tue May 13", "DRAFT\nTue May 13"],
]

for b in blogs:
    row = r(ws, row, [
        "Blog", b[0], b[1], b[2],
        "--", "--", "--", b[3],
        "--", "--", "--", "--"
    ], sblog)

row += 1

# ===== CAROUSEL SERIES =====
row = sec(ws, row, "CAROUSEL SERIES", scarousel)

carousels = [
    ["C01", "$850B Workforce Gap", "Wed Apr 30", "READY\nWed Apr 30\nRepost Fri May 16"],
    ["C02", "Investment Gap — 5 Stats", "Mon May 5", "TO BUILD\nMon May 5"],
    ["C03", "What Winning Organizations Do", "Mon May 12", "TO BUILD\nMon May 12"],
    ["C04", "347 Pioneers — Real Results", "Mon May 19", "TO BUILD\nMon May 19"],
    ["C05", "The $76 Trillion Opportunity", "Mon May 26", "TO BUILD\nMon May 26"],
    ["C06", "AI Is Not Replacing You", "Wed Jun 4", "TO BUILD\nWed Jun 4"],
    ["C07", "I Started at 65", "Mon Jun 9", "TO BUILD\nMon Jun 9"],
    ["C08", "Why 50-Year-Old Founders Win", "Mon Jun 16", "TO BUILD\nMon Jun 16"],
]

for ca in carousels:
    row = r(ws, row, [
        "Carousel", ca[0], ca[1], ca[2],
        "--", "--", "--", "--",
        "--", "--", "--", ca[3]
    ], scarousel)

row += 2

# ===== COLOR KEY =====
row = sec(ws, row, "COLOR KEY", hfill)
legend = [
    ("GREEN = POSTED / LIVE", green),
    ("BLUE = READY TO POST (file exists, date set)", blue),
    ("YELLOW = DRAFT (in WordPress, hit publish)", yellow),
    ("ORANGE = TO BUILD (needs Canva or production)", orange),
    ("RED = POST NOW / TONIGHT", red),
    ("GRAY = Not applicable for this content", gray),
]
for label, fill in legend:
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=2, value=label)
    cell.fill = fill
    cell.border = bdr
    cell.alignment = al_left
    if fill == red:
        cell.font = redf
    ws.row_dimensions[row].height = 25
    row += 1

# Series key
row += 1
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
ws.cell(row=row, column=2, value="SERIES COLOR (left column)").font = Font(bold=True, size=11)
row += 1
series_legend = [
    ("BLUE = $850 Billion Video Series", s850),
    ("GREEN = WordPress Blog Series", sblog),
    ("GOLD = Carousel Series", scarousel),
]
for label, fill in series_legend:
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=2, value=label)
    cell.fill = fill
    cell.font = Font(bold=True, color="FFFFFF")
    cell.border = bdr
    ws.row_dimensions[row].height = 25
    row += 1

ws.freeze_panes = "D2"

fp = r"C:\Users\USER\Desktop\LMT\850-Billion-Series\MASTER-CONTENT-GRID.xlsx"
wb.save(fp)
print("Saved: " + fp)
