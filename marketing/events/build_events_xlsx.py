import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Events 2026"

headers = [
    "Tier", "Event Name", "Type", "Date", "Location",
    "Why It Matters", "Action", "Status", "Brian Pitch Angle",
    "URL", "Contact / Notes", "Follow-up Date", "Outcome"
]
ws.append(headers)

events = [
    [1, "NAWB Forum 2026", "National Conference", "Mar 23-26, 2026", "Las Vegas, NV (The Cosmopolitan)",
     "Single biggest workforce decision-maker event in US. 1000+ workforce board EDs, employers, policy people.",
     "Attend / Apply to speak / Sponsor", "Not Started",
     "MBE-certified TX adult digital skills training with 3X completion - case study for ETPs serving older workers",
     "https://www.nawb.org/lp/the-forum-2026/", "nawbregistrar@nawb.org", "", ""],
    [1, "Austin SHRM 2026 Annual Conference", "Regional Conference", "TBD 2026", "Austin, TX",
     "400+ HR pros, 100+ Austin metro employers. Direct hiring/training decision makers. Sold out last year.",
     "Attend / Apply to speak / Sponsor", "Not Started",
     "Why your 50+ workers aren't the problem - your training is. Contrarian framework with data.",
     "https://www.austinhumanresource.org/Conference", "", "", ""],
    [1, "Texas Conference for Employers (TCE)", "State Conference Series", "Jan-Sep 2026 (multiple)", "Rotates: Austin, Dallas, Houston, San Antonio, Victoria",
     "TWC + Texas SHRM co-sponsored. Workforce decision makers + state agency staff. Where you meet WIOA contacts.",
     "Attend / Network with TWC staff / Sponsor", "Not Started",
     "I'd love 5 min on how your office handles training provider applications for adult digital skills",
     "https://www.twc.texas.gov/agency/texas-conference-employers", "Bring 50 capability sheets", "", ""],
    [1, "Texas Workforce Conference (Annual)", "State Conference", "Late 2026 TBA", "TX (rotates)",
     "TWC flagship annual. Workforce board EDs from all 28 Texas WDAs. Highest concentration of in-state decision makers.",
     "Attend / Apply to speak adult ed track", "Not Started",
     "Adult digital skills track - 3X completion data + WIOA partnership case study",
     "https://www.twc.texas.gov/agency/annual-texas-workforce-conference", "", "", ""],
    [1, "On Aging 2026", "National Conference", "April 20-23, 2026", "TBD (check ASA site)",
     "Largest aging-services conference. Workforce angle growing. Adjacent buyers + AARP, AGE of CT, senior svc orgs.",
     "Apply to exhibit / Attend workforce track", "Not Started",
     "Workforce + 50+ angle: scaling AI training for adults 50+ with 3X completion",
     "https://asaging.org/on-aging-exhibit/", "", "", ""],
    [2, "Fortune Workplace Innovation Summit", "National Conference", "May 19-20, 2026", "Atlanta, GA",
     "Fortune editors + Fortune 500 HR. Older workers/50+ talent theme. Press coverage if featured.",
     "Apply to speak / Attend / PR angle", "Not Started",
     "Texas case study: how MBE-certified training scales for adults 50+",
     "https://fortune.com/conferences/", "", "", ""],
    [2, "Age Pioneers Summit", "International Conference", "November 18, 2026", "London, UK",
     "Intl audience: CPOs, CHROs, age-inclusion leaders. Watch for streamed/virtual attendance.",
     "Submit session proposal / Virtual attend", "Not Started",
     "US perspective: state workforce funding for adult digital skills",
     "https://work-redefined.co/resources/age-pioneers", "", "", ""],
    [2, "CES AgeTech (AARP AgeTech Collaborative)", "National Tech Showcase", "Jan 2027 (next cycle)", "Las Vegas, NV",
     "AARP runs AgeTech Collaborative booth. 22+ startups. Massive press visibility.",
     "Apply to AARP AgeTech Collaborative cohort", "Not Started",
     "LMT as AgeTech innovation: 3X completion via human-centered AI training design",
     "https://agetechcollaborative.org", "6-9 month application cycle", "", ""],
    [2, "NAWB Webinars (monthly)", "Online", "Monthly", "Online",
     "Free way to stay current on workforce policy + face-time with NAWB staff in chat",
     "Subscribe / Attend 1+/month / Ask substantive Qs", "Not Started",
     "Substantive question per webinar to build name recognition",
     "https://www.nawb.org/events/", "", "", ""],
    [3, "Texas EDC Winter Conference", "Regional Conference", "Feb 25-27, 2026", "El Paso, TX",
     "Economic development directors from cities/counties. They fund and influence workforce programs.",
     "Attend / Present case study", "Not Started",
     "MBE workforce solution for ED directors: Austin pilot results",
     "https://texasedc.org/programs/conferences", "", "", ""],
    [3, "Texas EDC Mid-Year Conference", "Regional Conference", "Jun 17-19, 2026", "Plano, TX",
     "Closest to Austin. ED directors network for LMT contracts.",
     "Attend / Present case study", "Not Started",
     "Same as Winter - ED directors need workforce solutions",
     "https://texasedc.org/programs/conferences", "", "", ""],
    [3, "Texas EDC Annual Conference", "Regional Conference", "Oct 7-9, 2026", "Houston, TX",
     "Largest TX EDC event. Maximum reach across state.",
     "Attend / Apply to present", "Not Started",
     "Statewide ED case study: WIOA-funded adult digital skills",
     "https://texasedc.org/programs/conferences", "", "", ""],
    [3, "Texas SHRM (state-level)", "Annual + Chapter Events", "Various 2026", "Rotates Texas",
     "State-level HR community. Bigger reach than Austin SHRM alone.",
     "Follow events / Apply to speak", "Not Started",
     "Workforce retention math + actionable framework",
     "https://texasshrm.org/events", "", "", ""],
    [3, "SXSW Innovate Austin", "Festival/Conference", "March 2026", "Austin, TX (during SXSW)",
     "SXSW + Austin Chamber partnership. Workforce/innovation track. Free networking.",
     "Attend free Chamber events / Apply SXSW 2027", "Not Started",
     "Innovation panel: AI training designed for the workers everyone overlooks",
     "https://www.austinchamber.com/events/innovate-austin-march-2026", "Apps for SXSW 2027 open summer 2026", "", ""],
    [3, "Austin Chamber Annual Meeting", "Local Civic", "Various 2026", "Austin, TX",
     "Where Austin's top business leaders gather. Build/maintain executive class visibility.",
     "Attend / Network", "Not Started",
     "Brief intro + capability sheet to 1 new contact per event",
     "https://www.austinchamber.com/events", "", "", ""],
    [3, "Greater Austin Black Chamber Events", "Local Civic", "Monthly 2026", "Austin, TX",
     "MBE community + Black-owned business network. Brian's natural network.",
     "Become member / Attend monthly / Volunteer committee", "Not Started",
     "MBE-to-MBE collaboration on workforce contracts",
     "https://www.austinbcc.org/events/calendar/", "", "", ""],
    [3, "Greater Austin Hispanic Chamber", "Local Civic", "Monthly 2026", "Austin, TX",
     "MBE community + decision-maker access via workforce/education committees",
     "Become member / Attend monthly / Join workforce committee", "Not Started",
     "Adult digital skills for Hispanic 50+ workforce",
     "https://www.gahcc.org/", "", "", ""],
    [4, "WFS Capital Area Board Meeting", "Recurring Civic", "Monthly", "Austin, TX (in-person + virtual)",
     "MOST IMPORTANT MEETING FOR LMT. Approves WIOA training providers in Travis County. Public comment period.",
     "Attend / Public comment / Network with board", "Not Started",
     "3-min public comment introducing LMT + handing capability sheets to board members during break",
     "https://wfscapitalarea.com", "Find next meeting on website", "", ""],
    [4, "Austin City Council", "Recurring Civic", "Thursdays weekly", "Austin City Hall, 301 W 2nd St",
     "Workforce items appear monthly. Public comment available. Council members accessible.",
     "Attend on workforce items / Public comment", "Not Started",
     "Connect $850B series data to City workforce funding decisions",
     "https://www.austintexas.gov/department/city-council", "Sign up 24 hrs in advance", "", ""],
    [4, "Travis County Commissioners Court", "Recurring Civic", "Tuesdays weekly", "Travis County, Austin",
     "County workforce funding. Public comment available.",
     "Attend on relevant items", "Not Started",
     "County workforce strategy + adult digital skills",
     "https://www.traviscountytx.gov/commissioners-court", "", "", ""],
    [4, "Austin Commission on Aging", "Recurring Civic", "Monthly", "Austin, TX",
     "Direct line to aging-services policy in Austin. Workforce + 50+ angle.",
     "Attend / Apply to serve on commission", "Not Started",
     "Workforce angle on aging services - instant civic credibility if you serve",
     "https://austintexas.gov", "Search Commission on Aging schedule", "", ""],
    [4, "Mayor Committee on People with Disabilities", "Recurring Civic", "Monthly", "Austin, TX",
     "Workforce equity angle. ADA compliance for LMT. Disability employment networks.",
     "Attend on workforce items", "Not Started",
     "Adults 50+ + disability + workforce intersection",
     "https://austintexas.gov", "", "", ""],
    [5, "Austin Rotary Clubs", "Speaking Gig", "Weekly (book 3-6 mo out)", "Multiple Austin chapters",
     "Need a speaker every week. Great rehearsal + lead source. Every Rotarian = a business owner.",
     "Pitch all 8+ chapters", "Not Started",
     "Why your 50+ workers aren't the problem - 30 min + Q&A",
     "Search 'Austin Rotary Club' + chapter names", "", "", ""],
    [5, "Austin Lions Clubs", "Speaking Gig", "Weekly/biweekly", "Multiple Austin chapters",
     "Older demographic = direct LMT message alignment. Easy book.",
     "Pitch 5 chapters", "Not Started",
     "Same Rotary deck + community angle",
     "Search 'Lions Club Austin'", "", "", ""],
    [5, "Austin Kiwanis Clubs", "Speaking Gig", "Weekly", "Multiple Austin chapters",
     "Civic leaders, business owners, school connections.",
     "Pitch 3 chapters", "Not Started",
     "Workforce + school-to-work angle",
     "Search 'Kiwanis Club Austin'", "", "", ""],
    [5, "Austin Public Library Lecture Series", "Speaking Gig", "Various 2026", "Multiple branch locations",
     "Library workshops already exist. They WANT speakers on tech literacy and adult learning.",
     "Email each branch's adult programming coord", "Not Started",
     "AI for Adults 50+ - free 45-min workshop",
     "https://library.austintexas.gov/", "", "", ""],
    [5, "AGE of Central Texas", "Speaking Gig", "Quarterly", "Austin, TX",
     "Direct Pioneer recruitment + speaker stage at aging-focused org.",
     "Pitch workshop or speaker slot", "Not Started",
     "AI/digital skills workshop for AGE members",
     "https://www.ageofcentraltx.org", "", "", ""],
    [5, "Senior Services / Senior Centers", "Speaking Gig", "Ongoing", "Multiple locations",
     "End-user direct + community impact for grant/contract pitches.",
     "Pitch each senior center for quarterly workshop", "Not Started",
     "Tech 1-on-1 workshop intro to LMT curriculum",
     "https://seniorservicesofaustin.org", "", "", ""],
    [5, "Faith Communities (Austin churches)", "Speaking Gig", "Adult ed programs", "Austin churches",
     "Direct Pioneer recruitment + community trust + word of mouth.",
     "Pitch 5 churches", "Not Started",
     "Brian's natural network + faith community angle",
     "", "Especially Black church + Hispanic Catholic networks", "", ""],
    [6, "U.S. Conference of Mayors Winter Meeting", "National Conference", "January 2026", "Washington, DC",
     "Mayors of major cities + workforce staff. National influence.",
     "Apply to present / Attend", "Not Started",
     "City workforce strategy + adults 50+",
     "https://www.usmayors.org", "", "", ""],
    [6, "U.S. Conference of Mayors Annual", "National Conference", "June 2026", "Rotates",
     "Same as Winter, larger turnout.",
     "Apply to present / Attend", "Not Started",
     "Same - city workforce angle",
     "https://www.usmayors.org", "", "", ""],
    [6, "NACo Annual Conference", "National Conference", "July 2026", "Rotates",
     "County workforce program directors. Federal funding flows to counties.",
     "Attend / Apply to present", "Not Started",
     "County WIOA partnership case study",
     "https://www.naco.org", "", "", ""],
    [6, "U.S. DOL ETA Listening Sessions", "Federal Policy", "Quarterly", "Various / online",
     "Direct line to federal workforce funding policy. Voice can shape WIOA reauthorization.",
     "Subscribe / Attend / Submit comments", "Not Started",
     "Adult 50+ digital skills funding gaps in current WIOA",
     "https://www.dol.gov/agencies/eta", "", "", ""],
]

for event in events:
    ws.append(event)

header_fill = PatternFill(start_color="0E1C2F", end_color="0E1C2F", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
for col_num in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align
ws.row_dimensions[1].height = 32

tier_colors = {
    1: "FFE5CC", 2: "FFF5CC", 3: "E5F5E0",
    4: "E5F0FF", 5: "F0E5FF", 6: "FFE5E5",
}

for row_num in range(2, len(events) + 2):
    tier_val = ws.cell(row=row_num, column=1).value
    fill = PatternFill(start_color=tier_colors.get(tier_val, "FFFFFF"),
                       end_color=tier_colors.get(tier_val, "FFFFFF"),
                       fill_type="solid")
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.fill = fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)

widths = [6, 38, 22, 22, 32, 50, 32, 14, 50, 45, 35, 14, 30]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"

ws2 = wb.create_sheet("Legend")
ws2.append(["Tier", "Description", "Color"])
legend = [
    [1, "ACT NOW - high impact, near-term", "Light Orange"],
    [2, "National events - track and apply", "Light Yellow"],
    [3, "Texas regional - attend and network", "Light Green"],
    [4, "Recurring civic - free, weekly/monthly", "Light Blue"],
    [5, "Speaking-only - fast, free, builds reel", "Light Purple"],
    [6, "Federal - long-term, big-ticket", "Light Pink"],
]
for row in legend:
    ws2.append(row)
for col_num in range(1, 4):
    ws2.cell(row=1, column=col_num).fill = header_fill
    ws2.cell(row=1, column=col_num).font = header_font
for row_num, color_key in zip(range(2, 8), [1, 2, 3, 4, 5, 6]):
    fill = PatternFill(start_color=tier_colors[color_key], end_color=tier_colors[color_key], fill_type="solid")
    for col_num in range(1, 4):
        ws2.cell(row=row_num, column=col_num).fill = fill
ws2.column_dimensions["A"].width = 8
ws2.column_dimensions["B"].width = 60
ws2.column_dimensions["C"].width = 18

ws3 = wb.create_sheet("Status Values")
ws3.append(["Status"])
for s in ["Not Started", "Researching", "Applied / Pitched", "Confirmed", "Attending", "Attended", "Followed Up", "Closed"]:
    ws3.append([s])
ws3.column_dimensions["A"].width = 25

out_path = r"C:\Users\USER\Documents\50techbridge-content\marketing\events-tracker-2026.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
print(f"Total events: {len(events)}")
