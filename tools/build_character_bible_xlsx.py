"""
What If Brand Film — Character Bible
Consistent character system for ongoing story development.
All named characters across the full film series.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = "VIDEOS/what-if-psa/WHAT-IF-CHARACTER-BIBLE.xlsx"

# ── Styles ────────────────────────────────────────────────────────────────────
NAVY_FILL   = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
NAVY_FONT   = Font(bold=True, color="FFFFFF", size=11)
GOLD_FILL   = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
GOLD_FONT   = Font(bold=True, color="1F4E79", size=10)
GREEN_FILL  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
GREEN_FONT  = Font(bold=True, color="276221", size=10)
RED_FILL    = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
RED_FONT    = Font(color="9C0006", size=10)
BLUE_FILL   = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
BLUE_FONT   = Font(bold=True, color="1F4E79", size=10)
ALT_FILL    = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")
BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

def cell_set(ws, row, col, val, fill=None, font=None, wrap=True, height=None, bold=False, size=10, color="000000"):
    c = ws.cell(row=row, column=col)
    c.value = val
    c.border = BORDER
    c.alignment = Alignment(vertical="center", wrap_text=wrap)
    if fill: c.fill = fill
    if font:
        c.font = font
    else:
        c.font = Font(bold=bold, size=size, color=color)
    if height:
        ws.row_dimensions[row].height = height
    return c

def section_header(ws, row, text, cols=8):
    ws.merge_cells(f"A{row}:{get_column_letter(cols)}{row}")
    c = ws.cell(row=row, column=1)
    c.value = text
    c.fill = NAVY_FILL
    c.font = NAVY_FONT
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 28

def page_title(ws, text, sub, cols=8):
    ws.merge_cells(f"A1:{get_column_letter(cols)}1")
    ws["A1"].value = text
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34
    ws.merge_cells(f"A2:{get_column_letter(cols)}2")
    ws["A2"].value = sub
    ws["A2"].font = Font(italic=True, size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════════════════════
# SHEET 1 — CHARACTER ROSTER
# ═══════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Character Roster"

page_title(ws1,
    "What If Brand Film — Character Bible",
    "Consistent character system for ongoing story development | Learn More Technologies | 50+TechBridge")

# Headers
headers = ["Character", "Age", "Race/Ethnicity", "Role in Film", "World", "Scenes", "Anchor File", "Status"]
section_header(ws1, 4, "ALL CHARACTERS — WHAT IF BRAND FILM SERIES")
for col, h in enumerate(headers, 1):
    cell_set(ws1, 5, col, h, fill=GOLD_FILL, font=GOLD_FONT, height=32)

characters = [
    # name, age, race, role, world, scenes, anchor_file, status
    ("Marcus Johnson", "63", "Black / African American",
     "Primary protagonist. 28-year employee. Position eliminated by algorithm. His story opens the film.",
     "Machine World → Bridge", "Scenes 1, 2, 6", "marcus-johnson-anchor.png", "🔲 Generate anchor"),

    ("Carol Smith", "54", "White / Caucasian",
     "Secondary protagonist. Algorithm interview scene. Most qualified person in the room. Never gets to ask her three questions.",
     "Machine World", "Scene 3", "carol-smith-anchor.png", "🔲 Generate anchor"),

    ("Dorothy Williams", "64", "Black / African American",
     "The Turn character. Her real smile — the first warm moment in the film. Community center. AI changes what she knows about herself.",
     "Human World", "Scene 9", "dorothy-williams-anchor.png", "🔲 Generate anchor"),

    ("Rosa Gutierrez", "52", "Latina / Hispanic",
     "Three Workers scene. Left of frame. Reads the optimization dashboard. Her face: this has happened before.",
     "Machine World", "Scene 6", "rosa-gutierrez-anchor.png", "🔲 Generate anchor"),

    ("James Okafor", "58", "Black / African American",
     "Three Workers scene. Center — sharpest focus. The face of tired-not-surprised. Has been reading this room for years.",
     "Machine World", "Scene 6", "james-okafor-anchor.png", "🔲 Generate anchor"),

    ("Tom Briggs", "61", "White / Caucasian",
     "Three Workers scene. Right of frame. The veteran. Has watched younger colleagues make these slides for a decade.",
     "Machine World", "Scene 6", "tom-briggs-anchor.png", "🔲 Generate anchor"),

    ("Linda Chen", "59", "Asian American",
     "Pioneer montage. Presenting AI analysis to a room of younger colleagues. She built it. She knows it cold.",
     "Human World", "Scene 11 (montage)", "linda-chen-anchor.png", "🔲 Generate anchor"),

    ("Marcus Johnson (Reversed)", "63", "Black / African American",
     "Same character as Marcus — Human World version. Interviewer is leaning in. Knowledge is being seen.",
     "Human World", "Scene 12", "marcus-johnson-anchor.png", "♻️ Reuse Marcus anchor"),

    ("Brian McKinney", "Real person", "Black / African American",
     "Founder. Speaks direct to camera. Scene 10. Not AI-generated — filmed on iPhone or studio.",
     "Both worlds", "Scenes 10, 14", "REAL — film on camera", "📹 Film live"),
]

for i, row in enumerate(characters):
    r = i + 6
    for col, val in enumerate(row, 1):
        fill = ALT_FILL if i % 2 else None
        font = None
        if col == 5:  # World column
            if "Machine" in val: fill = RED_FILL; font = RED_FONT
            elif "Human" in val: fill = GREEN_FILL; font = GREEN_FONT
            elif "Both" in val: fill = GOLD_FILL; font = GOLD_FONT
        if col == 8:  # Status
            if "🔲" in val: font = Font(size=10, color="9C0006")
            elif "♻️" in val: font = Font(size=10, color="276221", bold=True)
            elif "📹" in val: font = Font(size=10, color="1F4E79", bold=True)
        cell_set(ws1, r, col, val, fill=fill, font=font, height=52)

ws1.column_dimensions["A"].width = 22
ws1.column_dimensions["B"].width = 6
ws1.column_dimensions["C"].width = 20
ws1.column_dimensions["D"].width = 44
ws1.column_dimensions["E"].width = 16
ws1.column_dimensions["F"].width = 20
ws1.column_dimensions["G"].width = 28
ws1.column_dimensions["H"].width = 22
ws1.freeze_panes = "A6"


# ═══════════════════════════════════════════════════════════════════
# SHEET 2 — CHARACTER BRIEFS (one per character)
# ═══════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Character Briefs")
page_title(ws2,
    "Character Briefs — Full Visual Detail",
    "Every field used in Midjourney prompts. Copy exactly — never retype from memory.")

briefs = [
    {
        "name": "Marcus Johnson",
        "fields": [
            ("Age", "63"),
            ("Race / Ethnicity", "Black / African American"),
            ("Hair", "Close-cropped salt-and-pepper natural hair, slightly receded at temples"),
            ("Facial hair", "Short salt-and-pepper beard, neatly kept — 3-4 day growth"),
            ("Eyes", "Dark brown, deep-set, carries weight in them"),
            ("Skin tone", "Deep brown, warm undertone"),
            ("Build", "Medium build, slightly broadened with age — carries himself with decades of quiet authority"),
            ("Wardrobe — Machine World", "Dark charcoal suit, white dress shirt, no tie — the suit of someone who dressed for a meeting that ended before it started. ID badge on lanyard."),
            ("Wardrobe — Human World", "Same suit, jacket open, more relaxed — the man who got his footing back"),
            ("Expression anchor", "Not defeated. Not angry. Somewhere between exhausted and resolute. He has been here before."),
            ("Distinctive detail", "Wedding ring, left hand. He has someone to go home to. That matters."),
            ("Voice/manner", "Measured. Precise. Does not raise his voice. Has never needed to."),
            ("Midjourney anchor prompt", "Black man, age 63, close-cropped salt-and-pepper natural hair, short salt-and-pepper beard, dark brown deep-set eyes, deep brown warm-toned skin, medium build with quiet authority, dark charcoal suit white dress shirt no tie, ID badge on lanyard, wedding ring left hand, composed dignified expression — exhausted resolve not defeat, character reference sheet front view three-quarter view side view, white studio background, soft directional lighting, photorealistic 8K --ar 3:2 --cw 100"),
        ]
    },
    {
        "name": "Carol Smith",
        "fields": [
            ("Age", "54"),
            ("Race / Ethnicity", "White / Caucasian"),
            ("Hair", "Silver-streaked dark chestnut brown, shoulder length, natural wave, side part — professional, not over-styled"),
            ("Eyes", "Hazel-green"),
            ("Skin tone", "Fair, natural fine lines at eyes and mouth — not hidden"),
            ("Build", "Medium build, straight posture. Carries herself with earned authority."),
            ("Wardrobe", "Deep navy fitted blazer, ivory button-down, small pearl stud earrings, simple gold watch on left wrist"),
            ("Expression anchor", "Composed readiness. She came prepared. She is waiting for the conversation to begin. It will not begin the way she expects."),
            ("Distinctive detail", "Leather portfolio open on the table. Three questions written inside she never gets to ask."),
            ("Midjourney anchor prompt", "White woman, age 54, silver-streaked dark chestnut brown shoulder-length hair natural wave side part, hazel-green eyes, fair skin natural fine lines, medium build straight posture, deep navy fitted blazer ivory button-down, small pearl stud earrings, simple gold watch left wrist, composed ready expression, character reference sheet front view three-quarter view side view, white studio background, soft neutral lighting, photorealistic 8K --ar 3:2 --cw 100"),
        ]
    },
    {
        "name": "Dorothy Williams",
        "fields": [
            ("Age", "64"),
            ("Race / Ethnicity", "Black / African American"),
            ("Hair", "Silver natural hair, short afro — fully grey, beautiful, not hidden"),
            ("Eyes", "Warm brown, reading glasses on face or hanging around neck"),
            ("Skin tone", "Rich medium-dark brown, warm undertone, natural lines of expression"),
            ("Build", "Compact, warm energy — the person everyone leans toward"),
            ("Wardrobe", "Burgundy cardigan over a patterned blouse. Reading glasses. Comfortable, real — not corporate."),
            ("Expression anchor", "The real smile. Not polite. The one that comes from understanding something you didn't expect to understand about yourself."),
            ("Distinctive detail", "Reading glasses — either on her face or on a chain around her neck. Basic laptop open in front of her."),
            ("Setting", "Community center table. Folding tables, modest laptops, warm fluorescent light that somehow feels human."),
            ("Midjourney anchor prompt", "Black woman, age 64, short silver natural afro, warm brown eyes, reading glasses, rich medium-dark brown warm-toned skin, natural expression lines, compact warm build, burgundy cardigan patterned blouse, seated at folding table with basic laptop open, expression of genuine private surprise becoming a real smile — not for camera, character reference sheet front view three-quarter view, warm community center lighting, photorealistic 8K --ar 3:2 --cw 100"),
        ]
    },
    {
        "name": "Rosa Gutierrez",
        "fields": [
            ("Age", "52"),
            ("Race / Ethnicity", "Latina / Hispanic"),
            ("Hair", "Dark brown, pulled back in a low bun — a few strands loose. Professional."),
            ("Eyes", "Dark brown"),
            ("Skin tone", "Warm medium brown, olive undertone"),
            ("Build", "Lean, held tension — she is reading the room very carefully"),
            ("Wardrobe", "Dark grey business jacket, simple dark blouse beneath. No jewelry beyond small gold hoops."),
            ("Expression anchor", "Her face is controlled. She has learned not to react visibly. But her jaw is set. This has happened before."),
            ("Distinctive detail", "She sits slightly forward — not relaxed. Prepared for impact."),
            ("Midjourney anchor prompt", "Latina woman, age 52, dark brown hair pulled back low bun a few strands loose, dark brown eyes, warm medium brown olive-undertoned skin, lean controlled posture, dark grey business jacket simple dark blouse, small gold hoop earrings, expression: controlled — jaw set, not afraid, has read this room before, character reference sheet front three-quarter view, cold blue-white corporate lighting, photorealistic 8K --ar 3:2 --cw 100"),
        ]
    },
    {
        "name": "James Okafor",
        "fields": [
            ("Age", "58"),
            ("Race / Ethnicity", "Black / African American — Nigerian heritage"),
            ("Hair", "Close-cropped, fully grey, natural"),
            ("Eyes", "Dark brown, very still"),
            ("Skin tone", "Deep brown, medium-dark, warm undertone"),
            ("Build", "Solid, experienced. Has been the most reliable person in every room he has ever entered."),
            ("Wardrobe", "Navy blue suit, white shirt, no tie — dressed up but not pretending"),
            ("Expression anchor", "Tired-not-surprised. The exhaustion of a man who has been watching this happen to other people for twenty years and now the number is pointing at him. He is not shocked. He is just — done pretending he didn't know it was coming."),
            ("Key scene", "He is the CENTER of the three workers. Sharpest focus. His face carries the scene."),
            ("Midjourney anchor prompt", "Black man, age 58, close-cropped fully grey natural hair, dark brown very still eyes, deep medium-dark brown warm-toned skin, solid experienced build, navy blue suit white shirt no tie, expression: tired not surprised — the specific exhaustion of someone who has watched this happen and is now the number, not defeated but done pretending, character reference sheet front view three-quarter view, cold blue-white corporate screen light, photorealistic 8K --ar 3:2 --cw 100"),
        ]
    },
    {
        "name": "Tom Briggs",
        "fields": [
            ("Age", "61"),
            ("Race / Ethnicity", "White / Caucasian"),
            ("Hair", "Silver-white, short, side-parted. Classic."),
            ("Eyes", "Blue, direct"),
            ("Skin tone", "Fair, weathered — someone who has spent time both outdoors and in boardrooms"),
            ("Build", "Tall, still straight-backed. Not gone yet."),
            ("Wardrobe", "Charcoal suit, blue tie, white shirt. He still dresses like the job respects him."),
            ("Expression anchor", "Veteran read. He has watched younger colleagues make these slides for a decade. He is not reading the dashboard for the first time. He is watching it happen to someone he knows."),
            ("Midjourney anchor prompt", "White man, age 61, silver-white short side-parted hair, direct blue eyes, fair weathered skin, tall straight-backed build, charcoal suit blue tie white shirt, expression: veteran recognition — he has watched this happen before and now watches it again, quiet and resigned but not broken, character reference sheet front view three-quarter view, cold corporate blue-white lighting, photorealistic 8K --ar 3:2 --cw 100"),
        ]
    },
    {
        "name": "Linda Chen",
        "fields": [
            ("Age", "59"),
            ("Race / Ethnicity", "Asian American — Chinese American"),
            ("Hair", "Black with silver threads at temples, chin length bob, clean and precise"),
            ("Eyes", "Dark brown, sharp, present"),
            ("Skin tone", "Light-medium, warm golden undertone"),
            ("Build", "Trim, upright posture — someone accustomed to commanding a room"),
            ("Wardrobe", "Deep teal blazer, ivory blouse. Simple gold pendant necklace. Professional, distinctive — she is not trying to disappear."),
            ("Expression anchor", "She knows this material. Her posture is that of someone who earned the floor, not borrowed it."),
            ("Scene", "Standing at front of conference room. Her AI analysis on screen behind her. Colleagues leaning in."),
            ("Midjourney anchor prompt", "Chinese American woman, age 59, black hair silver at temples chin-length bob, dark brown sharp present eyes, light-medium warm golden-toned skin, trim upright commanding posture, deep teal blazer ivory blouse simple gold pendant necklace, expression: quiet authority — she knows this and earned the floor, character reference sheet front view three-quarter view, warm professional conference room lighting, photorealistic 8K --ar 3:2 --cw 100"),
        ]
    },
]

row_num = 4
for char in briefs:
    section_header(ws2, row_num, f"CHARACTER: {char['name'].upper()}", cols=4)
    row_num += 1
    for field, value in char["fields"]:
        is_prompt = "prompt" in field.lower()
        fill = BLUE_FILL if is_prompt else (ALT_FILL if row_num % 2 == 0 else None)
        font = BLUE_FONT if is_prompt else None
        cell_set(ws2, row_num, 1, field, fill=fill, font=font, bold=not is_prompt, height=20 if not is_prompt else 60)
        cell_set(ws2, row_num, 2, value, fill=fill, font=Font(size=10, color="1F4E79" if is_prompt else "000000"), height=20 if not is_prompt else 60)
        ws2.merge_cells(f"B{row_num}:H{row_num}")
        row_num += 1
    row_num += 1  # spacer

ws2.column_dimensions["A"].width = 26
ws2.column_dimensions["B"].width = 90
ws2.freeze_panes = "A4"


# ═══════════════════════════════════════════════════════════════════
# SHEET 3 — SCENE ↔ CHARACTER TRACKING
# ═══════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Scene Tracking")
page_title(ws3,
    "Scene × Character Tracking — What If Film",
    "Which character appears in which scene | Reference image used | Status | File path")

track_headers = ["Scene #", "Scene Name", "Timing", "Character(s)", "Reference Image File", "Tool", "Output File", "Status"]
section_header(ws3, 4, "MACHINE WORLD SCENES (Cold — Blue-White Lighting)", cols=8)
for col, h in enumerate(track_headers, 1):
    cell_set(ws3, 5, col, h, fill=GOLD_FILL, font=GOLD_FONT, height=30)

machine_scenes = [
    ("1", "Cold Open — Data Screen", "0:00–0:12", "No character — text only", "N/A", "Canva", "scene-1-data-open.mp4", "🔲 Build in Canva"),
    ("2", "The March", "0:12–0:35", "Marcus Johnson", "marcus-johnson-anchor.png", "Galaxy AI / Kling", "scene-2-the-march.mp4", "🔲 Generate"),
    ("3", "The Hiring Algorithm", "0:35–0:58", "Carol Smith + Young Interviewer", "carol-smith-anchor.png", "Midjourney → Kling", "scene-3-algorithm-interview.mp4", "🔲 Generate"),
    ("4", "The Kiosk She Can't Navigate", "0:58–1:12", "Dorothy Williams (Machine World version)", "dorothy-williams-anchor.png", "Galaxy AI / Kling", "scene-4-kiosk-fail.mp4", "🔲 Generate"),
    ("5", "The Authority Screen", "1:12–1:28", "No named character — presenter silhouetted", "N/A", "Galaxy AI / Kling", "scene-5-boardroom-screen.mp4", "🔲 Generate"),
    ("6", "The Optimization Dashboard", "1:28–1:45", "Rosa Gutierrez + James Okafor + Tom Briggs", "3 anchor files", "Galaxy AI / Kling", "scene-6-dashboard.mp4", "🔲 Generate"),
    ("7", "Three Workers Reading the Room", "1:45–2:00", "Rosa + James (focus) + Tom", "james-okafor-anchor.png (primary)", "Kling", "scene-7-three-workers.mp4", "🔲 Generate"),
]

for i, row in enumerate(machine_scenes):
    r = i + 6
    for col, val in enumerate(row, 1):
        cell_set(ws3, r, col, val, fill=RED_FILL if i % 2 == 0 else PatternFill(), height=44)

section_header(ws3, 14, "THE TURN (Silence)", cols=8)
cell_set(ws3, 15, 1, "8")
cell_set(ws3, 15, 2, "The Break — One Warm Light")
cell_set(ws3, 15, 3, "2:00–2:03")
cell_set(ws3, 15, 4, "No character — light only")
cell_set(ws3, 15, 5, "N/A")
cell_set(ws3, 15, 6, "Galaxy AI / Kling")
cell_set(ws3, 15, 7, "scene-8-the-break.mp4")
cell_set(ws3, 15, 8, "🔲 Generate")
ws3.row_dimensions[15].height = 36

section_header(ws3, 17, "HUMAN WORLD SCENES (Warm — Gold/Natural Lighting)", cols=8)
for col, h in enumerate(track_headers, 1):
    cell_set(ws3, 18, col, h, fill=GOLD_FILL, font=GOLD_FONT, height=30)

human_scenes = [
    ("9", "Dorothy's Real Smile", "2:03–2:20", "Dorothy Williams", "dorothy-williams-anchor.png", "Midjourney → Kling", "scene-9-real-smile.mp4", "🔲 Generate"),
    ("10", "Brian — Direct to Camera", "2:20–3:10", "Brian McKinney (real)", "FILM LIVE — iPhone or studio", "iPhone / HeyGen", "scene-10-brian-speech.mp4", "📹 Film live"),
    ("11", "Linda — Pioneer Presenting", "3:10–3:25", "Linda Chen", "linda-chen-anchor.png", "Midjourney → Kling", "scene-11-linda-presenting.mp4", "🔲 Generate"),
    ("12", "Marcus — Reversed Interview", "3:25–3:40", "Marcus Johnson (Human World)", "marcus-johnson-anchor.png ♻️ reuse", "Kling", "scene-12-reversed-interview.mp4", "🔲 Generate"),
    ("13", "Community Center Montage", "3:40–4:10", "Dorothy + Others (background)", "dorothy-williams-anchor.png", "Kling / Galaxy AI", "scene-13-montage.mp4", "🔲 Generate"),
    ("14", "Brian — Final Close", "4:10–4:42", "Brian McKinney (real)", "FILM LIVE", "iPhone / HeyGen", "scene-14-brian-close.mp4", "📹 Film live"),
    ("15", "Final Title Card", "4:42–4:50", "No character — text", "N/A", "Canva", "scene-15-title-card.mp4", "🔲 Build in Canva"),
]

for i, row in enumerate(human_scenes):
    r = i + 19
    for col, val in enumerate(row, 1):
        cell_set(ws3, r, col, val, fill=GREEN_FILL if i % 2 == 0 else PatternFill(), height=44)

for col_idx, width in enumerate([6, 28, 12, 34, 30, 20, 30, 22], start=1):
    ws3.column_dimensions[get_column_letter(col_idx)].width = width
ws3.freeze_panes = "A6"


# ═══════════════════════════════════════════════════════════════════
# SHEET 4 — CONSISTENCY RULES (the law)
# ═══════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Consistency Rules")
page_title(ws4,
    "Character Consistency — Rules of the Road",
    "Follow these or the character drifts. Every rule has a reason.")

rules = [
    ("ANCHOR IMAGE", "Rule", "Explanation"),
    ("Never regenerate the anchor image", "ABSOLUTE",
     "Midjourney is non-deterministic. Regenerating with the same prompt produces a different face. The anchor image from Day 1 is the law. Store it. Back it up. Never delete it."),
    ("One anchor per character — not per scene", "ABSOLUTE",
     "All scenes across all episodes use the same anchor file. If you generate a new anchor for episode 2, you have a different character."),
    ("Store all anchor files in one place", "ABSOLUTE",
     "Path: C:/Users/USER/Documents/LMT-Video-Assets/CHARACTERS/[CharacterName]/[name]-anchor.png"),
    ("", "", ""),
    ("PROMPT DISCIPLINE", "Rule", "Explanation"),
    ("Save the exact Midjourney prompt as a .txt file", "REQUIRED",
     "Save the prompt that generated your anchor. If you ever need to generate more stills, you need this exact text. Never retype from memory — even one word changes the output."),
    ("Use --cref [anchor URL] --cw 100 in every still generation", "REQUIRED",
     "--cref locks the character reference. --cw 100 gives it maximum weight. Never lower --cw below 80 for a scene still."),
    ("Use --oref [anchor URL] --ow 80 for Omni Reference (MJ v7+)", "REQUIRED",
     "Omni Reference holds identity more tightly across different poses. Use both --cref and --oref together for strongest consistency."),
    ("", "", ""),
    ("SCENE GENERATION RULES", "Rule", "Explanation"),
    ("Change ONE variable at a time", "ABSOLUTE",
     "New location = same angle + same lighting. New angle = same location + same lighting. Changing two variables at once is the primary cause of identity drift."),
    ("Generate stills BEFORE video", "REQUIRED",
     "Confirm the character looks right in Midjourney before touching Kling. Fix drift in stills first — it is cheaper and faster than fixing it in video."),
    ("Always use the still as Kling's start frame", "REQUIRED",
     "Never text-prompt a character directly in Kling without a start frame. Start frame + Character ID = the two locks that hold identity in video."),
    ("Upload anchor to Kling Character ID every session", "REQUIRED",
     "Kling's Character ID feature uses your anchor image to lock the face across motion. Upload it at the start of every Kling session — it does not persist between sessions."),
    ("", "", ""),
    ("TWO-CHARACTER SCENES", "Rule", "Explanation"),
    ("Upload BOTH anchor images in Kling's multi-character reference", "REQUIRED",
     "Describe each character separately in the prompt. 'Character A: [description] stands on left. Character B: [description] stands on right.'"),
    ("Budget 3x more credits for two-character scenes", "REQUIRED",
     "Two-character scenes have a 40-50% reject rate from drift. Generate 3x what you need and pick the best."),
    ("Keep two-character interaction simple", "REQUIRED",
     "Complex movement between two characters = more drift. Wide shot, limited motion, implied interaction is safer than dynamic action between two generated characters."),
    ("", "", ""),
    ("FILE MANAGEMENT", "Rule", "Explanation"),
    ("Naming convention for all output files", "REQUIRED",
     "Format: scene-[number]-[character-shortname]-[take].mp4\nExample: scene-3-carol-take-2.mp4\nNever use generic names like 'video1.mp4'"),
    ("Shot log — track every accepted clip", "REQUIRED",
     "Columns: Scene | Character | Reference File Used | Kling Settings | Accepted Clip Filename | Date Generated\nUpdate after every session."),
    ("Back up anchor images to Google Drive immediately", "REQUIRED",
     "Path: LMT-Video-Assets/CHARACTERS/. Do not keep anchor files only on desktop — a crash loses the character."),
]

ws4.merge_cells("A4:C4")
ws4.cell(row=4, column=1).value = "THE RULES"
ws4.cell(row=4, column=1).fill = NAVY_FILL
ws4.cell(row=4, column=1).font = NAVY_FONT
ws4.cell(row=4, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws4.row_dimensions[4].height = 28

col_headers = ["Rule", "Priority", "Why / How to Apply"]
for col, h in enumerate(col_headers, 1):
    cell_set(ws4, 5, col, h, fill=GOLD_FILL, font=GOLD_FONT, height=30)

for i, (rule, priority, explanation) in enumerate(rules):
    r = i + 6
    if rule == "" and priority == "" and explanation == "":
        ws4.row_dimensions[r].height = 8
        continue
    if priority == "Rule":
        # section divider
        ws4.merge_cells(f"A{r}:C{r}")
        ws4.cell(row=r, column=1).value = rule
        ws4.cell(row=r, column=1).fill = BLUE_FILL
        ws4.cell(row=r, column=1).font = BLUE_FONT
        ws4.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws4.row_dimensions[r].height = 22
        continue
    fill = RED_FILL if priority == "ABSOLUTE" else ALT_FILL
    font_p = Font(bold=True, size=10, color="9C0006") if priority == "ABSOLUTE" else Font(bold=True, size=10, color="7D5A00")
    cell_set(ws4, r, 1, rule, fill=fill, height=52)
    cell_set(ws4, r, 2, priority, fill=fill, font=font_p, height=52)
    cell_set(ws4, r, 3, explanation, fill=fill, height=52)

ws4.column_dimensions["A"].width = 40
ws4.column_dimensions["B"].width = 12
ws4.column_dimensions["C"].width = 70
ws4.freeze_panes = "A6"


# ═══════════════════════════════════════════════════════════════════
# SHEET 5 — SHOT LOG (fill in as you generate)
# ═══════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Shot Log")
page_title(ws5,
    "Shot Log — Fill In As You Generate",
    "Track every accepted clip. Update after every session. This is how you find the right file 6 months from now.")

log_headers = ["Date", "Scene #", "Character", "Reference File Used", "Tool + Settings", "Take #", "Accepted?", "Output Filename", "Notes"]
section_header(ws5, 4, "SHOT LOG — UPDATE AFTER EVERY GENERATION SESSION", cols=9)
for col, h in enumerate(log_headers, 1):
    cell_set(ws5, 5, col, h, fill=GOLD_FILL, font=GOLD_FONT, height=32)

# 30 empty rows for logging
for r in range(6, 36):
    for col in range(1, 10):
        cell_set(ws5, r, col, "", height=28)
    ws5.row_dimensions[r].height = 28

for col_idx, width in enumerate([12, 8, 20, 30, 26, 8, 12, 34, 30], start=1):
    ws5.column_dimensions[get_column_letter(col_idx)].width = width
ws5.freeze_panes = "A6"


wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
