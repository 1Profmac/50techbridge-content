"""
AI Digital Film Series — Character Consistency + Tool Functions + Free Resources
Excel reference guide for LMT / Brian McKinney
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = "content/research/ai-film-character-consistency-guide-2026.xlsx"

GREEN_FILL  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
GREEN_FONT  = Font(bold=True, color="276221", size=10)
BLUE_FILL   = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
BLUE_FONT   = Font(bold=True, color="FFFFFF", size=11)
GOLD_FILL   = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
GOLD_FONT   = Font(bold=True, color="7D5A00", size=10)
SECTION_FILL= PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SECTION_FONT= Font(bold=True, color="1F4E79", size=11)
ALT_FILL    = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")
BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════════
# SHEET 1 — CHARACTER CONSISTENCY WORKFLOW
# ═══════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Character Consistency"

def header_row(ws, row, text, cols="A:G"):
    ws.merge_cells(f"{cols.split(':')[0]}{row}:{cols.split(':')[1]}{row}")
    c = ws.cell(row=row, column=1)
    c.value = text
    c.fill = SECTION_FILL
    c.font = SECTION_FONT
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 26

def title_block(ws, text, subtitle):
    ws.merge_cells("A1:G1")
    ws["A1"].value = text
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A2:G2")
    ws["A2"].value = subtitle
    ws["A2"].font = Font(italic=True, size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

def write_row(ws, row, data, bold=False, fill=None, font_color="000000", height=40):
    for col, val in enumerate(data, start=1):
        c = ws.cell(row=row, column=col)
        c.value = val
        c.border = BORDER
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.font = Font(bold=bold, size=10, color=font_color)
        if fill:
            c.fill = fill
    ws.row_dimensions[row].height = height

title_block(ws1,
    "AI Film Character Consistency — Step by Step Guide 2026",
    "How to create one character and keep them consistent across every scene with other characters")

# ── PHASE 1 ──────────────────────────────────────────
header_row(ws1, 4, "PHASE 1 — BUILD YOUR CHARACTER REFERENCE (before you touch video)")
write_row(ws1, 5, ["Step", "What to Do", "Tool", "How", "Cost", "Output", "Warning"], bold=True, fill=BLUE_FILL, font_color="FFFFFF", height=32)

phase1 = [
    ("1", "Write a detailed character description", "Claude Pro",
     "Name, age, skin tone, hair color/length, eye color, clothing style, any distinctive features. Be specific — 'brown hair' is not enough. 'Shoulder-length dark chestnut hair with natural wave, side part' is.",
     "$20/mo", "Character brief document", "Vague descriptions = inconsistent character across every tool"),
    ("2", "Generate character reference sheet (front, side, 3/4 view)", "Midjourney",
     "Prompt: [description] character reference sheet, front view, side view, 3/4 view, white background, consistent lighting, same outfit --cw 100 --ar 3:2\nRun 4 variations. Pick the best. This is your ANCHOR image.",
     "$30/mo", "3-angle character sheet PNG", "Never change the anchor image once locked. All scenes reference this one file."),
    ("3", "Lock clothing and distinctive features", "Midjourney",
     "Generate 5-8 stills of same character in different poses/expressions using --cref [anchor image URL] --cw 100. Confirm face holds across all variations before moving to video.",
     "$30/mo", "Character expression pack (5-8 images)", "If face drifts here, it will drift worse in video. Fix in Midjourney first."),
    ("4", "Save character reference pack to Google Drive", "Google Drive",
     "Folder: LMT-Video-Assets/[Film Name]/Characters/[CharacterName]/. Include: anchor sheet, expression pack, character brief TXT.",
     "Free", "Organized reference folder", "Never delete or edit anchor image. It is your source of truth."),
]

for i, row in enumerate(phase1):
    r = i + 6
    write_row(ws1, r, list(row), fill=ALT_FILL if i % 2 else None, height=60)

# ── PHASE 2 ──────────────────────────────────────────
header_row(ws1, 11, "PHASE 2 — GENERATE SCENES WITH YOUR CHARACTER (video generation)")
write_row(ws1, 12, ["Step", "What to Do", "Tool", "How", "Cost", "Output", "Warning"], bold=True, fill=BLUE_FILL, font_color="FFFFFF", height=32)

phase2 = [
    ("5", "Generate scene stills for each shot using character reference", "Midjourney",
     "Prompt: [scene description] --cref [anchor URL] --cw 90-100 --oref [anchor URL] --ow 80\nGenerate the still first. Confirm character looks right BEFORE animating.",
     "$30/mo", "Scene still images", "Change ONE variable per shot. New location = same angle/lighting. New angle = same location/lighting."),
    ("6", "Animate stills to video (5-second clips)", "Kling AI",
     "Upload your Midjourney still as the Start Frame in Kling. Use Character ID feature to lock the face. Set motion: subtle for dialogue, dynamic for action.\nKling 3.0: use AI Multi-Shot for multi-angle in one generation.",
     "$26/mo (Pro)", "5-10 sec video clips per scene", "Do not change the reference image between clips in the same scene. Same image = same character."),
    ("7", "Scenes with TWO characters together", "Kling AI",
     "Upload BOTH character reference images in Kling's multi-character reference feature. Describe each character separately in the prompt:\n'[Character A: description] stands on left. [Character B: description] stands on right.'\nKeep scene simple — complex interaction increases drift risk.",
     "$26/mo (Pro)", "Two-character scene clips", "Two-character scenes have higher drift risk. Generate extras — expect 30-40% reject rate. Budget more credits for these shots."),
    ("8", "Fix face drift between clips", "Runway (Acts as cleanup tool)",
     "If Kling drifts on a clip, upload the drifted clip to Runway with your anchor image as reference. Use video-to-video with low motion strength to pull the face back.",
     "$35/mo (Pro)", "Corrected clips", "Don't spend credits fixing bad clips in Kling. Switch to Runway for repairs — it's cheaper per fix."),
]

for i, row in enumerate(phase2):
    r = i + 13
    write_row(ws1, r, list(row), fill=ALT_FILL if i % 2 else None, height=70)

# ── PHASE 3 ──────────────────────────────────────────
header_row(ws1, 18, "PHASE 3 — MAINTAINING CONSISTENCY ACROSS EPISODES")
write_row(ws1, 19, ["Rule", "Why It Matters", "How to Enforce It"], bold=True, fill=BLUE_FILL, font_color="FFFFFF", height=32)

phase3 = [
    ("Never regenerate your anchor image", "Midjourney is non-deterministic — regenerating produces a different face even with the same prompt", "Lock anchor in ep 1. Store in Drive. Use that file for every episode forever."),
    ("Write a character prompt template", "Copy-paste errors introduce drift when you retype from memory", "Save the exact Midjourney prompt for each character as a text file. Copy-paste every time — never retype."),
    ("Same lighting in every scene for a character", "Lighting changes alter how the face reads across scenes — it looks like a different person", "Define lighting in character brief: 'soft natural side light from left.' Include in every prompt."),
    ("Wardrobe changes need new reference images", "If character changes outfit, generate a new reference sheet for that outfit — don't wing it", "New outfit = new Midjourney reference generation before video. Add to character folder."),
    ("Keep a Shot Log spreadsheet", "Tracks which reference image was used per scene so you can match it in reshoots", "Column: Scene | Reference Image File | Kling settings used | Accepted clip filename"),
    ("Generate 3x more clips than you need", "AI video has a 30-50% reject rate from drift, artifacts, wrong motion", "Plan for waste. If you need 10 clips per scene, generate 25-30."),
]

for i, row in enumerate(phase3):
    r = i + 20
    write_row(ws1, r, list(row), fill=ALT_FILL if i % 2 else None, height=52)

# Column widths for sheet 1
widths1 = [5, 28, 18, 48, 14, 22, 38]
for col_idx, width in enumerate(widths1, start=1):
    ws1.column_dimensions[get_column_letter(col_idx)].width = width
ws1.freeze_panes = "A6"


# ═══════════════════════════════════════════════════════
# SHEET 2 — TOOLS BY SECTION AND FUNCTION
# ═══════════════════════════════════════════════════════
ws2 = wb.create_sheet("Tools By Function")

ws2.merge_cells("A1:H1")
ws2["A1"].value = "AI Film Tools — Sections and Functions 2026"
ws2["A1"].font = Font(bold=True, size=14, color="1F4E79")
ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 32
ws2.merge_cells("A2:H2")
ws2["A2"].value = "Green = Already in LMT stack | Gold = Recommended next purchase | White = Optional / advanced"
ws2["A2"].font = Font(italic=True, size=10, color="666666")
ws2["A2"].alignment = Alignment(horizontal="center")
ws2.row_dimensions[2].height = 18

tools_by_function = [
    # section, tool, function, what it does, monthly, free tier, owned, priority
    ("WRITING & STORY", "Claude Pro", "Screenwriting / scripting", "Full screenplay, scene descriptions, character dialogue, shot lists, voiceover scripts", "$20/mo", "No — $20/mo only", True, "Have it"),
    ("WRITING & STORY", "ChatGPT Plus", "Script drafting (alternate)", "Alternative to Claude for story brainstorming. Not needed if you have Claude.", "$20/mo", "Yes — limited", False, "Skip — use Claude"),

    ("CHARACTER DESIGN", "Midjourney (Standard)", "Character reference sheet generation", "Generate anchor character images. Front/side/3/4 view sheets. Expression packs. --cref for consistency.", "$30/mo", "No", False, "Buy first — critical"),
    ("CHARACTER DESIGN", "Nano Banana (Google AI Studio)", "Free character consistency (multi-ref)", "Upload 3-5 character images, maintains identity across scenes. Free via Google AI Studio.", "$0", "Yes — fully free", False, "Try free first before Midjourney"),

    ("STORYBOARD & PRE-VIS", "Canva Pro", "Shot list slides, animatic, scene order", "Build storyboard deck, plan shot sequence, title cards", "$18/mo", "Yes — limited", True, "Have it"),
    ("STORYBOARD & PRE-VIS", "ScreenWeaver (free tool)", "Cinematic prompt generator", "Free tool that builds structured video prompts for Veo, Kling, Runway from your shot description", "$0", "Yes — fully free", False, "Use it now — free"),

    ("VIDEO GENERATION — PRIMARY", "Kling AI (Pro)", "Scene generation / character animation", "Best motion realism. Character ID locks face. Multi-Shot for multi-angle. Two-character scenes.", "$26/mo", "Yes — 66 credits/day (watermarked)", False, "Buy second — critical"),
    ("VIDEO GENERATION — SECONDARY", "Runway (Pro)", "Clip repair / style transfer / fast iteration", "Use to fix Kling drift. Video-to-video restyling. Faster iteration than Kling.", "$35/mo", "Yes — 125 one-time credits", False, "Buy after Kling"),
    ("VIDEO GENERATION — ENVIRONMENTS", "Luma AI Ray (Plus)", "Cinematic environments / establishing shots", "Best for photorealistic wide shots and locations. Character not in frame.", "$30/mo", "Yes — limited free credits", False, "Optional — add when budget allows"),
    ("VIDEO GENERATION — PIPELINE", "LTX Studio (Standard)", "Full pipeline workspace in one tool", "Script → storyboard → character lock → video → timeline editor. All in one. Best for beginners.", "$35/mo", "Yes — 8,000 compute seconds free", False, "Alternative to Kling+Runway if budget is tight"),

    ("VOICE & AUDIO", "ElevenLabs (Creator)", "Character voices / narration / voice cloning", "All character voices. Clone your own voice. 121K credits/mo. Commercial rights.", "$22/mo", "Yes — 10K credits/mo", True, "Have it"),
    ("VOICE & AUDIO", "Suno AI (Pro)", "Original music score / soundtrack", "Generate full film score from text prompt. 500 songs/mo. Commercial use.", "$10/mo", "Yes — 50 credits/day", False, "Buy third — affordable"),

    ("ASSEMBLY & EDITING", "ffmpeg", "Clip assembly / audio mix / caption burn-in", "Combine all clips into final film. Sync voiceover + music. Burn SRT captions. Free.", "$0", "Yes — fully free", True, "Have it"),
    ("ASSEMBLY & EDITING", "DaVinci Resolve (Free)", "Professional timeline editor / color grade", "Industry standard editor. Color grade. Multi-track audio. Free version covers everything needed.", "$0", "Yes — fully free", False, "Download now — free and powerful"),
    ("ASSEMBLY & EDITING", "Adobe Premiere Pro", "Professional timeline (paid alternative)", "Same as DaVinci but costs $22.99/mo. No advantage over free DaVinci for AI film work.", "$22.99/mo", "No", False, "Skip — DaVinci is free and equivalent"),

    ("UPSCALING", "Topaz Video AI", "4K upscale / artifact removal / frame interpolation", "AI video generates at 720p-1080p. Topaz upscales to 4K and removes AI artifacts.", "$299/yr", "No", False, "Add when producing final release cuts"),

    ("CAPTIONS & SUBTITLES", "Claude + ffmpeg", "SRT caption writing + burn-in", "Write captions in Claude, save as .srt, burn with ffmpeg. Already proven in PSA.", "$0 (included)", "Yes", True, "Have it"),

    ("DISTRIBUTION", "Metricool (Free)", "Schedule FB + IG + YouTube", "20 posts/mo free. All three platforms connected.", "$0", "Yes — 20 posts/mo", True, "Have it"),
    ("DISTRIBUTION", "YouTube Studio", "Native upload for long-form film", "Upload full film directly. Best algorithm treatment for native uploads.", "$0", "Yes", True, "Have it — use it"),
]

headers2 = ["Section", "Tool", "Function", "What It Does", "Monthly Cost", "Free Tier?", "In Stack?", "Priority / Action"]
write_row(ws2, 4, headers2, bold=True, fill=BLUE_FILL, font_color="FFFFFF", height=36)

for i, (section, tool, function, what, monthly, free_tier, owned, priority) in enumerate(tools_by_function):
    r = i + 5
    row_data = [section, tool, function, what, monthly, free_tier, "✅ Yes" if owned else "❌ No", priority]
    for col, val in enumerate(row_data, start=1):
        c = ws2.cell(row=r, column=col)
        c.value = val
        c.border = BORDER
        c.alignment = Alignment(vertical="center", wrap_text=True)
        if owned:
            c.fill = GREEN_FILL
            c.font = GREEN_FONT
        elif priority.startswith("Buy") or priority.startswith("Download") or priority.startswith("Try"):
            c.fill = GOLD_FILL
            c.font = GOLD_FONT
        else:
            c.fill = ALT_FILL if i % 2 else PatternFill()
            c.font = Font(size=10)
    ws2.row_dimensions[r].height = 52

widths2 = [22, 22, 24, 46, 14, 14, 10, 26]
for col_idx, width in enumerate(widths2, start=1):
    ws2.column_dimensions[get_column_letter(col_idx)].width = width
ws2.freeze_panes = "A5"


# ═══════════════════════════════════════════════════════
# SHEET 3 — FREE LEARNING RESOURCES
# ═══════════════════════════════════════════════════════
ws3 = wb.create_sheet("Free Learning Resources")

ws3.merge_cells("A1:F1")
ws3["A1"].value = "Free AI Filmmaking Learning Resources 2026"
ws3["A1"].font = Font(bold=True, size=14, color="1F4E79")
ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 32

ws3.merge_cells("A2:F2")
ws3["A2"].value = "Curious Refuge $149/mo membership still exists (curiousrefuge.com/curious-refuge-membership) — but free alternatives below cover 80% of the same content"
ws3["A2"].font = Font(italic=True, size=10, color="276221")
ws3["A2"].alignment = Alignment(horizontal="center")
ws3.row_dimensions[2].height = 20

resources_headers = ["Type", "Resource", "What You Learn", "URL / Where to Find", "Cost", "Best For"]
write_row(ws3, 4, resources_headers, bold=True, fill=BLUE_FILL, font_color="FFFFFF", height=32)

resources = [
    # Curious Refuge
    ("Curious Refuge (Paid)", "Curious Refuge $149/mo Membership", "Every course: AI filmmaking, animation, VFX, advertising. Live office hours. Career coaching.", "curiousrefuge.com/curious-refuge-membership", "$149/mo (annual billing)", "Full curriculum — best if you go deep"),
    ("Curious Refuge (Free)", "Curious Refuge Free Starter Course", "Intro to AI filmmaking tools, concepts, and pipeline overview", "curiousrefuge.com/start-here", "FREE", "First step before buying anything"),

    # YouTube — full courses
    ("YouTube — Full Course", "The Complete AI Filmmaking Course (Script to Finished Movie)", "Full script-to-screen pipeline. Tools, workflow, editing.", "Search YouTube: 'Complete AI Filmmaking Course Script to Finished Movie 2026'", "FREE", "Best single free course available"),
    ("YouTube — Full Course", "The Only AI Filmmaking Course You'll Ever Need (2026)", "Complete beginner walkthrough of AI film production", "Search YouTube: 'Only AI Filmmaking Course You'll Ever Need 2026'", "FREE", "Beginners — start here"),
    ("YouTube — Mini Course", "AI Filmmaking Mini-Course — Kling & Veo (Playlist)", "How to make cinematic AI movies using Kling and Google Veo 3.1", "Search YouTube: 'AI Filmmaking Mini-Course Kling Veo playlist'", "FREE", "Best for learning Kling workflow"),

    # Character consistency
    ("YouTube — Character", "How to Create Consistent Characters in Midjourney (Step by Step)", "Lock a character across scenes using --cref and Omni Reference", "Search YouTube: 'Create Consistent Characters Midjourney 2026 step by step'", "FREE", "Critical watch before starting any film"),
    ("YouTube — Character", "Kling AI Character Consistency Guide 2026", "Character ID feature, multi-character scenes, reference images in Kling", "Search YouTube: 'Kling AI Character Consistency 2026'", "FREE", "Watch after Midjourney character tutorial"),

    # Free tools with tutorials
    ("Free Tool + Tutorial", "ScreenWeaver — Free AI Prompt Generator", "Build structured cinematic prompts for Kling, Veo, Sora, Runway", "screenweaver.ai", "FREE", "Use before every video generation session"),
    ("Free Tool + Tutorial", "Nano Banana / Google AI Studio — Character Consistency", "Upload 3-5 character reference images, maintains identity — free alternative to paid tools", "aistudio.google.com", "FREE", "Test this before buying Midjourney"),
    ("Free Tool + Tutorial", "DaVinci Resolve — Full Free Course (YouTube)", "Professional video editing, color grading, audio mixing — no subscription", "Search YouTube: 'DaVinci Resolve complete beginner course 2026'", "FREE (software + tutorial)", "Learn editing without paying Adobe"),

    # Blogs / written guides
    ("Written Guide", "Sami Haraketi — Indie Filmmaker's Guide to Free AI Tools", "Complete written pipeline: script to screen using only free tools", "samiharaketi.com — search 'script to screen free AI tools 2026'", "FREE", "Read before spending any money"),
    ("Written Guide", "601 Media — AI Film Workflow: Text to Video No Camera No Crew", "Full workflow breakdown for no-budget AI films", "601media.com — search 'AI film workflow script to final cut no camera no crew'", "FREE", "Best written reference for solo creators"),
    ("Written Guide", "Midjourney --cref Character Reference Complete Guide", "Technical deep-dive on Midjourney character reference parameters", "prompting.systems — search 'midjourney cref consistent characters'", "FREE", "Bookmark — reference every time you build a character"),

    # Communities
    ("Community", "r/aivideo (Reddit)", "Real filmmaker workflows, tool comparisons, problem-solving, weekly showcases", "reddit.com/r/aivideo", "FREE", "Best community for troubleshooting"),
    ("Community", "Curious Refuge Discord (Free)", "Active AI filmmaker community. Tool tips. Workflow sharing.", "Access via curiousrefuge.com", "FREE", "Join now — free even without paid membership"),
    ("Community", "Kling AI Official Discord", "Direct access to Kling team. Feature updates. Workflow tips from power users.", "Access via kling.ai", "FREE", "Join when you subscribe to Kling"),
]

for i, row in enumerate(resources):
    r = i + 5
    for col, val in enumerate(row, start=1):
        c = ws3.cell(row=r, column=col)
        c.value = val
        c.border = BORDER
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.font = Font(size=10)
        c.fill = ALT_FILL if i % 2 else PatternFill()
    ws3.row_dimensions[r].height = 48

widths3 = [22, 36, 44, 44, 18, 30]
for col_idx, width in enumerate(widths3, start=1):
    ws3.column_dimensions[get_column_letter(col_idx)].width = width
ws3.freeze_panes = "A5"

wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
