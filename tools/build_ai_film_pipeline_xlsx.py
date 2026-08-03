"""
Build Complete AI Digital Movie Pipeline — Excel with green highlights for owned tools.
Script-to-screen breakdown for LMT / Brian McKinney reference.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = "content/research/ai-digital-movie-pipeline-2026.xlsx"

# ── Pipeline stages ──────────────────────────────────────────────────────────
stages = [
    # (stage, order, tool, purpose, monthly_mo, monthly_ann, annual, notes, owned)
    ("1 — SCRIPT & STORY", 1, "Claude Pro", "Write screenplay, scene descriptions, dialogue, character arcs",
     "$20", "$20", "$240", "Already in stack. Full script in one session. Export as PDF or Final Draft txt.", True),

    ("1 — SCRIPT & STORY", 2, "ChatGPT Pro (optional)", "Alternate script drafting / story structure",
     "$200", "$200", "$2,400", "Only needed if you want Sora 2 video access. Skip if using Claude.", False),

    ("2 — STORYBOARD & CONCEPT ART", 3, "Midjourney", "Generate character reference sheets, scene stills, visual style frames",
     "$30 (Standard)", "$24", "$288", "Standard plan (15 GPU hrs/mo) covers a short film. Lock character looks here before video.", False),

    ("2 — STORYBOARD & CONCEPT ART", 4, "Canva Pro", "Build animatic / shot list slides, title cards, scene order",
     "$18", "$12", "$144", "Already in stack. Use for storyboard deck and shot order planning.", True),

    ("3 — VIDEO GENERATION", 5, "Kling AI", "Primary video generation — motion realism, character consistency, cinematic shots",
     "$25.99 (Pro)", "$20.79", "$249.48", "Best motion realism in 2026. Pro plan (660 credits/mo) = ~33 x 5-sec clips. Entry tool for digital movies.", False),

    ("3 — VIDEO GENERATION", 6, "Runway (Gen-4.5)", "Secondary video — fast iteration, style transfer, image-to-video",
     "$35 (Pro)", "$28", "$336", "Best for quick re-takes and style consistency. Use alongside Kling.", False),

    ("3 — VIDEO GENERATION", 7, "Luma AI Ray 3", "Cinematic wide shots, photorealistic environments",
     "$30 (Plus)", "$25", "$300", "Strong for establishing shots and environments. Plus plan entry point.", False),

    ("3 — VIDEO GENERATION", 8, "LTX Studio", "Full pipeline workspace — script, storyboard, character lock, video, timeline editor in one tool",
     "$35 (Standard)", "$28", "$336", "Most complete single-tool pipeline. Includes Veo 2, character consistency, camera controls. Best for beginners.", False),

    ("4 — VOICE & AUDIO", 9, "ElevenLabs (Creator)", "All character voices, narrator, voice cloning",
     "$22", "$18.26", "$219", "Already in stack. Voice ID: uAs0vN0GLLpz7FM7JVkz. Creator plan covers a full short film.", True),

    ("4 — VOICE & AUDIO", 10, "Suno AI (Pro)", "Original music score and soundtrack",
     "$10", "$8", "$96", "Generate full film score from text prompt. Pro plan = 500 songs/mo. Commercial rights included.", False),

    ("5 — POST PRODUCTION", 11, "ffmpeg", "Assemble clips, sync audio, burn captions, mix audio tracks",
     "$0", "$0", "$0", "Already in stack. Free open source. Command-line. Used in LMT PSA.", True),

    ("5 — POST PRODUCTION", 12, "Adobe Premiere Pro", "Professional timeline edit, color grade, multi-track audio",
     "$34.49", "$22.99", "$263.88", "Industry standard. Not required if using LTX Studio or ffmpeg for short films.", False),

    ("5 — POST PRODUCTION", 13, "Adobe After Effects", "VFX, title sequences, compositing, motion graphics",
     "$34.49", "$22.99", "$263.88", "For advanced VFX only. Overkill for AI-generated short films.", False),

    ("5 — POST PRODUCTION", 14, "Topaz Video AI", "Upscale AI-generated footage from 720p to 4K, remove artifacts",
     "$39/mo or", "$24.92", "$299/yr", "Critical for theatrical-quality output. AI video generators output 720p-1080p — Topaz upscales to 4K.", False),

    ("5 — POST PRODUCTION", 15, "DaVinci Resolve (Free)", "Color grading, audio mixing, editing — free alternative to Premiere",
     "$0 (free tier)", "$0", "$0", "Free version covers 99% of indie film post needs. Studio version $295 one-time.", False),

    ("6 — CAPTIONS & DISTRIBUTION", 16, "Claude CLI + ffmpeg", "Generate SRT captions, burn into video, format for platforms",
     "$0 (included in Claude Pro)", "$0", "$0", "Already in stack. Same workflow used in PSA.", True),

    ("6 — CAPTIONS & DISTRIBUTION", 17, "Metricool", "Schedule and distribute to FB, IG, YouTube",
     "$0 (free plan)", "$0", "$0", "Already in stack. Free plan covers 20 posts/mo.", True),
]

headers = ["Stage", "#", "Tool", "Purpose", "Monthly\n(Month-to-Month)", "Monthly\n(Annual Plan)", "Annual Price", "Notes"]

# Colors
GREEN_FILL   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
GREEN_FONT   = Font(bold=True, color="276221", size=10)
HEADER_FILL  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
STAGE_FILL   = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
STAGE_FONT   = Font(bold=True, color="1F4E79", size=10)
ALT_FILL     = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")
BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "AI Film Pipeline 2026"

# Title
ws.merge_cells("A1:H1")
ws["A1"].value = "Complete AI Digital Movie Pipeline — Script to Screen 2026"
ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 32

# Subtitle
ws.merge_cells("A2:H2")
ws["A2"].value = "Green = Already in LMT stack  |  Prices as of August 2026  |  Built for solo creator / no crew"
ws["A2"].font = Font(italic=True, size=10, color="666666")
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 18

ws.append([])

# Headers (row 4)
ws.append(headers)
for col_idx in range(1, len(headers) + 1):
    cell = ws.cell(row=4, column=col_idx)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER
ws.row_dimensions[4].height = 36

# Data rows
for i, (stage, order, tool, purpose, monthly_mo, monthly_ann, annual, notes, owned) in enumerate(stages):
    row_num = i + 5
    ws.append([stage, order, tool, purpose, monthly_mo, monthly_ann, annual, notes])

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)

        if owned:
            cell.fill = GREEN_FILL
            cell.font = GREEN_FONT
        elif i % 2 == 1:
            cell.fill = ALT_FILL
            cell.font = Font(size=10)
        else:
            cell.font = Font(size=10)

    ws.row_dimensions[row_num].height = 52

# Column widths
col_widths = [22, 4, 26, 38, 18, 18, 14, 54]
for col_idx, width in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.freeze_panes = "A5"

# ── Cost Summary ──────────────────────────────────────────────────────────────
summary_row = len(stages) + 7

def srow(r, col, val, bold=False, color="000000", size=10):
    c = ws.cell(row=r, column=col)
    c.value = val
    c.font = Font(bold=bold, size=size, color=color)

srow(summary_row, 1, "COST SUMMARY", bold=True, size=12, color="1F4E79")

rows = [
    ("LMT Current Stack (Claude + ElevenLabs + Canva + ffmpeg + Metricool)", "~$42/month", "$504/year"),
    ("Minimum viable digital movie stack (add Kling + Midjourney + Suno)", "~$78/month", "$936/year"),
    ("Full professional pipeline (add Runway + Topaz + LTX Studio)", "~$180/month", "$2,160/year"),
    ("Curious Refuge / full studio tier (add Adobe CC)", "~$375–$500/month", "$4,500–$6,000/year"),
    ("Gap from current LMT stack to minimum viable movie stack", "~$36/month", "$432/year"),
]

srow(summary_row + 1, 1, "Stack", bold=True, color="1F4E79")
srow(summary_row + 1, 2, "Monthly Cost", bold=True, color="1F4E79")
srow(summary_row + 1, 3, "Annual Cost", bold=True, color="1F4E79")

for offset, (label, monthly, annual) in enumerate(rows):
    r = summary_row + 2 + offset
    srow(r, 1, label)
    srow(r, 2, monthly, bold=True)
    srow(r, 3, annual, bold=True)

# Script section
script_row = summary_row + len(rows) + 4
srow(script_row, 1, "SCRIPT STRUCTURE — AI SHORT FILM (5–15 min)", bold=True, size=12, color="1F4E79")

script_steps = [
    ("1", "LOGLINE", "One sentence. Character + conflict + stakes. Write in Claude first."),
    ("2", "TREATMENT", "1–2 page prose summary. Scene by scene. No dialogue yet."),
    ("3", "SCENE BREAKDOWN", "List every scene: INT/EXT, location, characters, action, mood, camera note."),
    ("4", "VISUAL STYLE GUIDE", "Define look in Midjourney. Generate 10–15 reference frames before touching video tools."),
    ("5", "CHARACTER SHEETS", "Generate consistent character reference in Midjourney. Lock face, clothes, lighting style."),
    ("6", "SHOT LIST", "Every scene: wide/medium/close. Camera move. Duration. Which AI tool generates it."),
    ("7", "VOICEOVER / DIALOGUE SCRIPT", "Final dialogue with ElevenLabs notes: pace, emotion, pause markers."),
    ("8", "MUSIC CUE SHEET", "Scene-by-scene music notes for Suno: mood, tempo, instrumentation, duration."),
    ("9", "GENERATE VIDEO", "Scene by scene. Kling or Runway from Midjourney reference frames."),
    ("10", "ASSEMBLE IN TIMELINE", "LTX Studio or Premiere or ffmpeg. Sync VO + music + video."),
    ("11", "CAPTIONS", "Claude writes SRT. ffmpeg burns in. Review for timing."),
    ("12", "UPSCALE", "Topaz Video AI: 720p → 4K. Final export at H.264 or H.265."),
    ("13", "DISTRIBUTE", "YouTube native upload. Metricool for FB + IG. LinkedIn manual."),
]

srow(script_row + 1, 1, "Step", bold=True, color="1F4E79")
srow(script_row + 1, 2, "Phase", bold=True, color="1F4E79")
srow(script_row + 1, 3, "What to Do", bold=True, color="1F4E79")

for offset, (step, phase, what) in enumerate(script_steps):
    r = script_row + 2 + offset
    srow(r, 1, step)
    srow(r, 2, phase, bold=True)
    srow(r, 3, what)
    ws.row_dimensions[r].height = 20

wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
