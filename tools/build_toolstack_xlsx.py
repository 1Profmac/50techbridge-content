"""
Build Curious Refuge AI Filmmaking Toolstack — Excel with green highlights for owned tools.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = "content/research/curious-refuge-toolstack-pricing-2026.xlsx"

tools = [
    {
        "order": 1, "tool": "Midjourney", "category": "Image Generation",
        "purpose": "Character & scene stills, concept art, visual development",
        "monthly_billing": "$10 / $30 / $60 / $120",
        "annual_billing": "$8 / $24 / $48 / $96",
        "annual_price": "$96 – $1,152",
        "notes": "No free tier. Standard ($30/mo) is working filmmaker plan. 15 GPU hrs/mo.",
        "owned": False
    },
    {
        "order": 2, "tool": "Runway", "category": "Video Generation",
        "purpose": "Image-to-video, cinematic motion, fast editing",
        "monthly_billing": "$15 / $35 / $95",
        "annual_billing": "$12 / $28 / $76",
        "annual_price": "$144 – $912",
        "notes": "1 sec Gen-4.5 = 25 credits. Pro (2,250 credits/mo) minimum for production.",
        "owned": False
    },
    {
        "order": 3, "tool": "Luma AI (Dream Machine / Ray)", "category": "Video Generation",
        "purpose": "Cinematic motion from stills, Ray 3 video",
        "monthly_billing": "$30 / $90 / $300",
        "annual_billing": "$25 / $75 / $250",
        "annual_price": "$300 – $3,000",
        "notes": "Credits do not roll over. Ray-2 API: $0.08/sec. Plus plan entry point.",
        "owned": False
    },
    {
        "order": 4, "tool": "Sora 2 (OpenAI)", "category": "Video Generation",
        "purpose": "Premium cinematic sequences, narrative flow",
        "monthly_billing": "$200 (ChatGPT Pro bundle)",
        "annual_billing": "$200 (ChatGPT Pro bundle)",
        "annual_price": "$2,400",
        "notes": "Web app discontinued April 2026. API only until Sept 24, 2026. $0.10/sec (720p).",
        "owned": False
    },
    {
        "order": 5, "tool": "ElevenLabs", "category": "Voice / Audio",
        "purpose": "Voiceover generation, voice cloning",
        "monthly_billing": "$0 / $5 / $22 / $99",
        "annual_billing": "$0 / $4.15 / $18.26 / $82",
        "annual_price": "$0 – $986",
        "notes": "Creator ($22/mo) covers most production. 121K credits/mo. Commercial rights included.",
        "owned": True
    },
    {
        "order": 6, "tool": "Adobe After Effects", "category": "Post Production",
        "purpose": "Camera movements, upscaling, compositing",
        "monthly_billing": "$34.49",
        "annual_billing": "$22.99",
        "annual_price": "$263.88",
        "notes": "Single app plan. Annual commitment required for $22.99/mo rate.",
        "owned": False
    },
    {
        "order": 7, "tool": "Adobe Premiere Pro", "category": "Post Production",
        "purpose": "Final edit, assembly, color",
        "monthly_billing": "$34.49",
        "annual_billing": "$22.99",
        "annual_price": "$263.88",
        "notes": "Single app. Often bundled with After Effects via Creative Cloud.",
        "owned": False
    },
    {
        "order": 8, "tool": "Adobe Creative Cloud (All Apps)", "category": "Post Production",
        "purpose": "Full Adobe suite — Premiere + After Effects + Photoshop + 20+ apps",
        "monthly_billing": "$69.99",
        "annual_billing": "$54.99",
        "annual_price": "$659.88",
        "notes": "Best value if using multiple Adobe tools. Includes 1TB storage.",
        "owned": False
    },
    {
        "order": 9, "tool": "Canva Pro", "category": "Design / Slides",
        "purpose": "Slide design, graphics, video assembly",
        "monthly_billing": "$18.00",
        "annual_billing": "$12.00",
        "annual_price": "$144.00",
        "notes": "Free tier available but limited. Pro needed for brand kit and video features.",
        "owned": True
    },
    {
        "order": 10, "tool": "Metricool", "category": "Scheduling",
        "purpose": "Schedule FB + IG + YouTube. Analytics.",
        "monthly_billing": "$0 (Free plan)",
        "annual_billing": "$0 (Free plan)",
        "annual_price": "$0",
        "notes": "Free plan: 20 posts/month. FB + IG + YT connected. LinkedIn requires $20/mo paid.",
        "owned": True
    },
    {
        "order": 11, "tool": "ffmpeg", "category": "Assembly",
        "purpose": "Video assembly, audio mixing, caption burn-in",
        "monthly_billing": "$0",
        "annual_billing": "$0",
        "annual_price": "$0",
        "notes": "Free and open source. Command-line only. Used in LMT PSA production.",
        "owned": True
    },
    {
        "order": 12, "tool": "Claude CLI", "category": "Scripting / Writing",
        "purpose": "Write scripts, captions, ffmpeg commands, voiceover copy",
        "monthly_billing": "$20 (Pro)",
        "annual_billing": "$20 (Pro)",
        "annual_price": "$240",
        "notes": "Claude Pro includes extended context and Claude Sonnet 4.6 access.",
        "owned": True
    },
]

headers = ["#", "Tool", "Category", "Purpose", "Monthly\n(Month-to-Month)", "Monthly\n(Annual Plan)", "Annual Price", "Notes"]

# Colors
GREEN_FILL  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
GREEN_FONT  = Font(bold=True, color="276221")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL    = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "AI Filmmaking Toolstack"

# Title row
ws.merge_cells("A1:H1")
title_cell = ws["A1"]
title_cell.value = "Curious Refuge AI Filmmaking Toolstack — Pricing 2026"
title_cell.font = Font(bold=True, size=14, color="1F4E79")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

# Subtitle row
ws.merge_cells("A2:H2")
sub_cell = ws["A2"]
sub_cell.value = "Green = Tools already in LMT stack  |  Prices as of August 2026"
sub_cell.font = Font(italic=True, size=10, color="666666")
sub_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 18

ws.append([])  # spacer

# Header row (row 4)
ws.append(headers)
for col_idx, _ in enumerate(headers, start=1):
    cell = ws.cell(row=4, column=col_idx)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER
ws.row_dimensions[4].height = 36

# Data rows
for i, tool in enumerate(tools):
    row_num = i + 5
    row_data = [
        tool["order"],
        tool["tool"],
        tool["category"],
        tool["purpose"],
        tool["monthly_billing"],
        tool["annual_billing"],
        tool["annual_price"],
        tool["notes"],
    ]
    ws.append(row_data)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)

        if tool["owned"]:
            cell.fill = GREEN_FILL
            cell.font = GREEN_FONT
        elif i % 2 == 1:
            cell.fill = ALT_FILL

    ws.row_dimensions[row_num].height = 48

# Column widths
col_widths = [4, 28, 18, 38, 22, 20, 14, 52]
for col_idx, width in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# Freeze header
ws.freeze_panes = "A5"

# Summary block below data
summary_row = len(tools) + 7
ws.cell(row=summary_row, column=1).value = "COST COMPARISON"
ws.cell(row=summary_row, column=1).font = Font(bold=True, size=11, color="1F4E79")

rows = [
    ("Full Curious Refuge Stack (entry level)", "~$375 – $500/month"),
    ("LMT Current Stack (ElevenLabs + Canva + ffmpeg + Claude)", "~$42/month"),
    ("Gap to full Curious Refuge stack", "~$333 – $458/month"),
]
for offset, (label, value) in enumerate(rows):
    ws.cell(row=summary_row + 1 + offset, column=1).value = label
    ws.cell(row=summary_row + 1 + offset, column=1).font = Font(size=10)
    ws.cell(row=summary_row + 2 + offset - 1, column=2).value = value
    ws.cell(row=summary_row + 2 + offset - 1, column=2).font = Font(bold=True, size=10)

wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
